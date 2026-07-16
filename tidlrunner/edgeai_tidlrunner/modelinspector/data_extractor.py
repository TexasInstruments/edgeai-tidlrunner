#!/usr/bin/env python3
# Copyright (c) 2018-2025, Texas Instruments
# All Rights Reserved.

"""
Data Extractor - Extract TIDL artifact data to JSON

This script discovers and parses all TIDL compilation artifacts from work_dirs/
and outputs a single compressed JSON file containing all extracted data.

Usage:
    python data_extractor.py <work_dirs/> <output.json>

Example:
    python data_extractor.py work_dirs/ model_data.json

Output:
    - Single JSON file with all parsed data (metadata, model, compilation)
"""

import onnx
import json
import sys
import os
import re
import gzip
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional

# Python 3.10+ compatibility: collections.Callable moved to collections.abc.Callable
import collections
import collections.abc
if not hasattr(collections, 'Callable'):
    collections.Callable = collections.abc.Callable

from bs4 import BeautifulSoup

try:
    import onnx_graphsurgeon as gs
    HAS_GRAPHSURGEON = True
except ImportError:
    HAS_GRAPHSURGEON = False
    print("WARNING: onnx-graphsurgeon not installed. Using raw ONNX API (slower).")
    print("  Install with: pip install onnx-graphsurgeon")


class ActivationDataParser:
    """Parser for layer activation binary files with histogram and scatter plot generation

    Uses layer_info.txt files directly to build activation mappings
    """

    def __init__(self, model_dir: str, frame_idx: int = 0, tidl_data: Optional[Dict] = None):
        """
        Initialize activation data parser using layer_info.txt files

        Args:
            model_dir: Direct path to model directory (e.g., work_dirs/compile/AM69A/cl_onnx_model/)
            frame_idx: Frame index to use for activation data (default: 0)
            tidl_data: Optional TIDL parsed data containing layer_index to onnx_node_index mappings
        """
        self.model_dir = Path(model_dir)
        self.frame_idx = frame_idx
        self.mapping = {}
        self.data_cache = {}
        self.layer_to_onnx_node = {}  # Maps (subgraph_id, layer_index) -> onnx_node_index

        self._load_from_layer_info()

        # Build layer_index to onnx_node_index mapping (lightweight, just IDs)
        if tidl_data:
            self._build_layer_to_onnx_mapping(tidl_data)

    def _load_from_layer_info(self):
        """Build activation mapping from layer_info.txt files"""
        import glob

        print(f"Building activation mapping from layer_info.txt files...")

        layer_info_pattern = os.path.join(self.model_dir, 'tidl/artifacts/tempDir/subgraph_*_tidl_net.bin.layer_info.txt')
        layer_info_files = glob.glob(layer_info_pattern, recursive=False)

        if not layer_info_files:
            print(f"  WARNING: No layer_info.txt files found in {layer_info_pattern}")
            return

        print(f"  Found {len(layer_info_files)} layer_info.txt files")

        notidl_outputs_dir = self._find_notidl_outputs()
        tidl_traces_dir = self._find_tidl_traces()

        if not notidl_outputs_dir or not tidl_traces_dir:
            print(f"  WARNING: Could not find output directories")
            return

        total_layers = 0
        total_mapped = 0

        for layer_info_path in sorted(layer_info_files):
            match = re.search(r'subgraph_(\d+)_tidl_net\.bin\.layer_info\.txt', layer_info_path)
            if not match:
                continue

            subgraph_id = int(match.group(1))

            if subgraph_id not in self.mapping:
                self.mapping[subgraph_id] = {}

            with open(layer_info_path, 'r') as f:
                for line in f:
                    parts = line.strip().split()
                    if len(parts) < 3:
                        continue

                    tidl_id_col0 = parts[0]
                    tidl_id_col1 = parts[1]
                    onnx_layer_name = parts[2]

                    if tidl_id_col0 != tidl_id_col1:
                        continue

                    tidl_layer_id = tidl_id_col0
                    total_layers += 1

                    notidl_path = self._build_notidl_path(notidl_outputs_dir, onnx_layer_name)
                    tidl_path = self._build_tidl_path(tidl_traces_dir, subgraph_id, tidl_layer_id)

                    if notidl_path and tidl_path and os.path.exists(notidl_path) and os.path.exists(tidl_path):
                        self.mapping[subgraph_id][tidl_layer_id] = {
                            'notidl': notidl_path,
                            'tidl': tidl_path,
                            'onnx_name': onnx_layer_name
                        }
                        total_mapped += 1

        print(f"  Total layers in layer_info.txt: {total_layers}")
        print(f"  Layers with activation files: {total_mapped}")

        for sg_id in sorted(self.mapping.keys()):
            print(f"    Subgraph {sg_id}: {len(self.mapping[sg_id])} layers")

    def _build_layer_to_onnx_mapping(self, tidl_data: Dict):
        """Build lightweight mapping from (subgraph_id, layer_index) to onnx_node_index"""
        print("  Building layer_index to onnx_node_index mapping...")
        for subgraph_id, subgraph_info in tidl_data.items():
            if 'layers' not in subgraph_info:
                continue

            for layer in subgraph_info['layers']:
                layer_index = layer.get('layer_index')
                onnx_node_index = layer.get('onnx_node_index')

                if layer_index is not None and onnx_node_index is not None:
                    key = (subgraph_id, str(layer_index))
                    self.layer_to_onnx_node[key] = onnx_node_index

        print(f"    Mapped {len(self.layer_to_onnx_node)} layer indices to ONNX node indices")

    def _find_notidl_outputs(self) -> Optional[str]:
        """Find NotIDL outputs directory"""
        import glob
        for subdir in ['notidl', 'notidl32']:
            pattern = os.path.join(self.model_dir, f'{subdir}/outputs_/{self.frame_idx}')
            matches = glob.glob(pattern, recursive=False)
            if matches:
                return matches[0]
        return None

    def _find_tidl_traces(self) -> Optional[str]:
        """Find TIDL traces directory — checks both tidl/ and tidl32/"""
        import glob
        for subdir in ['tidl', 'tidl32']:
            pattern = os.path.join(self.model_dir, f'{subdir}/traces_/{self.frame_idx}')
            matches = glob.glob(pattern, recursive=False)
            if matches:
                print(f"  Found TIDL traces in {subdir}/traces_/{self.frame_idx}")
                return matches[0]
        return None

    def _build_notidl_path(self, notidl_dir: str, onnx_layer_name: str) -> Optional[str]:
        """Build path to NotIDL output file"""
        onnx_layer_id = onnx_layer_name.replace("/", "_")
        notidl_path = os.path.join(notidl_dir, f"{onnx_layer_id}.bin")

        if os.path.exists(notidl_path):
            return notidl_path

        for filename in os.listdir(notidl_dir):
            if onnx_layer_id in filename and filename.endswith('.bin'):
                return os.path.join(notidl_dir, filename)

        return None

    def _build_tidl_path(self, tidl_dir: str, subgraph_id: int, tidl_layer_id: str) -> Optional[str]:
        """Build path to TIDL trace file"""
        import glob

        tidl_id_padded = tidl_layer_id.zfill(4)
        pattern = os.path.join(tidl_dir, f"tidl_trace_subgraph_{subgraph_id}_{tidl_id_padded}_*_float.bin")
        matches = glob.glob(pattern)

        return matches[0] if matches else None

    def _load_bin_file(self, bin_path: str) -> Optional[np.ndarray]:
        """Load binary file with caching"""
        if bin_path in self.data_cache:
            return self.data_cache[bin_path]

        try:
            if not os.path.exists(bin_path):
                print(f"    Warning: File not found: {bin_path}")
                return None

            data = np.fromfile(bin_path, dtype=float)
            self.data_cache[bin_path] = data
            return data

        except Exception as e:
            print(f"    Error loading {bin_path}: {e}")
            return None

    def _smart_sample_scatter_points(self, tidl_data: np.ndarray, notidl_data: np.ndarray,
                                      max_points: int = 2000) -> Tuple[np.ndarray, np.ndarray]:
        """Smart sampling using best-fit line as reference to preserve outliers

        Args:
            tidl_data: TIDL (quantized) activation values
            notidl_data: Original FP32 activation values
            max_points: Maximum number of points to keep (default 4000)

        Returns:
            Tuple of (sampled_tidl_data, sampled_notidl_data)
        """
        total_points = len(tidl_data)

        # Adaptive minimum: ensure we have enough points for visualization
        # For small datasets, keep more; for large datasets, aim for max_points
        if total_points <= 2000:
            return tidl_data, notidl_data
        elif total_points <= 8000:
            min_points = int(total_points * 0.4)  # Keep at least 40% for small-medium datasets
        else:
            min_points = max(2000, int(max_points * 0.5))  # At least 2000 or 50% of max_points

        # If already under max limit, return all points
        if total_points <= max_points:
            return tidl_data, notidl_data

        try:
            # Calculate best-fit line through all points
            # Try sklearn first (more robust), fallback to numpy
            try:
                from sklearn.linear_model import LinearRegression
                X = tidl_data.reshape(-1, 1)
                y = notidl_data
                model = LinearRegression()
                model.fit(X, y)
                slope = float(model.coef_[0])
                intercept = float(model.intercept_)
            except ImportError:
                # Fallback to numpy polyfit
                coeffs = np.polyfit(tidl_data, notidl_data, 1)
                slope = float(coeffs[0])
                intercept = float(coeffs[1])

            # Calculate perpendicular distance from best-fit line
            # Distance = |y - (mx + b)| / sqrt(1 + m^2)
            predicted = slope * tidl_data + intercept
            distances = np.abs(notidl_data - predicted) / np.sqrt(1 + slope**2)

            # Adaptive thresholds based on percentiles
            p50 = np.percentile(distances, 50)
            p75 = np.percentile(distances, 75)
            p90 = np.percentile(distances, 90)
            p95 = np.percentile(distances, 95)

            # Zone-based sampling with different rates
            # Format: (lower_bound, upper_bound, sampling_rate, zone_name)
            zones = [
                (p95, np.inf, 1.0, 'critical'),    # Keep ALL outliers
                (p90, p95, 0.8, 'poor'),            # Keep 80% of poor points
                (p75, p90, 0.4, 'fair'),            # Keep 40% of fair points
                (p50, p75, 0.15, 'good'),           # Keep 15% of good points
                (0.0, p50, 0.15, 'excellent')       # Keep 5% of excellent points
            ]

            # Build index list for each zone
            selected_indices = []
            np.random.seed(42)  # Reproducible sampling

            for lower, upper, rate, zone_name in zones:
                # Find points in this zone
                mask = (distances >= lower) & (distances < upper)
                zone_indices = np.where(mask)[0]

                if len(zone_indices) == 0:
                    continue

                # Sample from this zone
                n_to_sample = int(np.ceil(len(zone_indices) * rate))
                if n_to_sample > len(zone_indices):
                    n_to_sample = len(zone_indices)

                sampled_indices = np.random.choice(zone_indices, n_to_sample, replace=False)
                selected_indices.extend(sampled_indices)

            # Convert to array
            selected_indices = np.array(selected_indices)

            # If we don't have enough points, fill up by sampling more from better zones
            if len(selected_indices) < min_points:
                # Get all unselected indices
                all_indices = np.arange(total_points)
                selected_set = set(selected_indices)
                unselected_indices = np.array([i for i in all_indices if i not in selected_set])

                if len(unselected_indices) > 0:
                    # How many more points do we need?
                    needed = min_points - len(selected_indices)
                    needed = min(needed, len(unselected_indices))

                    # Sample from unselected points, preferring those closer to outliers
                    unselected_distances = distances[unselected_indices]
                    # Sort by distance (descending) and take the top 'needed' points
                    sorted_unselected = unselected_indices[np.argsort(-unselected_distances)]
                    additional_indices = sorted_unselected[:needed]

                    selected_indices = np.concatenate([selected_indices, additional_indices])

            # Limit to max_points if we have too many
            if len(selected_indices) > max_points:
                # Prioritize by distance (keep worst outliers)
                sorted_by_distance = selected_indices[np.argsort(-distances[selected_indices])]
                selected_indices = sorted_by_distance[:max_points]

            # Validation: ensure we kept the worst outlier
            max_distance_idx = np.argmax(distances)
            if max_distance_idx not in selected_indices:
                # Replace a random good point with the max outlier
                good_zone_mask = distances[selected_indices] < p75
                if np.any(good_zone_mask):
                    good_indices_in_selection = np.where(good_zone_mask)[0]
                    replace_idx = np.random.choice(good_indices_in_selection)
                    selected_indices[replace_idx] = max_distance_idx

            # Return sampled data
            sampled_tidl = tidl_data[selected_indices]
            sampled_notidl = notidl_data[selected_indices]

            # Print sampling summary
            reduction_pct = (1 - len(selected_indices) / total_points) * 100
            print(f"    Smart sampling: {total_points} -> {len(selected_indices)} points ({reduction_pct:.1f}% reduction, min: {min_points})")
            print(f"    Best-fit line: y = {slope:.4f}*x + {intercept:.4f}")

            return sampled_tidl, sampled_notidl

        except Exception as e:
            print(f"    Warning: Smart sampling failed ({e}), using random sampling")
            # Fallback to simple random sampling
            np.random.seed(42)
            indices = np.random.choice(total_points, max_points, replace=False)
            return tidl_data[indices], notidl_data[indices]

    def _sample_data(self, data: np.ndarray, max_samples: int = 50000) -> np.ndarray:
        """Sample data for visualization"""
        if len(data) <= max_samples:
            return data

        np.random.seed(42)
        indices = np.random.choice(len(data), max_samples, replace=False)
        return data[indices]

    def _sanitize_float(self, value: float) -> float:
        """Convert Infinity/NaN to None for JSON compatibility"""
        if np.isnan(value) or np.isinf(value):
            return None
        return value

    def _calculate_statistics(self, notidl_data: np.ndarray, tidl_data: np.ndarray) -> Dict[str, float]:
        """Calculate statistics needed for plot generation"""
        try:
            notidl_min = float(np.min(notidl_data))
            notidl_max = float(np.max(notidl_data))
            notidl_mean = float(np.mean(notidl_data))
            notidl_std = float(np.std(notidl_data))

            tidl_min = float(np.min(tidl_data))
            tidl_max = float(np.max(tidl_data))
            tidl_mean = float(np.mean(tidl_data))
            tidl_std = float(np.std(tidl_data))

            return {
                'total_points': len(notidl_data),
                'notidl_min': self._sanitize_float(notidl_min),
                'notidl_max': self._sanitize_float(notidl_max),
                'notidl_mean': self._sanitize_float(notidl_mean),
                'notidl_std': self._sanitize_float(notidl_std),
                'tidl_min': self._sanitize_float(tidl_min),
                'tidl_max': self._sanitize_float(tidl_max),
                'tidl_mean': self._sanitize_float(tidl_mean),
                'tidl_std': self._sanitize_float(tidl_std)
            }

        except Exception as e:
            print(f"    Error calculating statistics: {e}")
            return {}

    def _generate_histogram_json(self, notidl_data: np.ndarray, tidl_data: np.ndarray,
                              stats: Dict[str, float]) -> Dict[str, Any]:
        """Generate histogram data for visualization"""

        try:
            notidl_data = np.asarray(notidl_data).flatten()
            tidl_data = np.asarray(tidl_data).flatten()

            # Filter out NaN and inf values to prevent histogram errors
            notidl_data = notidl_data[np.isfinite(notidl_data)]
            tidl_data = tidl_data[np.isfinite(tidl_data)]

            # Check if we have valid data after filtering
            if len(notidl_data) == 0 or len(tidl_data) == 0:
                print(f"    Warning: No finite values available for histogram generation")
                return {}

            notidl_counts, notidl_edges = np.histogram(notidl_data, bins=100)
            tidl_counts, tidl_edges = np.histogram(tidl_data, bins=100)

            notidl_centers = (notidl_edges[:-1] + notidl_edges[1:]) / 2
            tidl_centers = (tidl_edges[:-1] + tidl_edges[1:]) / 2

            notidl_centers = np.round(notidl_centers, 4)
            tidl_centers = np.round(tidl_centers, 4)

            notidl_centers_list = notidl_centers.tolist()
            notidl_counts_list = notidl_counts.tolist()
            tidl_centers_list = tidl_centers.tolist()
            tidl_counts_list = tidl_counts.tolist()

        except Exception as e:
            print(f"    Error generating histogram bins: {e}")
            import traceback
            traceback.print_exc()
            return {'traces': [], 'layout': {}}

        return {
            'tidl_bins': tidl_centers_list,
            'tidl_counts': tidl_counts_list,
            'notidl_bins': notidl_centers_list,
            'notidl_counts': notidl_counts_list,
        }

    def _generate_scatter_plot_d3(self, notidl_data: np.ndarray, tidl_data: np.ndarray,
                                   stats: Dict[str, float]) -> Dict[str, Any]:
        """Generate scatter plot data for D3 visualization"""

        try:
            notidl_flat = np.asarray(notidl_data, dtype=np.float64).flatten()
            tidl_flat = np.asarray(tidl_data, dtype=np.float64).flatten()

            total_points = len(notidl_flat)

            # Filter out corrupted values: NaN, Inf, or extreme outliers
            # These appear in layers like GlobalAveragePool when TIDL output is misread
            valid_mask = (
                np.isfinite(notidl_flat) & np.isfinite(tidl_flat) &
                (np.abs(notidl_flat) < 1e10) & (np.abs(tidl_flat) < 1e10)
            )
            if valid_mask.sum() == 0:
                # All values corrupted — return empty scatter with warning
                return {'points': [], 'stats': {'total_points': total_points, 'displayed_points': 0,
                                                'warning': 'Corrupted activation data'}}
            notidl_flat = notidl_flat[valid_mask]
            tidl_flat = tidl_flat[valid_mask]

            notidl_rounded = np.round(notidl_flat, 4)
            tidl_rounded = np.round(tidl_flat, 4)

            # Apply smart sampling if too many points
            if total_points > 2000:
                tidl_sampled, notidl_sampled = self._smart_sample_scatter_points(
                    tidl_rounded, notidl_rounded, max_points=2000
                )
                tidl_rounded = tidl_sampled
                notidl_rounded = notidl_sampled

            all_values = np.concatenate([notidl_rounded, tidl_rounded])
            min_val = float(np.min(all_values))
            max_val = float(np.max(all_values))

            padding = (max_val - min_val) * 0.05
            axis_min = min_val - padding
            axis_max = max_val + padding

            scatter_data = {
                'x': [float(notidl_rounded[i]) for i in range(len(notidl_rounded))],
                'y': [float(tidl_rounded[i]) for i in range(len(tidl_rounded))],
                'sample_size': len(tidl_rounded),
                'total_points': total_points,
                'axis': {
                    'min': axis_min,
                    'max': axis_max
                },
            }

            return scatter_data

        except Exception as e:
            print(f"    Error generating D3 scatter plot: {e}")
            import traceback
            traceback.print_exc()
            return {'x': [], 'y': [], 'sample_size': 0, 'total_points': 0}

    def process_layer(self, subgraph_id: int, tidl_layer_id: str) -> Optional[Dict[str, Any]]:
        """Process a single layer and generate plot data

        Args:
            subgraph_id: Subgraph ID
            tidl_layer_id: TIDL layer ID within subgraph

        Returns:
            Dict with stats, histogram, and scatter plot data, or None if processing failed
        """

        if subgraph_id not in self.mapping:
            return None

        if tidl_layer_id not in self.mapping[subgraph_id]:
            return None

        layer_data = self.mapping[subgraph_id][tidl_layer_id]

        notidl_path = layer_data['notidl']
        tidl_path = layer_data['tidl']

        notidl_data = self._load_bin_file(notidl_path)
        tidl_data = self._load_bin_file(tidl_path)

        if notidl_data is None or tidl_data is None:
            return None

        if len(notidl_data) != len(tidl_data):
            print(f"  Warning: Size mismatch for subgraph {subgraph_id} layer {tidl_layer_id} (NoTIDL: {len(notidl_data)}, TIDL: {len(tidl_data)})")
            return None

        try:
            stats = self._calculate_statistics(notidl_data, tidl_data)
            histogram_data = self._generate_histogram_json(notidl_data, tidl_data, stats)
            scatter_data = self._generate_scatter_plot_d3(notidl_data, tidl_data, stats)

        except Exception as e:
            print(f"  Error processing layer {tidl_layer_id}: {e}")
            import traceback
            traceback.print_exc()
            return None

        del notidl_data, tidl_data

        return {
            'histogram': histogram_data,
            'scatter': scatter_data,
            'metrics': stats,
            'bin_files': {
                'tidl': tidl_path,
                'notidl': notidl_path
            }
        }

    def process_all_layers(self, subgraph_node_mapping: Dict[int, List[int]] = None) -> Dict[str, Any]:
        """
        Process all layers and generate activation data

        Args:
            subgraph_node_mapping: NOT USED - kept for backward compatibility

        Returns:
            Dict with keys "subgraphId_tidlLayerId" -> plot data
        """
        if not self.mapping:
            print("  No activation mapping available, skipping activation analysis")
            return {}

        print("\nProcessing activation data...")

        total_layers = sum(len(layers) for layers in self.mapping.values())
        print(f"  Found {total_layers} TIDL layers with activation files")

        activation_data = {}
        total_processed = 0
        total_failed = 0
        subgraph_stats = {}

        for subgraph_id in sorted(self.mapping.keys()):
            sorted_tidl_ids = sorted(self.mapping[subgraph_id].keys(), key=lambda x: int(x))
            layer_count = len(sorted_tidl_ids)
            print(f"\n  Subgraph {subgraph_id}: Processing {layer_count} layers")

            for tidl_layer_id in sorted_tidl_ids:
                result = self.process_layer(subgraph_id, tidl_layer_id)

                if result:
                    # Key matches JSON layer_id (HTML layer_index = TIDL net layer number)
                    key = f"{subgraph_id}_{tidl_layer_id}"
                    activation_data[key] = result
                    total_processed += 1
                    subgraph_stats[subgraph_id] = subgraph_stats.get(subgraph_id, 0) + 1

                    if total_processed <= 10 or total_processed % 10 == 0:
                        layer_info = self.mapping[subgraph_id][tidl_layer_id]
                        onnx_name = layer_info.get('onnx_name', 'unknown')
                        print(f"    Layer {tidl_layer_id} ({onnx_name}): Processed")
                else:
                    total_failed += 1

        print(f"\nActivation processing complete:")
        for sg_id in sorted(subgraph_stats.keys()):
            print(f"  Subgraph {sg_id}: {subgraph_stats[sg_id]} layers processed")
        print(f"  Total Processed: {total_processed}")
        print(f"  Total Failed: {total_failed}")
        print(f"  Data size: {len(json.dumps(activation_data)) / (1024*1024):.2f} MB")

        return activation_data


class MetricsParser:
    """Parser for analyze.xlsx metrics file"""

    def __init__(self, xlsx_path: str):
        """
        Initialize metrics parser

        Args:
            xlsx_path: Path to analyze.xlsx file
        """
        self.xlsx_path = Path(xlsx_path)
        self.metrics_data = {}

        if self.xlsx_path.exists():
            self._load_metrics()
            print(f"Loaded metrics from: {xlsx_path}")
        else:
            print(f"WARNING: Metrics file not found: {xlsx_path}")

    def _load_metrics(self):
        """Load metrics from Excel file"""
        try:
            import openpyxl
        except ImportError:
            print("  Error: openpyxl not installed. Run: pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(self.xlsx_path, read_only=True)

            for sheet_name in wb.sheetnames:
                if sheet_name.startswith('diff_notidl_tidl_'):
                    sheet = wb[sheet_name]

                    for row_idx in range(2, sheet.max_row + 1):
                        row = [cell.value for cell in sheet[row_idx]]

                        if len(row) >= 12 and row[1] is not None:
                            # Group by the Subgraph column (row[0]) — not by sheet name
                            subgraph_id = str(row[0]) if row[0] is not None else '0'
                            # Detect new format (with SNR_dB as first metric column) vs old format
                            has_snr = len(row) >= 13 and isinstance(row[4], (int, float)) and str(row[4]) not in ('', 'None')
                            if has_snr:
                                # New format: SNR_dB, MeanAbsRelDiff, MeanAbsDiff, MedianAbsDiff, MaxAbsDiff, ...
                                metric_entry = {
                                    'subgraph': subgraph_id,
                                    'serial_num': row[1],
                                    'onnx_layer': row[2],
                                    'tidl_layer_id': row[3],
                                    'snr_db': float(row[4]) if row[4] is not None else 0.0,
                                    'mean_abs_rel_diff': float(row[5]) if row[5] is not None else 0.0,
                                    'mean_abs_diff': float(row[6]) if row[6] is not None else 0.0,
                                    'median_abs_diff': float(row[7]) if row[7] is not None else 0.0,
                                    'max_abs_diff': float(row[8]) if row[8] is not None else 0.0,
                                    'mean_abs_diff_median': float(row[10]) if len(row) > 10 and row[10] is not None else 0.0,
                                    'median_abs_diff_median': float(row[11]) if len(row) > 11 and row[11] is not None else 0.0,
                                    'max_abs_diff_median': float(row[12]) if len(row) > 12 and row[12] is not None else 0.0
                                }
                            else:
                                # Old format: MeanAbsRelDiff, MeanAbsDiff, MedianAbsDiff, MaxAbsDiff, ...
                                metric_entry = {
                                    'subgraph': subgraph_id,
                                    'serial_num': row[1],
                                    'onnx_layer': row[2],
                                    'tidl_layer_id': row[3],
                                    'mean_abs_rel_diff': float(row[4]) if row[4] is not None else 0.0,
                                    'mean_abs_diff': float(row[5]) if row[5] is not None else 0.0,
                                    'median_abs_diff': float(row[6]) if row[6] is not None else 0.0,
                                    'max_abs_diff': float(row[7]) if row[7] is not None else 0.0,
                                    'mean_abs_diff_median': float(row[9]) if len(row) > 9 and row[9] is not None else 0.0,
                                    'median_abs_diff_median': float(row[10]) if len(row) > 10 and row[10] is not None else 0.0,
                                    'max_abs_diff_median': float(row[11]) if len(row) > 11 and row[11] is not None else 0.0
                                }
                            if subgraph_id not in self.metrics_data:
                                self.metrics_data[subgraph_id] = []
                            self.metrics_data[subgraph_id].append(metric_entry)

            for sg_id, metrics in self.metrics_data.items():
                print(f"  Loaded {len(metrics)} metrics for subgraph {sg_id}")

            wb.close()

        except Exception as e:
            print(f"  Error loading metrics: {e}")
            self.metrics_data = {}

    def get_metrics(self) -> Dict[str, List[Dict[str, Any]]]:
        """Get all metrics data organized by subgraph"""
        return self.metrics_data


class TIDLNetLogParser:
    """Parser for subgraph_X_tidl_net.bin_netLog.txt files"""

    def __init__(self, filepath: str):
        self.filepath = filepath

    def parse(self) -> Dict[str, Any]:
        """Parse netLog.txt file and extract MACS information"""
        print(f"  Parsing netLog: {os.path.basename(self.filepath)}")

        layer_macs = {}
        total_gmacs = 0.0

        try:
            with open(self.filepath, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            in_table = False

            for line in lines:
                line_stripped = line.strip()

                if '----' in line_stripped and len(line_stripped) > 50:
                    in_table = True
                    continue

                if in_table and '----' in line_stripped and len(line_stripped) > 50:
                    in_table = False
                    continue

                if in_table and line_stripped and not line_stripped.startswith('Num|'):
                    parts = [p.strip() for p in line.split('|')]

                    if len(parts) >= 11:
                        try:
                            layer_num = int(parts[0])
                            layer_name = parts[1].strip()
                            output_name = parts[2].strip()
                            inbuf_str = parts[6].strip()
                            outbuf_str = parts[7].strip()
                            macs_str = parts[-2]  # trailing | creates empty last element

                            # Parse inbuf IDs (non-'x' values are real buffer IDs)
                            inbuf_ids = [int(x) for x in inbuf_str.split() if x != 'x']
                            outbuf_id = int(outbuf_str.split()[0]) if outbuf_str else layer_num

                            macs = int(macs_str) if macs_str.isdigit() else 0
                            gmacs = macs / 1_000_000_000.0

                            layer_macs[layer_num] = {
                                'layer_name': layer_name,
                                'output_name': output_name,
                                'inbuf_ids': inbuf_ids,
                                'outbuf_id': outbuf_id,
                                'macs': macs,
                                'gmacs': gmacs
                            }

                        except (ValueError, IndexError) as e:
                            continue

                if 'Total Giga Macs' in line:
                    match = re.search(r'Total Giga Macs\s*:\s*([\d.]+)', line)
                    if match:
                        total_gmacs = float(match.group(1))

            print(f"    Found MACS info for {len(layer_macs)} layers, Total: {total_gmacs} GMACS")

            return {
                'layer_macs': layer_macs,
                'total_gmacs': total_gmacs
            }

        except Exception as e:
            print(f"    Warning: Failed to parse netLog file: {e}")
            return {
                'layer_macs': {},
                'total_gmacs': 0.0
            }



class TIDLSubgraphParser:
    """Parser for subgraph_X_tidl_net.bin.html files"""

    def __init__(self, base_dir: str, node_support: Dict[int, Dict[str, Any]] = None):
        self.base_dir = base_dir
        self.node_support = node_support or {}

    def _parse_layer_info_file(self, subgraph_id: int) -> Dict[int, str]:
        """Parse layer_info.txt to get TIDL layer index → ONNX tensor name mapping"""
        layer_info_path = os.path.join(self.base_dir, f'subgraph_{subgraph_id}_tidl_net.bin.layer_info.txt')
        mapping = {}  # tidl_layer_index → onnx_tensor_name
        if not os.path.exists(layer_info_path):
            return mapping
        with open(layer_info_path, 'r') as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) < 3:
                    continue
                tidl_idx = int(parts[0])
                buffer_id = int(parts[1])
                tensor_name = parts[2]
                # Skip buffer reuse entries and model input
                if tidl_idx != buffer_id or tensor_name == 'images':
                    continue
                # Strip _netFormat suffix to get the base ONNX tensor name
                if tensor_name.endswith('_netFormat'):
                    tensor_name = tensor_name[:-10]
                mapping[tidl_idx] = tensor_name
        return mapping

    def _map_tidl_to_onnx_node(self, tidl_output_name: str) -> Optional[int]:
        """
        Map TIDL layer output tensor name to ONNX node index

        Args:
            tidl_output_name: Output tensor name from TIDL layer

        Returns:
            ONNX node index, or None if not found
        """
        if not self.node_support:
            return None

        onnx_node_name = tidl_output_name
        if onnx_node_name.endswith('_output_0'):
            onnx_node_name = onnx_node_name[:-9]
        if onnx_node_name.endswith('_netFormat'):
            onnx_node_name = onnx_node_name[:-10]

        # First try: match by node name
        for node_idx, node_data in self.node_support.items():
            if node_data.get('node_name') == onnx_node_name:
                return node_idx

        # Second try: match by output tensor name in the ONNX layer details
        if hasattr(self, 'tensor_to_node_map') and self.tensor_to_node_map:
            if onnx_node_name in self.tensor_to_node_map:
                return self.tensor_to_node_map[onnx_node_name]

        return None

    def find_subgraph_files(self) -> List[Tuple[int, str, str]]:
        """Find all subgraph HTML files and their corresponding netLog files"""
        pattern = re.compile(r'subgraph_(\d+)_tidl_net\.bin\.html')
        files = []

        for filename in os.listdir(self.base_dir):
            match = pattern.match(filename)
            if match:
                subgraph_idx = int(match.group(1))
                html_filepath = os.path.join(self.base_dir, filename)

                netlog_filename = f'subgraph_{subgraph_idx}_tidl_net.bin_netLog.txt'
                netlog_filepath = os.path.join(self.base_dir, netlog_filename)

                if not os.path.exists(netlog_filepath):
                    netlog_filepath = None

                files.append((subgraph_idx, html_filepath, netlog_filepath))

        files.sort(key=lambda x: x[0])
        return files

    def _parse_output_params(self, text: str, output_dict: Dict[str, Any]):
        """Helper to parse output parameters from text"""
        if not text or not text.strip():
            return

        for param_match in re.finditer(r'(\w+(?:/\w+)?)=([^\s]+)', text):
            param_key = param_match.group(1)
            param_val = param_match.group(2)

            if param_val.isdigit():
                output_dict[param_key] = int(param_val)
            elif param_val.replace('.', '', 1).replace('-', '', 1).isdigit():
                output_dict[param_key] = float(param_val)
            else:
                output_dict[param_key] = param_val

        for array_match in re.finditer(r'(\w+(?:/\w+)?)=\[([^\]]+)\]', text):
            array_key = array_match.group(1)
            array_val = array_match.group(2)

            try:
                output_dict[array_key] = [int(x.strip()) for x in array_val.split(',')]
            except ValueError:
                output_dict[array_key] = array_val

        for paren_match in re.finditer(r'(\w+(?:/\w+)?)\(([^)]+)\)', text):
            paren_key = paren_match.group(1)
            paren_val = paren_match.group(2)
            output_dict[paren_key] = f"({paren_val})"

    def _build_reverse_onnx_mapping(self, tidl_layers: List[Dict]) -> Dict[int, List[int]]:
        """
        Build reverse mapping: onnx_node_index -> [list of tidl_layer_indices]

        Returns:
            Dict mapping ONNX node indices to lists of TIDL layer indices
            Example: {1: [2], 48: [33, 34, 35]}  # 48 is expanded into 3 TIDL layers
        """
        from collections import defaultdict
        reverse_map = defaultdict(list)
        for tidl_idx, layer in enumerate(tidl_layers):
            onnx_idx = layer.get('onnx_node_index')
            if onnx_idx is not None:
                reverse_map[onnx_idx].append(tidl_idx)
        return dict(reverse_map)

    def _detect_layer_expansion(self, reverse_map: Dict[int, List[int]],
                               tidl_layers: List[Dict]) -> Dict[int, Dict]:
        """
        Detect when one ONNX layer expands into multiple TIDL layers.

        Strategy: If one onnx_node_index maps to multiple TIDL layers, it's expansion.
        Only the LAST computational (non-DataConvert) TIDL layer should show activation.

        Returns:
            Dict mapping ONNX node index to expansion info:
            {
                48: {
                    'expanded_tidl_indices': [33, 34],
                    'primary_tidl_index': 33,  # Computational layer (InnerProduct)
                    'last_tidl_index': 33      # Shows activation (primary computational)
                }
            }
        """
        expansion_map = {}

        for onnx_idx, tidl_indices in reverse_map.items():
            if len(tidl_indices) > 1:
                # Expansion detected
                # Find the primary computational layer (not DataLayer/DataConvert)
                primary_idx = None
                last_computational_idx = None

                for tidl_idx in tidl_indices:
                    layer_type = tidl_layers[tidl_idx].get('layer_type', '')

                    # Skip non-computational layers
                    if layer_type in ['TIDL_DataLayer', 'TIDL_DataConvertLayer']:
                        continue

                    if primary_idx is None:
                        primary_idx = tidl_idx

                    last_computational_idx = tidl_idx  # Keep updating to get the last

                expansion_map[onnx_idx] = {
                    'expanded_tidl_indices': tidl_indices,
                    'primary_tidl_index': primary_idx or tidl_indices[0],
                    # Show activation on the last computational layer, not DataConvert
                    'last_tidl_index': last_computational_idx or max(tidl_indices)
                }

        return expansion_map

    def _build_onnx_graph(self):
        """Build ONNX graph structure: node inputs and outputs from layer_details"""
        onnx_graph = {}

        if not hasattr(self, 'onnx_layer_details'):
            return onnx_graph

        for layer_name, layer_info in self.onnx_layer_details.items():
            node_idx = layer_info.get('node_index')
            if node_idx is None:
                continue
            onnx_graph[node_idx] = {
                'name': layer_name,
                'inputs': layer_info.get('input', []),
                'outputs': layer_info.get('output', [])
            }

        return onnx_graph

    def _detect_layer_fusion(self, tidl_layers: List[Dict], subgraph_id: int) -> Dict[int, Dict]:
        """
        Detect when multiple ONNX layers are fused into one TIDL layer.

        Strategy:
        1. Build map of which TIDL layer produces which tensor
        2. Build ONNX graph structure (node inputs/outputs)
        3. For each TIDL layer, trace backwards from output tensor through ONNX graph
        4. Stop when we reach a tensor that's produced by another TIDL layer
        5. All ONNX nodes in between are fused into this TIDL layer

        Returns:
            Dict mapping TIDL layer index to fusion info with ALL fused ONNX nodes
        """
        fusion_map = {}

        # Parse layer_info.txt to get TIDL layer → output tensor mapping
        layer_info_mapping = self._parse_layer_info_file(subgraph_id)

        # Build reverse map: tensor → TIDL layer that produces it
        tidl_tensor_map = {}
        for tidl_idx, tensor_name in layer_info_mapping.items():
            if tensor_name:
                tidl_tensor_map[tensor_name] = tidl_idx

        # Build ONNX graph structure
        onnx_graph = self._build_onnx_graph()

        # Check if we have necessary data
        if not hasattr(self, 'tensor_to_node_map') or not hasattr(self, 'onnx_layer_names'):
            return fusion_map

        # For each TIDL layer, find all ONNX nodes that are fused into it
        for tidl_idx, layer in enumerate(tidl_layers):
            # Get the output tensor for this TIDL layer
            output_tensor = layer_info_mapping.get(tidl_idx)
            if not output_tensor:
                continue

            # Skip non-computational layers
            if layer.get('layer_type') in ['TIDL_DataLayer', 'TIDL_DataConvertLayer']:
                continue

            # Find which ONNX node produces this output tensor
            output_onnx_idx = self.tensor_to_node_map.get(output_tensor)
            if output_onnx_idx is None:
                continue

            # Determine the starting runtime (TIDL or TVM/ARM)
            # supported=True → TIDL, supported=False → TVM/ARM
            start_runtime = self.node_support.get(output_onnx_idx, {}).get('supported', True)

            # Trace backwards through ONNX graph to find all fused nodes
            fused_onnx_nodes = set()
            visited = set()

            def trace_backwards(onnx_node_idx):
                """Recursively trace backwards, stopping when runtime changes"""
                if onnx_node_idx in visited:
                    return
                visited.add(onnx_node_idx)

                fused_onnx_nodes.add(onnx_node_idx)

                if onnx_node_idx not in onnx_graph:
                    return

                input_tensors = onnx_graph[onnx_node_idx].get('inputs', [])

                for input_tensor in input_tensors:
                    # Stop: tensor is output of another TIDL layer in this subgraph
                    if input_tensor in tidl_tensor_map:
                        continue

                    # Stop: model input or constant (no producer node)
                    producer_node_idx = self.tensor_to_node_map.get(input_tensor)
                    if producer_node_idx is None:
                        continue

                    # Stop: runtime changes (TIDL→TVM, TVM→TIDL, TVM→ARM, etc.)
                    producer_runtime = self.node_support.get(producer_node_idx, {}).get('supported', True)
                    if producer_runtime != start_runtime:
                        continue

                    trace_backwards(producer_node_idx)

            # Start tracing from the output node
            trace_backwards(output_onnx_idx)

            # Convert set to sorted list
            fused_indices = sorted(list(fused_onnx_nodes))

            # Get ONNX node names
            fused_names = []
            for idx in fused_indices:
                if idx < len(self.onnx_layer_names):
                    fused_names.append(self.onnx_layer_names[idx])
                elif idx in self.node_support:
                    fused_names.append(self.node_support[idx].get('node_name', f'Node_{idx}'))
                else:
                    fused_names.append(f'Node_{idx}')

            # Only add to fusion_map if more than 1 ONNX node
            if len(fused_indices) > 1:
                fusion_map[tidl_idx] = {
                    'fused_onnx_indices': fused_indices,
                    'fused_onnx_names': fused_names,
                    'fusion_type': self._classify_fusion_type(fused_names)
                }

        return fusion_map

    def _classify_fusion_type(self, onnx_names: List[str]) -> str:
        """Classify the type of fusion based on ONNX layer names"""
        types = []
        for name in onnx_names:
            # Extract operation type from node name (e.g., "Conv_0" -> "CONV")
            if '_' in name:
                op_type = name.split('_')[0].upper()
            else:
                op_type = name.upper()
            types.append(op_type)
        return '_'.join(types)

    def parse_layer_info(self, title_text: str) -> Dict[str, Any]:
        """Parse TIDL layer information from SVG node title"""
        layer_info = {
            'raw_text': title_text,
            'layer_index': None,
            'layer_type': None,
            'layer_name': None,
            'parameters': {}
        }

        lines = title_text.strip().split('\n')
        if not lines:
            return layer_info

        first_line = lines[0].strip()
        layer_match = re.match(r'Layer\s+(\d+):\s+(\w+)\s+"([^"]+)"', first_line)
        if layer_match:
            layer_info['layer_index'] = int(layer_match.group(1))
            layer_info['layer_type'] = layer_match.group(2)
            layer_info['layer_name'] = layer_match.group(3)

        current_section = None
        current_output = None

        for line in lines[1:]:
            line = line.strip()
            if not line:
                continue

            if line.startswith('actParams:'):
                current_section = 'actParams'
                layer_info['parameters']['actParams'] = {}
                remainder = line[len('actParams:'):].strip()
                if remainder:
                    for param_match in re.finditer(r'(\w+(?:/\w+)?)[:=]([^\s]+)', remainder):
                        param_key = param_match.group(1)
                        param_val = param_match.group(2)
                        layer_info['parameters']['actParams'][param_key] = param_val
                continue
            elif line.startswith('Outputs:'):
                current_section = 'outputs'
                if 'outputs' not in layer_info['parameters']:
                    layer_info['parameters']['outputs'] = []

                remainder = line[len('Outputs:'):].strip()
                if remainder:
                    output_match = re.match(r'\[(\d+)\]', remainder)
                    if output_match:
                        current_output = {'index': int(output_match.group(1))}
                        layer_info['parameters']['outputs'].append(current_output)
                        remainder = remainder[output_match.end():].strip()
                    elif not layer_info['parameters']['outputs']:
                        current_output = {'index': 0}
                        layer_info['parameters']['outputs'].append(current_output)
                    else:
                        current_output = layer_info['parameters']['outputs'][-1]

                    if remainder:
                        self._parse_output_params(remainder, current_output)
                continue

            if current_section == 'actParams':
                for param_match in re.finditer(r'(\w+(?:/\w+)?)[:=]([^\s]+)', line):
                    param_key = param_match.group(1)
                    param_val = param_match.group(2)
                    layer_info['parameters']['actParams'][param_key] = param_val

            elif current_section == 'outputs':
                if current_output is None:
                    if not layer_info['parameters']['outputs']:
                        current_output = {'index': 0}
                        layer_info['parameters']['outputs'].append(current_output)
                    else:
                        current_output = layer_info['parameters']['outputs'][-1]

                output_match = re.match(r'\[(\d+)\]', line)
                if output_match:
                    current_output = {'index': int(output_match.group(1))}
                    layer_info['parameters']['outputs'].append(current_output)
                    remainder = line[output_match.end():].strip()
                    if remainder:
                        self._parse_output_params(remainder, current_output)
                else:
                    self._parse_output_params(line, current_output)

            else:
                self._parse_output_params(line, layer_info['parameters'])

        params = layer_info['parameters']
        if 'kernelH/W' in params:
            layer_info['kernelShape'] = params['kernelH/W']

        if 'strideH/W' in params:
            layer_info['strides'] = params['strideH/W']

        if 'padH/W' in params:
            layer_info['pads'] = params['padH/W']

        if 'dilationH/W' in params:
            layer_info['dilations'] = params['dilationH/W']

        if 'numGroups' in params:
            layer_info['groups'] = params['numGroups']

        if 'numInChannels' in params:
            layer_info['numInChannels'] = params['numInChannels']
        if 'numOutChannels' in params:
            layer_info['numOutChannels'] = params['numOutChannels']

        return layer_info

    def _extract_graph_structure(self, soup, layers, subgraph_id=0) -> Tuple[List[Dict], List[Dict]]:
        """Extract graph structure (nodes + edges) from SVG for visualization"""
        graph_nodes = []
        graph_edges = []

        # Parse layer_info.txt for authoritative TIDL → ONNX tensor name mapping
        layer_info_mapping = self._parse_layer_info_file(subgraph_id)

        layer_map = {layer['layer_index']: layer for layer in layers}

        svg_nodes = soup.find_all('g', class_='node')

        for node_elem in svg_nodes:
            title = node_elem.find('title')
            if not title or not title.string:
                continue

            first_line = title.string.strip().split('\n')[0]
            layer_match = re.match(r'Layer\s+(\d+):', first_line)
            if not layer_match:
                continue

            layer_idx = int(layer_match.group(1))
            layer_info = layer_map.get(layer_idx)

            if not layer_info:
                continue

            output_dims = []
            if 'parameters' in layer_info and 'outputs' in layer_info['parameters']:
                for output in layer_info['parameters']['outputs']:
                    if 'dims' in output:
                        output_dims.append(output['dims'])

            output_shape_str = 'N/A'
            if output_dims:
                formatted_shapes = [f"[{','.join(map(str, dims))}]" for dims in output_dims]
                output_shape_str = ', '.join(formatted_shapes)

            input_shape_str = 'N/A'
            if 'numInChannels' in layer_info:
                input_shape_str = f"Channels: {layer_info['numInChannels']}"

            # Get ONNX mapping from layer_info.txt (authoritative source)
            # Only map computational layers — skip DataLayer and DataConvertLayer
            layer_type = layer_info['layer_type'] or 'Unknown'
            non_computational_types = {'TIDL_DataLayer', 'TIDL_DataConvertLayer'}
            onnx_tensor_name = layer_info_mapping.get(layer_idx, '') if layer_type not in non_computational_types else ''
            onnx_node_index = None
            onnx_name = ''
            if onnx_tensor_name and hasattr(self, 'tensor_to_node_map'):
                # Get the node that produces this output tensor
                output_node_idx = self.tensor_to_node_map.get(onnx_tensor_name)
                if output_node_idx is not None:
                    # For fused layers, trace back to find the first computational node
                    # by checking if the previous node's output is the current node's input
                    onnx_node_index = output_node_idx

                    # Trace backwards to find the first node in a potential fusion chain
                    # For Conv+Relu fusion, we want Conv (the earlier node)
                    if output_node_idx > 0 and hasattr(self, 'onnx_layer_names'):
                        # Check if this could be a fused activation (Relu, Clip, etc.)
                        current_name = self.onnx_layer_names[output_node_idx] if output_node_idx < len(self.onnx_layer_names) else ''
                        if 'Relu' in current_name or 'Clip' in current_name:
                            # Likely a fused activation, map to the previous computational node
                            onnx_node_index = output_node_idx - 1

                    if hasattr(self, 'onnx_layer_names'):
                        onnx_name = self.onnx_layer_names[onnx_node_index] if onnx_node_index < len(self.onnx_layer_names) else ''

            node = {
                'id': f"tidl_layer_{layer_idx}",
                'index': layer_idx,
                'name': layer_info['layer_name'][:20] if layer_info['layer_name'] else f"Layer{layer_idx}",
                'full_name': layer_info['layer_name'] or f"Layer_{layer_idx}",
                'type': layer_type,
                'tidl_supported': True,
                'inputshape': input_shape_str,
                'outputshape': output_shape_str,
                'layer_data': layer_info,
                'onnx_node_index': onnx_node_index,
                'onnx_name': onnx_name
            }

            graph_nodes.append(node)

        # Build lookup maps for shape propagation
        node_map = {node['index']: node for node in graph_nodes}
        output_dims_map = {}
        for node in graph_nodes:
            ld = node.get('layer_data', {})
            params = ld.get('parameters', {})
            if 'outputs' in params:
                for out in params['outputs']:
                    if 'dims' in out:
                        output_dims_map[node['index']] = out['dims']
                        break  # use first output with dims

        svg_edges = soup.find_all('g', class_='edge')

        for edge_elem in svg_edges:
            title = edge_elem.find('title')
            if not title or not title.string:
                continue

            edge_text = title.string

            arrow_idx = edge_text.find('->')
            if arrow_idx == -1:
                continue

            source_part = edge_text[:arrow_idx]
            source_match = re.search(r'Layer\s+(\d+):', source_part)
            if not source_match:
                continue
            source_idx = int(source_match.group(1))

            target_part = edge_text[arrow_idx+2:]
            target_match = re.search(r'Layer\s+(\d+):', target_part)
            if not target_match:
                continue
            target_idx = int(target_match.group(1))

            edge = {
                'source': f"tidl_layer_{source_idx}",
                'target': f"tidl_layer_{target_idx}",
                'source_node_id': source_idx,
                'target_node_id': target_idx
            }

            graph_edges.append(edge)

        # Propagate input shapes from predecessor output shapes via edges
        incoming = {}
        for edge in graph_edges:
            target_idx = edge['target_node_id']
            source_idx = edge['source_node_id']
            incoming.setdefault(target_idx, []).append(source_idx)

        for target_idx, source_indices in incoming.items():
            target_node = node_map.get(target_idx)
            if not target_node:
                continue

            input_shapes = []
            for src_idx in source_indices:
                dims = output_dims_map.get(src_idx)
                if dims:
                    input_shapes.append(f"[{','.join(map(str, dims))}]")

            if input_shapes:
                if len(input_shapes) == 1:
                    target_node['inputshape'] = input_shapes[0]
                else:
                    target_node['inputshape'] = ', '.join(input_shapes)

        return graph_nodes, graph_edges

    def parse_subgraph_html(self, filepath: str, netlog_filepath: str = None, subgraph_id: int = 0) -> Dict[str, Any]:
        """Parse a single subgraph HTML file and its netLog file"""
        print(f"  Parsing: {os.path.basename(filepath)}")

        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()

        soup = BeautifulSoup(content, 'html.parser')

        layers = []

        svg_nodes = soup.find_all('g', class_='node')

        for node in svg_nodes:
            title = node.find('title')
            if title and title.string:
                layer_info = self.parse_layer_info(title.string)
                if layer_info['layer_index'] is not None:
                    layers.append(layer_info)

        layers.sort(key=lambda x: x['layer_index'])

        print(f"    Found {len(layers)} TIDL layers")

        macs_data = {'layer_macs': {}, 'total_gmacs': 0.0}
        if netlog_filepath and os.path.exists(netlog_filepath):
            netlog_parser = TIDLNetLogParser(netlog_filepath)
            macs_data = netlog_parser.parse()

            for layer in layers:
                layer_idx = layer['layer_index']
                layer_type = layer.get('layer_type', '')

                if layer_idx in macs_data['layer_macs']:
                    layer['macs'] = macs_data['layer_macs'][layer_idx]['macs']
                    layer['gmacs'] = macs_data['layer_macs'][layer_idx]['gmacs']

                    # Only map to ONNX nodes for computational layers
                    # DataLayer/DataConvertLayer don't correspond to ONNX nodes
                    if layer_type not in ['TIDL_DataLayer', 'TIDL_DataConvertLayer']:
                        tidl_output_name = macs_data['layer_macs'][layer_idx].get('output_name', '')
                        onnx_node_idx = self._map_tidl_to_onnx_node(tidl_output_name)
                        if onnx_node_idx is not None:
                            layer['onnx_node_index'] = onnx_node_idx

        # Note: ONNX mapping is now done in _extract_graph_structure using layer_info.txt

        # Detect fusion and expansion patterns
        reverse_map = self._build_reverse_onnx_mapping(layers)
        expansion_map = self._detect_layer_expansion(reverse_map, layers)
        fusion_map = self._detect_layer_fusion(layers, subgraph_id)

        if fusion_map:
            print(f"    Detected {len(fusion_map)} fused layers")
        if expansion_map:
            print(f"    Detected {len(expansion_map)} expanded ONNX nodes")

        graph_nodes, graph_edges = self._extract_graph_structure(soup, layers, subgraph_id)

        print(f"    Extracted graph: {len(graph_nodes)} nodes, {len(graph_edges)} edges")

        return {
            'filepath': filepath,
            'netlog_filepath': netlog_filepath,
            'num_layers': len(layers),
            'layers': layers,
            'total_gmacs': macs_data['total_gmacs'],
            'netlog_layer_macs': macs_data['layer_macs'],
            'graph_nodes': graph_nodes,
            'graph_edges': graph_edges,
            'fusion_map': fusion_map,
            'expansion_map': expansion_map
        }

    def parse_all_subgraphs(self) -> Dict[int, Dict[str, Any]]:
        """Parse all subgraph HTML files and netLog files"""
        print("\nParsing TIDL subgraph HTML and netLog files...")

        subgraph_files = self.find_subgraph_files()

        if not subgraph_files:
            print("  No subgraph HTML files found")
            return {}

        print(f"  Found {len(subgraph_files)} subgraph HTML files")

        tidl_data = {}

        for subgraph_idx, html_filepath, netlog_filepath in subgraph_files:
            try:
                tidl_data[subgraph_idx] = self.parse_subgraph_html(html_filepath, netlog_filepath, subgraph_idx)
            except Exception as e:
                print(f"  Warning: Failed to parse {html_filepath}: {e}")
                continue

        print(f"Successfully parsed {len(tidl_data)} subgraph files")

        return tidl_data


class GraphVizParser:
    """Parser for graphvzInfo.txt"""

    def __init__(self, filepath: str):
        self.filepath = filepath

    def parse(self) -> Dict[int, Dict[str, Any]]:
        """Parse graphvizInfo.txt and extract support status"""
        node_support = {}

        print(f"Parsing graphvizInfo.txt: {self.filepath}")

        with open(self.filepath, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue

                parts = line.split(maxsplit=3)
                if len(parts) < 4:
                    continue

                try:
                    node_id = int(parts[0])
                    node_name = parts[1]
                    node_type = parts[2]

                    diag_match = re.search(r'diagInfo\s+(.+)$', line)
                    diag_info = diag_match.group(1) if diag_match else ''

                    is_supported = 'SUPPORTED' in diag_info and 'UNSUPPORTED' not in diag_info

                    node_support[node_id] = {
                        'supported': is_supported,
                        'diagInfo': diag_info,
                        'node_name': node_name,
                        'node_type': node_type
                    }

                except (ValueError, IndexError) as e:
                    print(f"  Warning: Could not parse line: {line[:50]}...")
                    continue

        supported_count = sum(1 for n in node_support.values() if n['supported'])
        unsupported_count = len(node_support) - supported_count

        print(f"Parsed {len(node_support)} nodes")
        print(f"  Supported: {supported_count}")
        print(f"  Unsupported: {unsupported_count}")

        return node_support


class AllowedNodeParser:
    """Parser for allowednode.txt"""

    def __init__(self, filepath: str):
        self.filepath = filepath

    def parse(self) -> List[Dict[str, Any]]:
        """Parse allowednode.txt and extract subgraph information"""
        print(f"Parsing allowednode.txt: {self.filepath}")

        with open(self.filepath, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f if line.strip()]

        if not lines:
            print("  Warning: Empty allowednode.txt")
            return []

        try:
            num_subgraphs = int(lines[0])
            print(f"  Number of subgraphs: {num_subgraphs}")

            subgraphs = []
            line_idx = 1

            for sg_idx in range(num_subgraphs):
                if line_idx >= len(lines):
                    break

                num_nodes = int(lines[line_idx])
                line_idx += 1

                nodes = []
                for _ in range(num_nodes):
                    if line_idx >= len(lines):
                        break
                    nodes.append(int(lines[line_idx]))
                    line_idx += 1

                subgraphs.append({
                    'id': sg_idx,
                    'nodes': nodes
                })

                print(f"  Subgraph {sg_idx}: {num_nodes} nodes")

            print(f"Parsed {len(subgraphs)} subgraphs")
            return subgraphs

        except (ValueError, IndexError) as e:
            print(f"  Error parsing allowednode.txt: {e}")
            return []


def calculate_node_depths_and_positions(nodes, edges, width=1200, height=800):
    """
    Calculate optimal x,y positions for nodes using Netron-style layout

    Args:
        nodes: Dict of node_name -> node_data
        edges: List of edge dictionaries
        width: Canvas width
        height: Canvas height

    Returns:
        Updated nodes dict with x, y, depth, horizontal_position
    """

    node_list = list(nodes.keys())
    node_to_idx = {name: idx for idx, name in enumerate(node_list)}

    children = {name: [] for name in node_list}
    parents = {name: [] for name in node_list}

    for edge in edges:
        src_name = edge.get('source_node_name')
        tgt_name = edge.get('target_node_name')

        if src_name in children and tgt_name in parents:
            children[src_name].append(tgt_name)
            parents[tgt_name].append(src_name)

    roots = [name for name in node_list if len(parents[name]) == 0]
    if not roots:
        roots = [node_list[0]]

    print(f"  Found {len(roots)} root nodes")

    depths = {}

    def assign_depth(node_name, depth):
        current_depth = depths.get(node_name, -1)
        if depth > current_depth:
            depths[node_name] = depth
            for child_name in children[node_name]:
                assign_depth(child_name, depth + 1)

    actual_roots = []
    constant_nodes = []

    for name in roots:
        node_data = nodes.get(name, {})
        node_type = node_data.get('type', '')

        if node_type in ['Constant', 'Initializer'] or 'Constant' in name:
            constant_nodes.append(name)
        else:
            actual_roots.append(name)

    if not actual_roots:
        actual_roots = roots

    print(f"  Actual roots: {len(actual_roots)}")
    print(f"  Constant nodes: {len(constant_nodes)}")

    for root in actual_roots:
        assign_depth(root, 0)

    for const_name in constant_nodes:
        if const_name not in depths:
            child_depths = [depths.get(child, 0) for child in children[const_name]]
            if child_depths:
                depths[const_name] = max(0, min(child_depths) - 1)
            else:
                depths[const_name] = 0

    for name in node_list:
        if name not in depths:
            depths[name] = 0

    max_depth = max(depths.values()) if depths else 0
    print(f"  Max depth: {max_depth}")

    depth_groups = {}
    for name, depth in depths.items():
        if depth not in depth_groups:
            depth_groups[depth] = []
        depth_groups[depth].append(name)

    horizontal_positions = {}

    for depth in sorted(depth_groups.keys()):
        nodes_at_depth = depth_groups[depth]

        if depth == 0:
            for i, name in enumerate(sorted(nodes_at_depth)):
                horizontal_positions[name] = i
        else:
            node_scores = []
            for name in nodes_at_depth:
                parent_positions = [
                    horizontal_positions.get(p, 0)
                    for p in parents[name]
                    if p in horizontal_positions
                ]
                avg_pos = sum(parent_positions) / len(parent_positions) if parent_positions else 0
                node_scores.append((name, avg_pos))

            node_scores.sort(key=lambda x: x[1])

            for i, (name, _) in enumerate(node_scores):
                horizontal_positions[name] = i

    max_width = max(len(nodes) for nodes in depth_groups.values()) if depth_groups else 1
    print(f"  Max width: {max_width}")

    VERTICAL_SPACING = 150
    HORIZONTAL_SPACING = 200
    PADDING = 100

    for name in node_list:
        depth = depths.get(name, 0)
        h_pos = horizontal_positions.get(name, 0)

        layer_width = len(depth_groups.get(depth, [])) * HORIZONTAL_SPACING
        start_x = (width - layer_width) / 2 + HORIZONTAL_SPACING / 2

        nodes[name]['x'] = start_x + h_pos * HORIZONTAL_SPACING
        nodes[name]['y'] = PADDING + depth * VERTICAL_SPACING
        nodes[name]['depth'] = depth
        nodes[name]['horizontal_position'] = h_pos

    print(f"  Calculated positions for {len(nodes)} nodes")

    return nodes


class ONNXParser:
    """Parser for ONNX models using GraphSurgeon for better readability and efficiency"""

    def __init__(self, model_path: str):
        self.model_path = model_path
        self.model = None
        self.gs_graph = None
        self.use_gs = HAS_GRAPHSURGEON

    def load_model(self):
        """Load ONNX model using GraphSurgeon or raw ONNX"""
        print(f"Loading ONNX model: {self.model_path}")
        self.model = onnx.load(self.model_path)
        onnx.checker.check_model(self.model)

        if self.use_gs:
            self.gs_graph = gs.import_onnx(self.model)
            print(f"Model loaded with GraphSurgeon (nodes: {len(self.gs_graph.nodes)}, tensors: {len(self.gs_graph.tensors())})")
        else:
            print("Model loaded with raw ONNX API")

    def get_tensor_shape(self, tensor) -> List:
        """Extract shape from GraphSurgeon tensor or ONNX ValueInfo

        Args:
            tensor: gs.Variable, gs.Constant, or onnx ValueInfoProto

        Returns:
            List of dimensions (ints or 'var' for dynamic)
        """
        if self.use_gs and (isinstance(tensor, gs.Variable) or isinstance(tensor, gs.Constant)):
            if tensor.shape is None:
                # Try to infer shape from values for Constants
                if isinstance(tensor, gs.Constant) and tensor.values is not None:
                    try:
                        import numpy as np
                        values = np.asarray(tensor.values)
                        return list(values.shape)
                    except:
                        pass
                return []
            shape = []
            for dim in tensor.shape:
                if isinstance(dim, int):
                    shape.append(dim)
                else:
                    shape.append('var')
            return shape
        else:
            if not hasattr(tensor, 'type') or not tensor.type.tensor_type.HasField('shape'):
                return []
            shape = []
            for dim in tensor.type.tensor_type.shape.dim:
                if dim.HasField('dim_value'):
                    shape.append(int(dim.dim_value))
                elif dim.HasField('dim_param'):
                    shape.append('var')
                else:
                    shape.append('var')
            return shape

    def format_shapes(self, shapes: List[List]) -> str:
        """Format shapes for display"""
        if not shapes:
            return 'N/A'

        formatted = []
        for shape in shapes:
            if not shape:
                continue

            shape_str = ','.join(map(str, shape))
            formatted.append(f"[{shape_str}]")

        return ', '.join(formatted) if formatted else 'N/A'

    def get_tensor_dtype(self, tensor) -> str:
        """Get data type of a tensor"""
        try:
            if hasattr(tensor, 'type') and tensor.type:
                if tensor.type.HasField('tensor_type'):
                    elem_type = tensor.type.tensor_type.elem_type
                    dtype_map = {
                        1: 'float32', 2: 'uint8', 3: 'int8', 4: 'uint16',
                        5: 'int16', 6: 'int32', 7: 'int64', 8: 'string',
                        9: 'bool', 10: 'float16', 11: 'float64', 12: 'uint32',
                        13: 'uint64', 14: 'complex64', 15: 'complex128',
                        16: 'bfloat16'
                    }
                    return dtype_map.get(elem_type, f'unknown({elem_type})')
            elif hasattr(tensor, 'data_type'):
                dtype_map = {
                    1: 'float32', 2: 'uint8', 3: 'int8', 4: 'uint16',
                    5: 'int16', 6: 'int32', 7: 'int64', 8: 'string',
                    9: 'bool', 10: 'float16', 11: 'float64', 12: 'uint32',
                    13: 'uint64', 14: 'complex64', 15: 'complex128',
                    16: 'bfloat16'
                }
                return dtype_map.get(tensor.data_type, f'unknown({tensor.data_type})')
        except:
            pass
        return 'unknown'

    def extract_node_attributes(self, node) -> Dict[str, Any]:
        """Extract attributes from GraphSurgeon node or ONNX node"""
        if self.use_gs and isinstance(node, gs.Node):
            attrs = {}
            for key, val in node.attrs.items():
                if isinstance(val, np.ndarray):
                    attrs[key] = val.tolist()
                elif isinstance(val, (list, tuple)):
                    attrs[key] = list(val)
                elif isinstance(val, bytes):
                    attrs[key] = val.decode('utf-8')
                elif isinstance(val, (int, float, bool, str, type(None))):
                    attrs[key] = val
                else:
                    # gs.Constant, gs.Graph, onnx.TensorProto, etc. — not JSON-serializable;
                    # the 'gs' substring check used here previously was unreliable because
                    # 'onnx_graphsurgeon' does not contain 'gs' as a substring.
                    attrs[key] = str(val)
            return attrs
        else:
            attrs = {}
            for attr in node.attribute:
                name = attr.name
                if attr.type == onnx.AttributeProto.INTS:
                    attrs[name] = list(attr.ints)
                elif attr.type == onnx.AttributeProto.INT:
                    attrs[name] = int(attr.i)
                elif attr.type == onnx.AttributeProto.FLOATS:
                    attrs[name] = list(attr.floats)
                elif attr.type == onnx.AttributeProto.FLOAT:
                    attrs[name] = float(attr.f)
                elif attr.type == onnx.AttributeProto.STRING:
                    attrs[name] = attr.s.decode('utf-8')
                elif attr.type == onnx.AttributeProto.STRINGS:
                    attrs[name] = [s.decode('utf-8') for s in attr.strings]
                elif attr.type == onnx.AttributeProto.TENSOR:
                    attrs[name] = 'Tensor'
            return attrs

    def parse(self) -> Dict[str, Any]:
        """Parse ONNX model and extract all information using GraphSurgeon or raw ONNX"""
        self.load_model()

        model_name = Path(self.model_path).stem

        if self.use_gs:
            return self._parse_with_graphsurgeon(model_name)
        else:
            return self._parse_with_onnx(model_name)

    def _parse_with_graphsurgeon(self, model_name: str) -> Dict[str, Any]:
        """Parse using GraphSurgeon API"""
        graph = self.gs_graph

        total_params = 0
        for tensor in graph.tensors().values():
            if isinstance(tensor, gs.Constant) and tensor.values is not None:
                total_params += tensor.values.size

        print(f"Total parameters: {total_params:,}")

        tensor_metadata = {}
        tensor_dict = graph.tensors()
        constant_names = set()

        for tensor_name, tensor in tensor_dict.items():
            if isinstance(tensor, gs.Constant):
                constant_names.add(tensor_name)
                shape = self.get_tensor_shape(tensor)
                dtype = 'unknown'
                if tensor.values is not None and hasattr(tensor.values, 'dtype'):
                    dtype_map = {
                        'float32': 'float32', 'uint8': 'uint8', 'int8': 'int8',
                        'int16': 'int16', 'int32': 'int32', 'int64': 'int64',
                        'float16': 'float16', 'float64': 'float64', 'bool': 'bool'
                    }
                    dtype = dtype_map.get(str(tensor.values.dtype), str(tensor.values.dtype))

                tensor_metadata[tensor_name] = {
                    'shape': shape,
                    'dtype': dtype,
                    'is_constant': True
                }

        for inp in graph.inputs:
            if inp.name not in constant_names:
                shape = self.get_tensor_shape(inp)
                dtype = 'unknown'
                if hasattr(inp, 'dtype') and inp.dtype is not None:
                    dtype_map = {
                        1: 'float32', 2: 'uint8', 3: 'int8', 6: 'int32', 7: 'int64',
                        10: 'float16', 11: 'float64', 9: 'bool'
                    }
                    dtype = dtype_map.get(inp.dtype, f'type_{inp.dtype}')

                tensor_metadata[inp.name] = {
                    'shape': shape,
                    'dtype': dtype,
                    'is_constant': False
                }
                print(f"Input: {inp.name} -> {shape} ({dtype})")

        for out in graph.outputs:
            shape = self.get_tensor_shape(out)
            dtype = 'unknown'
            if hasattr(out, 'dtype') and out.dtype is not None:
                dtype_map = {
                    1: 'float32', 2: 'uint8', 3: 'int8', 6: 'int32', 7: 'int64',
                    10: 'float16', 11: 'float64', 9: 'bool'
                }
                dtype = dtype_map.get(out.dtype, f'type_{out.dtype}')

            if out.name not in tensor_metadata:
                tensor_metadata[out.name] = {
                    'shape': shape,
                    'dtype': dtype,
                    'is_constant': False
                }
            print(f"Output: {out.name} -> {shape} ({dtype})")

        for tensor_name, tensor in tensor_dict.items():
            if tensor_name not in tensor_metadata and not isinstance(tensor, gs.Constant):
                shape = self.get_tensor_shape(tensor)
                dtype = 'unknown'
                if hasattr(tensor, 'dtype') and tensor.dtype is not None:
                    dtype_map = {
                        1: 'float32', 2: 'uint8', 3: 'int8', 6: 'int32', 7: 'int64',
                        10: 'float16', 11: 'float64', 9: 'bool'
                    }
                    dtype = dtype_map.get(tensor.dtype, f'type_{tensor.dtype}')

                tensor_metadata[tensor_name] = {
                    'shape': shape,
                    'dtype': dtype,
                    'is_constant': False
                }

        print(f"Total tensors tracked: {len(tensor_metadata)} (constants: {len(constant_names)})")

        shape_lookup = {name: meta['shape'] for name, meta in tensor_metadata.items()}

        model_details = {
            'name': model_name,
            'weights': total_params,
            'no_of_layers': len(graph.nodes),
            'input_shape': [
                {
                    'name': inp.name,
                    'shape': self.get_tensor_shape(inp)
                } for inp in graph.inputs
            ],
            'output_shape': [
                {
                    'name': out.name,
                    'shape': self.get_tensor_shape(out)
                } for out in graph.outputs
            ]
        }

        # Get model opset version for schema lookup
        opset_version = 18  # default
        try:
            if hasattr(onnx_model, 'opset_import'):
                for opset in onnx_model.opset_import:
                    if opset.domain == '' or opset.domain == 'ai.onnx':
                        opset_version = opset.version
                        break
        except Exception:
            pass

        def get_input_param_names(op_type, num_inputs):
            """Get formal ONNX input parameter names for an operator."""
            try:
                import onnx as _onnx
                schema = _onnx.defs.get_schema(op_type, opset_version)
                names = [inp.name for inp in schema.inputs]
                while len(names) < num_inputs:
                    names.append(f'input_{len(names)}')
                return names
            except Exception:
                return [f'input_{i}' for i in range(num_inputs)]

        output_to_node = {}
        for idx, node in enumerate(graph.nodes):
            for out_tensor in node.outputs:
                output_to_node[out_tensor.name] = idx

        layer_details = {}
        edges = []

        print(f"\nProcessing {len(graph.nodes)} nodes...")

        for idx, node in enumerate(graph.nodes):
            node_name = node.name if node.name else f"{node.op}_{idx}"
            attrs = self.extract_node_attributes(node)

            input_names = [inp_tensor.name for inp_tensor in node.inputs]
            output_names = [out_tensor.name for out_tensor in node.outputs]

            # Get formal input parameter names for this op type
            param_names = get_input_param_names(node.op, len(input_names))

            input_metadata = []
            for inp_idx, inp_name in enumerate(input_names):
                meta = tensor_metadata.get(inp_name, {})
                input_metadata.append({
                    'tensor_name': inp_name,
                    'param_name': param_names[inp_idx] if inp_idx < len(param_names) else f'input_{inp_idx}',
                    'shape': meta.get('shape', []),
                    'dtype': meta.get('dtype', 'unknown'),
                    'is_constant': meta.get('is_constant', False)
                })

            output_metadata = []
            for out_name in output_names:
                meta = tensor_metadata.get(out_name, {})
                output_metadata.append({
                    'tensor_name': out_name,
                    'shape': meta.get('shape', []),
                    'dtype': meta.get('dtype', 'unknown'),
                    'is_constant': False
                })

            layer_details[node_name] = {
                'layer_name': node_name,
                'node_index': idx,  # Add node index for TIDL support matching
                'type': node.op,
                'input': input_names,
                'output': output_names,
                'input_metadata': input_metadata,
                'output_metadata': output_metadata,
                'attributes': attrs,
                'x': 0,
                'y': 0,
                'depth': 0,
                'horizontal_position': 0
            }

            for input_name in input_names:
                source_idx = output_to_node.get(input_name)
                if source_idx is not None and source_idx != idx:
                    source_node = graph.nodes[source_idx]
                    source_name = source_node.name if source_node.name else f"{source_node.op}_{source_idx}"

                    edges.append({
                        'source_node_id': source_idx,
                        'target_node_id': idx,
                        'source_node_name': source_name,
                        'target_node_name': node_name,
                        'connection_info': {
                            'tensor': input_name,
                            'shape': shape_lookup.get(input_name, [])
                        }
                    })

            if (idx + 1) % 10 == 0:
                print(f"  Processed {idx + 1}/{len(graph.nodes)} nodes")

        print(f"Extracted {len(layer_details)} layers and {len(edges)} edges")

        print("\nCalculating node positions...")
        layer_details = calculate_node_depths_and_positions(
            layer_details,
            edges,
            width=1200,
            height=800
        )

        return {
            'model_details': model_details,
            'layer_details': layer_details,
            'edges': edges
        }

    def _parse_with_onnx(self, model_name: str) -> Dict[str, Any]:
        """Fallback: Parse using raw ONNX API"""
        graph = self.model.graph

        total_params = 0
        for initializer in graph.initializer:
            dims = list(initializer.dims)
            if dims:
                size = 1
                for dim in dims:
                    size *= dim
                total_params += size

        print(f"Total parameters: {total_params:,}")

        tensor_metadata = {}
        initializer_names = set()

        for init in graph.initializer:
            numeric_dims = [int(d) for d in init.dims]
            # If dims is empty, try to infer shape from data size
            if not numeric_dims:
                # Check various data fields to infer shape
                data_length = 0
                if init.float_data:
                    data_length = len(init.float_data)
                elif init.int32_data:
                    data_length = len(init.int32_data)
                elif init.int64_data:
                    data_length = len(init.int64_data)
                elif init.raw_data:
                    # Calculate based on data type size
                    type_sizes = {1: 4, 2: 1, 3: 1, 6: 4, 7: 8, 10: 2, 11: 8}  # float32, uint8, int8, int32, int64, float16, float64
                    if init.data_type in type_sizes:
                        data_length = len(init.raw_data) // type_sizes[init.data_type]

                if data_length > 0:
                    numeric_dims = [data_length]

            initializer_names.add(init.name)
            tensor_metadata[init.name] = {
                'shape': numeric_dims,
                'dtype': self.get_tensor_dtype(init),
                'is_constant': True
            }

        for inp in graph.input:
            if inp.name not in initializer_names:
                shape = self.get_tensor_shape(inp)
                tensor_metadata[inp.name] = {
                    'shape': shape,
                    'dtype': self.get_tensor_dtype(inp),
                    'is_constant': False
                }
                print(f"Input: {inp.name} -> {shape} ({self.get_tensor_dtype(inp)})")

        for out in graph.output:
            shape = self.get_tensor_shape(out)
            if out.name not in tensor_metadata:
                tensor_metadata[out.name] = {
                    'shape': shape,
                    'dtype': self.get_tensor_dtype(out),
                    'is_constant': False
                }
            print(f"Output: {out.name} -> {shape} ({self.get_tensor_dtype(out)})")

        for vi in graph.value_info:
            shape = self.get_tensor_shape(vi)
            if vi.name not in tensor_metadata:
                tensor_metadata[vi.name] = {
                    'shape': shape,
                    'dtype': self.get_tensor_dtype(vi),
                    'is_constant': False
                }

        print(f"Total tensors tracked: {len(tensor_metadata)} (constants: {len(initializer_names)})")

        shape_lookup = {name: meta['shape'] for name, meta in tensor_metadata.items()}

        model_details = {
            'name': model_name,
            'weights': total_params,
            'no_of_layers': len(graph.node),
            'input_shape': [
                {
                    'name': inp.name,
                    'shape': self.get_tensor_shape(inp)
                } for inp in graph.input
            ],
            'output_shape': [
                {
                    'name': out.name,
                    'shape': self.get_tensor_shape(out)
                } for out in graph.output
            ]
        }

        # Get model opset version for schema lookup
        opset_version = 18  # default
        try:
            if hasattr(onnx_model, 'opset_import'):
                for opset in onnx_model.opset_import:
                    if opset.domain == '' or opset.domain == 'ai.onnx':
                        opset_version = opset.version
                        break
        except Exception:
            pass

        def get_input_param_names(op_type, num_inputs):
            """Get formal ONNX input parameter names for an operator."""
            try:
                import onnx as _onnx
                schema = _onnx.defs.get_schema(op_type, opset_version)
                names = [inp.name for inp in schema.inputs]
                # For variadic inputs, extend with indexed names
                while len(names) < num_inputs:
                    names.append(f'input_{len(names)}')
                return names
            except Exception:
                return [f'input_{i}' for i in range(num_inputs)]

        output_to_node = {}
        for idx, node in enumerate(graph.node):
            for output in node.output:
                output_to_node[output] = idx

        layer_details = {}
        edges = []

        print(f"\nProcessing {len(graph.node)} nodes...")

        for idx, node in enumerate(graph.node):
            node_name = node.name if node.name else f"{node.op_type}_{idx}"
            attrs = self.extract_node_attributes(node)

            # Get formal input parameter names for this op type
            param_names = get_input_param_names(node.op_type, len(node.input))

            input_metadata = []
            for inp_idx, inp_name in enumerate(node.input):
                meta = tensor_metadata.get(inp_name, {})
                input_metadata.append({
                    'tensor_name': inp_name,
                    'param_name': param_names[inp_idx] if inp_idx < len(param_names) else f'input_{inp_idx}',
                    'shape': meta.get('shape', []),
                    'dtype': meta.get('dtype', 'unknown'),
                    'is_constant': meta.get('is_constant', False)
                })

            output_metadata = []
            for out_name in node.output:
                meta = tensor_metadata.get(out_name, {})
                output_metadata.append({
                    'tensor_name': out_name,
                    'shape': meta.get('shape', []),
                    'dtype': meta.get('dtype', 'unknown'),
                    'is_constant': False
                })

            layer_details[node_name] = {
                'layer_name': node_name,
                'node_index': idx,  # Add node index for TIDL support matching
                'type': node.op_type,
                'input': list(node.input),
                'output': list(node.output),
                'input_metadata': input_metadata,
                'output_metadata': output_metadata,
                'attributes': attrs,
                'x': 0,
                'y': 0,
                'depth': 0,
                'horizontal_position': 0
            }

            for input_name in node.input:
                source_idx = output_to_node.get(input_name)
                if source_idx is not None and source_idx != idx:
                    source_node = graph.node[source_idx]
                    source_name = source_node.name if source_node.name else f"{source_node.op_type}_{source_idx}"

                    edges.append({
                        'source_node_id': source_idx,
                        'target_node_id': idx,
                        'source_node_name': source_name,
                        'target_node_name': node_name,
                        'connection_info': {
                            'tensor': input_name,
                            'shape': shape_lookup.get(input_name, [])
                        }
                    })

            if (idx + 1) % 10 == 0:
                print(f"  Processed {idx + 1}/{len(graph.node)} nodes")

        print(f"Extracted {len(layer_details)} layers and {len(edges)} edges")

        print("\nCalculating node positions...")
        layer_details = calculate_node_depths_and_positions(
            layer_details,
            edges,
            width=1200,
            height=800
        )

        return {
            'model_details': model_details,
            'layer_details': layer_details,
            'edges': edges
        }


def build_hierarchical_tree(layer_details: Dict[str, Any], edges: List[Dict]) -> Dict[str, Any]:
    """Build hierarchical tree structure from layer details based on node naming"""
    print("\nBuilding hierarchical tree structure...")

    node_topo_indices = {name: idx for idx, name in enumerate(layer_details.keys())}

    tree_dict = {}

    for node_name, node_data in layer_details.items():
        node_path = []
        for part in node_name.split('/'):
            node_path.extend(part.split('.'))

        if not node_path:
            continue

        node_info = {
            "name": node_name,
            "op": node_data.get('type', 'Unknown'),
            "is_leaf": True,
            "node_details": {
                "name": node_name,
                "op": node_data.get('type', 'Unknown'),
                "inputs": node_data.get('input', []),
                "outputs": node_data.get('output', []),
                "input_metadata": node_data.get('input_metadata', []),
                "output_metadata": node_data.get('output_metadata', []),
                "attributs": node_data.get('attributes', {})
            },
            "topo_idx": node_topo_indices.get(node_name, 0)
        }

        current_level = tree_dict

        for i, part in enumerate(node_path):
            is_last = (i == len(node_path) - 1)

            if part not in current_level:
                if is_last:
                    current_level[part] = node_info
                else:
                    current_level[part] = {
                        "is_leaf": False,
                        "children": {}
                    }
            elif not is_last:
                current_level[part]["is_leaf"] = False

            if not is_last:
                if "children" not in current_level[part]:
                    current_level[part]["children"] = {}
                current_level = current_level[part]["children"]

    print(f"  Built tree with {len(tree_dict)} top-level modules")

    tree_dict = flatten_single_child_modules(tree_dict)

    return tree_dict


def flatten_single_child_modules(tree: Dict, parent_key: str = '') -> Dict:
    """Recursively flatten modules with only one child module"""
    flattened = {}

    for key, value in tree.items():
        if not value.get("is_leaf", False) and "children" in value:
            value["children"] = flatten_single_child_modules(value["children"], key)
            children = value["children"]

            if len(children) == 1:
                child_key = list(children.keys())[0]
                child_value = children[child_key]

                combined_key = f"{key}/{child_key}"

                if not child_value.get("is_leaf", False) and "children" in child_value:
                    flattened[combined_key] = child_value
                else:
                    flattened[combined_key] = child_value
            else:
                min_topo_idx = float('inf')

                def find_min_topo_idx(node):
                    nonlocal min_topo_idx
                    if node.get("is_leaf", False):
                        if "topo_idx" in node:
                            min_topo_idx = min(min_topo_idx, node["topo_idx"])
                    elif "children" in node:
                        for child_node in node["children"].values():
                            find_min_topo_idx(child_node)

                find_min_topo_idx(value)
                value["min_topo_idx"] = min_topo_idx if min_topo_idx != float('inf') else 0
                flattened[key] = value
        else:
            flattened[key] = value

    return flattened


def discover_files_from_workdir(model_dir_path: str) -> Dict[str, str]:
    """Auto-discover all required files from model directory structure"""
    discovered = {}

    print(f"\nAuto-discovering files in: {model_dir_path}")
    print("=" * 70)

    import glob

    onnx_files = glob.glob(os.path.join(model_dir_path, 'model/*.onnx'), recursive=False)
    if onnx_files:
        discovered['onnx'] = onnx_files[0]
        print(f"[FOUND] ONNX model: {os.path.relpath(discovered['onnx'])}")
    else:
        print("[NOT FOUND] ONNX model not found")

    graphviz_files = glob.glob(os.path.join(model_dir_path, 'artifacts/tempDir/graphvizInfo.txt'), recursive=False)
    if graphviz_files:
        discovered['graphviz'] = graphviz_files[0]
        print(f"[FOUND] graphvizInfo: {os.path.relpath(discovered['graphviz'])}")
    else:
        print("[NOT FOUND] graphvizInfo.txt not found")

    allowednode_files = glob.glob(os.path.join(model_dir_path, 'artifacts/allowedNode.txt'), recursive=False)
    if allowednode_files:
        discovered['allowednode'] = allowednode_files[0]
        print(f"[FOUND] allowedNode: {os.path.relpath(discovered['allowednode'])}")
    else:
        print("[NOT FOUND] allowedNode.txt not found")

    subgraph_dir = None
    subgraph_html_files = glob.glob(os.path.join(model_dir_path, 'artifacts/tempDir/subgraph_*_tidl_net.bin.html'), recursive=False)
    if subgraph_html_files:
        subgraph_dir = os.path.dirname(subgraph_html_files[0])
        discovered['subgraph_dir'] = subgraph_dir
        print(f"[FOUND] {len(subgraph_html_files)} subgraph HTML files in: {os.path.relpath(subgraph_dir)}")
    else:
        # Check if SVG files exist (TIDL Tools < 11.02 generates SVG instead of HTML)
        subgraph_svg_files = glob.glob(os.path.join(model_dir_path, 'artifacts/tempDir/subgraph_*_tidl_net.bin.svg'), recursive=False)
        if subgraph_svg_files:
            svg_dir = os.path.relpath(os.path.dirname(subgraph_svg_files[0]))
            error_msg = f"""
================================================================================
ERROR: TIDL Tools Version Incompatibility
================================================================================

Found {len(subgraph_svg_files)} SVG subgraph file(s) in: {svg_dir}

Your TIDL Tools version (< 11.02) generates SVG files for subgraphs.
Model Inspector requires HTML files which are only available in TIDL Tools 11.02+

Detected Issue:
  - Old TIDL Tools versions (11.01, 11.00, and earlier) generate SVG format
  - Model Inspector only supports HTML format (TIDL Tools 11.02+)

Solution:
  1. Upgrade to TIDL Tools version 11.02 or later
  2. Recompile your model with the new version
  3. Run the inspect command again

Note: Model Inspector cannot parse SVG format - HTML format is required.

================================================================================
"""
            raise RuntimeError(error_msg.strip())
        else:
            print("[NOT FOUND] Subgraph HTML files not found")

    xlsx_files = glob.glob(os.path.join(model_dir_path, 'analyze.xlsx'), recursive=False)
    if xlsx_files:
        discovered['xlsx'] = xlsx_files[0]
        print(f"[FOUND] analyze.xlsx: {os.path.relpath(discovered['xlsx'])}")
    else:
        print("[NOT FOUND] analyze.xlsx not found")

    mapping_files = glob.glob(os.path.join(model_dir_path, 'layer_output_mapping*.yaml'), recursive=False)
    if mapping_files:
        notidl_tidl_files = [f for f in mapping_files if 'notidl_tidl.yaml' in f and 'tidl32' not in f]
        if notidl_tidl_files:
            discovered['activation_yaml'] = notidl_tidl_files[0]
        else:
            discovered['activation_yaml'] = mapping_files[0]
        print(f"[FOUND] activation mapping: {os.path.relpath(discovered['activation_yaml'])}")
    else:
        print("WARNING: Activation mapping YAML not found (optional)")

    print("=" * 70)
    return discovered


def load_config_data(model_dir_path: str) -> Dict[str, Any]:
    """Load configuration data from model directory (config.yaml and result.yaml)"""
    config_data = {}

    import glob
    config_files = glob.glob(os.path.join(model_dir_path, '**/config.yaml'), recursive=True)
    result_files = glob.glob(os.path.join(model_dir_path, '**/result.yaml'), recursive=True)

    config_path = None
    result_path = None

    for path in config_files:
        if '/tidl/' not in path and '/tidl32/' not in path and '/notidl/' not in path:
            config_path = path
            break
    if not config_path:
        for path in config_files:
            if '/tidl/' in path:
                config_path = path
                break
    if not config_path and config_files:
        config_path = config_files[0]

    for path in result_files:
        if '/tidl/' not in path and '/tidl32/' not in path and '/notidl/' not in path:
            result_path = path
            break
    if not result_path:
        for path in result_files:
            if '/tidl/' in path:
                result_path = path
                break
    if not result_path and result_files:
        result_path = result_files[0]

    if config_path and os.path.exists(config_path):
        try:
            import yaml
            with open(config_path, 'r') as f:
                config = yaml.safe_load(f)

            config_data['target_device'] = config.get('session', {}).get('target_device', 'Unknown')
            config_data['task_type'] = config.get('common', {}).get('task_type', 'Unknown')
            config_data['tensor_bits'] = config.get('session', {}).get('runtime_options', {}).get('tensor_bits', 'Unknown')

            metric_ref = config.get('model_info', {}).get('metric_reference', {})
            if 'accuracy_top1%' in metric_ref:
                config_data['accuracy'] = f"{metric_ref['accuracy_top1%']}%"
            else:
                config_data['accuracy'] = 'N/A'

            print(f"Loaded config data from: {config_path}")
        except Exception as e:
            print(f"WARNING: Failed to load config.yaml: {e}")

    if result_path and os.path.exists(result_path):
        try:
            import yaml
            with open(result_path, 'r') as f:
                result = yaml.safe_load(f)

            result_data = result.get('result', {})
            config_data['num_frames'] = result_data.get('num_frames', 'N/A')
            config_data['num_subgraphs'] = result_data.get('num_subgraphs', 'N/A')
            config_data['perfsim_ddr_transfer_mb'] = result_data.get('perfsim_ddr_transfer_mb', 'N/A')
            config_data['perfsim_gmacs'] = result_data.get('perfsim_gmacs', 'N/A')
            config_data['perfsim_time_ms'] = result_data.get('perfsim_time_ms', 'N/A')

            print(f"Loaded result data from: {result_path}")
            print(f"  -> num_frames: {config_data.get('num_frames')}")
            print(f"  -> num_subgraphs: {config_data.get('num_subgraphs')}")
            print(f"  -> perfsim_ddr_transfer_mb: {config_data.get('perfsim_ddr_transfer_mb')}")
            print(f"  -> perfsim_gmacs: {config_data.get('perfsim_gmacs')}")
            print(f"  -> perfsim_time_ms: {config_data.get('perfsim_time_ms')}")
        except Exception as e:
            print(f"WARNING: Failed to load result.yaml: {e}")

    return config_data


# ---------------------------------------------------------------------------
# EVM hardware performance (reads /tmp/tidl_trace_subgraph_<N>_perf.csv)
# ---------------------------------------------------------------------------

def is_evm_perf_locked(json_path: str) -> bool:
    """Return True if the JSON already contains real EVM hardware perf data.
    Detected by presence of infer_time_subgraph_ms in metadata (written by EVM infer).
    Once locked, PC simulation data must not overwrite it."""
    try:
        with open(json_path, 'r') as fh:
            d = json.load(fh)
        meta = d.get('metadata', {})
        return (
            meta.get('performance_source') == 'evm_hardware' or
            meta.get('infer_time_subgraph_ms') is not None
        )
    except Exception:
        return False


def load_accuracy_from_result_yaml(run_dir: str) -> Dict[str, Any]:
    """Read accuracy and timing fields from result.yaml in run_dir.
    Returns an empty dict if the file is missing or unreadable.
    Only call this when running on EVM (caller is responsible for the check).
    """
    import glob as _glob
    import yaml as _yaml

    result_files = _glob.glob(os.path.join(run_dir, '**/result.yaml'), recursive=True)
    result_path = None
    for p in result_files:
        if '/tidl/' not in p and '/tidl32/' not in p and '/notidl/' not in p:
            result_path = p
            break
    if not result_path and result_files:
        result_path = result_files[0]

    if not result_path:
        return {}

    try:
        with open(result_path, 'r') as fh:
            data = _yaml.safe_load(fh)

        # Only use timing when the result was generated on real EVM hardware.
        # On PC (compile or evaluate), timing comes from simulation and num_frames
        # may refer to calibration frames — not real inference time.
        target_machine = data.get('session', {}).get('target_machine', '')
        timing_valid = (target_machine == 'evm')

        r = data.get('result', {})
        accuracy = {}

        if timing_valid:
            for key in ('infer_time_subgraph_ms', 'infer_time_core_ms',
                        'infer_time_invoke_ms', 'num_frames'):
                if key in r:
                    accuracy[key] = r[key]
        else:
            print(f'  NOTE: result.yaml is from target_machine={target_machine!r} '
                  f'— skipping timing fields (only valid on EVM)')

        # Accuracy metrics are valid from any pipeline (evaluate always writes them)
        for key, val in r.items():
            if key.lower().startswith('accuracy') and isinstance(val, (int, float)):
                accuracy[key] = val

        accuracy['_result_path'] = result_path
        print(f'  Loaded result.yaml from: {result_path}')
        return accuracy
    except Exception as exc:
        print(f'  WARNING: Failed to read result.yaml: {exc}')
        return {}

# C7x DSP clock on J721E/AM6xA in MHz — used to convert cycles → µs.
# Override with env var C7X_DSP_FREQ_MHZ if your board runs at a different speed.
_C7X_FREQ_MHZ = int(os.environ.get('C7X_DSP_FREQ_MHZ', '1000'))


def load_evm_perf_csv(csv_path: str) -> Dict[int, Dict[str, Any]]:
    """Parse a /tmp/tidl_trace_subgraph_<N>_perf.csv written by TIDL on the EVM.

    Returns {layer_id: {cycle fields}} for each data row.
    The trailing 'Sum of Layer Cycles …' line is silently skipped.
    Column headers in the file have extra whitespace — all keys are stripped.
    """
    import csv as _csv
    result: Dict[int, Dict[str, Any]] = {}
    try:
        with open(csv_path, 'r') as fh:
            reader = _csv.reader(fh)
            raw_headers = next(reader)
            headers = [h.strip() for h in raw_headers]
            for raw_row in reader:
                row = {headers[i]: v.strip()
                       for i, v in enumerate(raw_row) if i < len(headers)}
                layer_str = row.get('Layer', '')
                if not layer_str.lstrip('-').isdigit():
                    continue  # skip blank lines and the final "Sum" line
                layer_id = int(layer_str)

                def _int(col: str) -> int:
                    v = row.get(col, '').strip()
                    return int(v) if v.lstrip('-').isdigit() else 0

                result[layer_id] = {
                    'layer_cycles':       _int('Layer Cycles'),
                    'kernel_cycles':      _int('kernelOnlyCycles'),
                    'core_loop_cycles':   _int('coreLoopCycles'),
                    'layer_setup_cycles': _int('LayerSetupCycles'),
                    'dma_pipeup_cycles':  _int('dmaPipeupCycles'),
                    'ddr_read_bytes':     _int('DDRBWReadInBytes'),
                    'ddr_write_bytes':    _int('DDRBWWriteInBytes'),
                }
    except Exception as exc:
        print(f'WARNING: Failed to load EVM perf CSV {csv_path}: {exc}')
    return result


def update_with_evm_perf(json_path: str) -> bool:
    """Update inspector JSON in-place with real hardware cycle data from the EVM.

    Scans /tmp/ for tidl_trace_subgraph_<N>_perf.csv files, matches each row
    to the corresponding TIDL layer by layer_id, and writes updated performance
    fields back to *json_path*.  Also returns True if anything was updated.
    """
    import glob as _glob
    import re as _re

    csv_files = _glob.glob('/tmp/tidl_trace_subgraph_*_perf.csv')
    if not csv_files:
        print('  No /tmp/tidl_trace_subgraph_*_perf.csv files found on this device')
        return False

    with open(json_path, 'r', encoding='utf-8') as fh:
        data = json.load(fh)

    updated = False
    for csv_path in sorted(csv_files):
        m = _re.search(r'tidl_trace_subgraph_(\d+)_perf\.csv', csv_path)
        if not m:
            continue
        sg_num = int(m.group(1))

        perf_map = load_evm_perf_csv(csv_path)
        if not perf_map:
            continue

        # Subgraph key may be 'tidl_0' or '0'
        subgraphs = data.get('runtime', {}).get('subgraphs', {})
        subgraph = subgraphs.get(f'tidl_{sg_num}') or subgraphs.get(str(sg_num))
        if not subgraph:
            print(f'  WARNING: subgraph {sg_num} not found in JSON, skipping')
            continue

        matched = 0
        for layer in subgraph.get('layers', []):
            lid = layer.get('layer_id')
            if lid is None or lid not in perf_map:
                continue

            p = perf_map[lid]

            # Replace performance with only the two reliable EVM metrics.
            # Everything else (proctime_us, core_loop_cycles, io_cycles, memory)
            # is set null — HTML hides charts for null fields automatically.
            layer['performance'] = {
                'layer_cycles':     p['layer_cycles'],
                'kernel_cycles':    p['kernel_cycles'],
                'core_loop_cycles': None,
                'proctime_us':      None,
                'io_cycles':        None,
                'memory':           {'l2_kb': None, 'msmc_kb': None, 'ddr_kb': None, 'total_kb': None},
            }
            matched += 1
            updated = True

        print(f'  Subgraph {sg_num}: matched {matched}/{len(perf_map)} layers '
              f'from {os.path.basename(csv_path)}')

    if updated:
        meta = data.setdefault('metadata', {})

        # Stamp the source so html_generator knows this JSON has real EVM cycle data.
        # Stored in metadata (not root) to keep the 3-key top-level structure clean.
        meta['performance_source'] = 'evm_hardware'

        # Load result.yaml timing and write flat into metadata — only when the
        # result.yaml was generated by an actual infer/evaluate run (not compile).
        run_dir = os.path.dirname(os.path.dirname(json_path))  # inspector/ -> model_dir
        acc = load_accuracy_from_result_yaml(run_dir)
        if acc:
            timing_written = False
            for key in ('infer_time_subgraph_ms', 'infer_time_core_ms', 'infer_time_invoke_ms'):
                raw = acc.get(key)
                if raw is not None:
                    # Values are already per-frame averages (basert_wrapper divides
                    # the running sum by num_frames before writing result.yaml).
                    # Do NOT divide again here.
                    meta[key] = round(float(raw), 3)
                    timing_written = True
            if acc.get('num_frames'):
                meta['num_frames'] = acc['num_frames']
            if timing_written:
                print(f'  EVM timing written to metadata')

        # Clean up any old-format keys left from previous runs
        data.pop('performance_source', None)  # remove from root if present
        meta.pop('evm_timing', None)
        meta.pop('evm_accuracy', None)
        for sg_info in data.get('runtime', {}).get('subgraphs', {}).values():
            sg_info.pop('evm_execution_time_ms_per_frame', None)

        with open(json_path, 'w', encoding='utf-8') as fh:
            json.dump(data, fh, indent=2)
        print(f'  EVM perf data written to {json_path}')

    return updated


def load_proctime_data(model_dir_path: str) -> Dict[int, List[Dict[str, Any]]]:
    """Load proctime data from performance CSV files for each subgraph"""
    proctime_data = {}

    import glob
    csv_files = glob.glob(os.path.join(model_dir_path, '**/tempDir/subgraph_*/tempDir_subgraph_*_tidl_net_*.csv'), recursive=True)

    preferred_files = []
    for path in csv_files:
        if '/tidl/' in path:
            preferred_files.append(path)

    if not preferred_files:
        preferred_files = csv_files

    processed_subgraphs = set()

    for csv_path in preferred_files:
        try:
            import re
            match = re.search(r'subgraph_(\d+)_tidl_net', csv_path)
            if not match:
                continue

            subgraph_num = int(match.group(1))

            # Skip if already processed this subgraph
            if subgraph_num in processed_subgraphs:
                continue

            import csv
            with open(csv_path, 'r') as f:
                reader = csv.DictReader(f)
                layer_data = []

                for row in reader:
                    try:
                        layer_num = int(row[' lyrNum '].strip())
                        layer_type = row[' LyrType'].strip()
                        proctime_str = row['procTime(us)'].strip()

                        proctime = float(proctime_str) if proctime_str else 0.0

                        layer_data.append({
                            'layer_num': layer_num,
                            'layer_type': layer_type,
                            'proctime': proctime
                        })
                    except (ValueError, KeyError) as e:
                        continue

                if layer_data:
                    proctime_data[subgraph_num] = layer_data
                    processed_subgraphs.add(subgraph_num)
                    print(f"Loaded proctime data for subgraph {subgraph_num}: {len(layer_data)} layers")

        except Exception as e:
            print(f"WARNING: Failed to load proctime CSV {csv_path}: {e}")

    return proctime_data


def load_cycles_data(model_dir_path: str) -> Dict[int, List[Dict[str, Any]]]:
    """Load kernel and core loop cycles data from performance CSV files for each subgraph"""
    cycles_data = {}

    import glob
    csv_files = glob.glob(os.path.join(model_dir_path, '**/tempDir/subgraph_*/tempDir_subgraph_*_tidl_net_*.csv'), recursive=True)

    preferred_files = []
    for path in csv_files:
        if '/tidl/' in path:
            preferred_files.append(path)

    if not preferred_files:
        preferred_files = csv_files

    processed_subgraphs = set()

    for csv_path in preferred_files:
        try:
            import re
            match = re.search(r'subgraph_(\d+)_tidl_net', csv_path)
            if not match:
                continue

            subgraph_num = int(match.group(1))

            # Skip if already processed this subgraph
            if subgraph_num in processed_subgraphs:
                continue

            import csv
            with open(csv_path, 'r') as f:
                reader = csv.DictReader(f)
                layer_data = []

                for row in reader:
                    try:
                        layer_num = int(row[' lyrNum '].strip())
                        layer_type = row[' LyrType'].strip()
                        kernel_only_cycles_str = row['kernelOnlyCycles'].strip()
                        core_loop_cycles_str = row['coreLoopCycles'].strip()

                        kernel_only_cycles = float(kernel_only_cycles_str) if kernel_only_cycles_str else 0.0
                        core_loop_cycles = float(core_loop_cycles_str) if core_loop_cycles_str else 0.0

                        layer_cycles_str = row.get('LayerCycles', '0').strip()
                        # CSV header has typo 'IOcyles' with leading spaces
                        io_cycles_str = '0'
                        for key in row:
                            if 'IOcyles' in key or 'IOcycles' in key:
                                io_cycles_str = row[key].strip()
                                break
                        layer_cycles = float(layer_cycles_str) if layer_cycles_str else 0.0
                        io_cycles = float(io_cycles_str) if io_cycles_str else 0.0

                        layer_data.append({
                            'layer_num': layer_num,
                            'layer_type': layer_type,
                            'kernelOnlyCycles': kernel_only_cycles,
                            'coreLoopCycles': core_loop_cycles,
                            'layerCycles': layer_cycles,
                            'ioCycles': io_cycles
                        })
                    except (ValueError, KeyError) as e:
                        continue

                if layer_data:
                    cycles_data[subgraph_num] = layer_data
                    processed_subgraphs.add(subgraph_num)
                    print(f"Loaded cycles data for subgraph {subgraph_num}: {len(layer_data)} layers")

        except Exception as e:
            print(f"WARNING: Failed to load cycles CSV {csv_path}: {e}")

    return cycles_data


def load_memory_data(model_dir_path: str) -> Dict[int, List[Dict[str, Any]]]:
    """Load memory usage data from performance CSV files for each subgraph"""
    memory_data = {}

    import glob
    csv_files = glob.glob(os.path.join(model_dir_path, '**/tempDir/subgraph_*/tempDir_subgraph_*_tidl_net_*.csv'), recursive=True)

    preferred_files = []
    for path in csv_files:
        if '/tidl/' in path:
            preferred_files.append(path)

    if not preferred_files:
        preferred_files = csv_files

    processed_subgraphs = set()

    for csv_path in preferred_files:
        try:
            import re
            match = re.search(r'subgraph_(\d+)_tidl_net', csv_path)
            if not match:
                continue

            subgraph_num = int(match.group(1))

            # Skip if already processed this subgraph
            if subgraph_num in processed_subgraphs:
                continue

            import csv
            with open(csv_path, 'r') as f:
                reader = csv.DictReader(f)
                layer_data = []

                for row in reader:
                    try:
                        layer_num = int(row[' lyrNum '].strip())
                        layer_type = row[' LyrType'].strip()

                        in_vol = float(row['inVol(KB)'].strip()) if row['inVol(KB)'].strip() else 0.0
                        out_vol = float(row['outVol(KB)'].strip()) if row['outVol(KB)'].strip() else 0.0
                        wt_vol = float(row['wtVol(KB)'].strip()) if row['wtVol(KB)'].strip() else 0.0

                        src_mem_in = row[' srcMem-IN'].strip()
                        dst_mem_in = row[' dstMem-IN'].strip()
                        src_mem_out = row['srcMem-OUT'].strip()
                        dst_mem_out = row['dstMem-OUT'].strip()
                        src_mem_wt = row[' srcMem-WT'].strip()
                        dst_mem_wt = row[' dstMem-WT'].strip()

                        l2_usage = 0.0
                        msmc_usage = 0.0
                        ddr_usage = 0.0

                        for mem_loc, vol in [(src_mem_in, in_vol), (dst_mem_in, in_vol),
                                             (src_mem_out, out_vol), (dst_mem_out, out_vol),
                                             (src_mem_wt, wt_vol), (dst_mem_wt, wt_vol)]:
                            if 'L2' in mem_loc:
                                l2_usage += vol / 2
                            elif 'MSMC' in mem_loc or 'L3' in mem_loc:
                                msmc_usage += vol / 2
                            elif 'DDR' in mem_loc:
                                ddr_usage += vol / 2

                        layer_data.append({
                            'layer_num': layer_num,
                            'layer_type': layer_type,
                            'l2_usage': l2_usage,
                            'msmc_usage': msmc_usage,
                            'ddr_usage': ddr_usage,
                            'total_usage': l2_usage + msmc_usage + ddr_usage
                        })
                    except (ValueError, KeyError) as e:
                        continue

                if layer_data:
                    memory_data[subgraph_num] = layer_data
                    processed_subgraphs.add(subgraph_num)
                    print(f"Loaded memory data for subgraph {subgraph_num}: {len(layer_data)} layers")

        except Exception as e:
            print(f"WARNING: Failed to load memory CSV {csv_path}: {e}")

    return memory_data


def main(work_dirs_path, output_json_path, extract_activations=False):
    """Main function to extract all artifact data to JSON"""
    if len(sys.argv) < 3:
        print("=" * 70)
        print("Data Extractor - Extract TIDL Artifacts to JSON")
        print("=" * 70)
        print("\nUsage: python data_extractor.py <model_dir/> <output.json> [--act_data=false]")
        print("\nArguments:")
        print("  model_dir/   - Direct path to model directory (e.g., work_dirs/compile/AM69A/cl_onnx_model_name/)")
        print("  output.json  - Output JSON file path (will be compressed)")
        print("  --act_data   - Extract activations data to separate file (enabled by default, use --act_data=false to disable)")
        print("\nExample:")
        print("  python data_extractor.py work_dirs/compile/AM69A/cl-ort-resnet18/ model_data.json")
        print("  python data_extractor.py work_dirs/compile/AM69A/cl-ort-resnet18/ model_data.json --act_data=false")
        print("\nThe script will automatically discover and parse:")
        print("  - ONNX model from <model_dir>/model/*.onnx")
        print("  - GraphViz info from <model_dir>/artifacts/tempDir/graphvizInfo.txt")
        print("  - Allowed nodes from <model_dir>/artifacts/allowedNode.txt")
        print("  - Subgraph files from <model_dir>/artifacts/tempDir/")
        print("  - Metrics from <model_dir>/analyze.xlsx")
        print("  - Activation data from layer_info.txt and binary files (enabled by default)")
        print("  - Config/Result from <model_dir>/tidl/*.yaml")
        print("  - Performance data from <model_dir>/tidl/artifacts/tempDir/**/*.csv")
        print("=" * 70)
        sys.exit(1)

    # Use the function parameters (passed when called programmatically)
    model_dir_path = work_dirs_path
    # output_json_path already set from parameter

    # Check if intermediate JSON already exists
    if os.path.exists(output_json_path):
        file_size = os.path.getsize(output_json_path) / (1024 * 1024)
        print("=" * 70)
        print("Intermediate JSON already exists")
        print("=" * 70)
        print(f"File: {output_json_path} ({file_size:.2f} MB)")
        print("Skipping data extraction (file already present)")
        print("\nTo regenerate, delete the existing file and run again:")
        print(f"  rm {output_json_path}")
        print("=" * 70)
        return

    if not os.path.exists(model_dir_path):
        print(f"ERROR: Model directory not found: {model_dir_path}")
        sys.exit(1)

    if not os.path.isdir(model_dir_path):
        print(f"ERROR: {model_dir_path} is not a directory")
        sys.exit(1)

    discovered = discover_files_from_workdir(model_dir_path)

    # ONNX model is required; TIDL artifacts are optional and skipped gracefully
    if 'onnx' not in discovered:
        print(f"\nERROR: ONNX model not found")
        sys.exit(1)

    onnx_path = discovered['onnx']
    graphviz_path = discovered.get('graphviz')
    allowednode_path = discovered.get('allowednode')
    subgraph_dir = discovered.get('subgraph_dir')

    # Log missing optional TIDL artifacts as info (not error)
    if not graphviz_path:
        print("[INFO] GraphViz info not found (optional) - TIDL support info will be unavailable")
    if not allowednode_path:
        print("[INFO] allowedNode.txt not found (optional) - using all nodes as allowed")
    if not subgraph_dir:
        print("[INFO] Subgraph directory not found (optional) - assuming single subgraph")

    print("\n" + "=" * 70)
    print("Data Extractor - Parsing Artifacts")
    print("=" * 70)
    print(f"ONNX Model:      {onnx_path}")
    print(f"GraphViz Info:   {graphviz_path}")
    print(f"Allowed Nodes:   {allowednode_path}")
    print(f"Subgraph Dir:    {subgraph_dir}")
    print(f"Output JSON:     {output_json_path}")
    print("=" * 70)

    try:
        print("\n[1/8] Parsing ONNX model...")
        onnx_parser = ONNXParser(onnx_path)
        model_data = onnx_parser.parse()

        tree_structure = build_hierarchical_tree(
            model_data['layer_details'],
            model_data['edges']
        )
        model_data['tree_structure'] = tree_structure

        subgraph_data = {
            'subgraphs': [],
            'node_support': {}
        }

        print("\n[2/8] Parsing GraphViz info...")
        if graphviz_path and os.path.exists(graphviz_path):
            graphviz_parser = GraphVizParser(graphviz_path)
            subgraph_data['node_support'] = graphviz_parser.parse()
        else:
            print(f"WARNING: GraphViz info not found (optional)")
            subgraph_data['node_support'] = {}

        print("\n[3/8] Parsing allowed nodes...")
        if allowednode_path and os.path.exists(allowednode_path):
            allowednode_parser = AllowedNodeParser(allowednode_path)
            subgraph_data['subgraphs'] = allowednode_parser.parse()
        else:
            print(f"WARNING: allowedNode.txt not found (optional)")
            subgraph_data['subgraphs'] = {}

        print("\n[4/8] Parsing TIDL subgraph HTML files...")
        # Build tensor name → ONNX node index map for TIDL-to-ONNX mapping
        tensor_to_node_map = {}
        for idx, (layer_name, layer_info) in enumerate(model_data.get('layer_details', {}).items()):
            for out_tensor in layer_info.get('output', []):
                tensor_to_node_map[out_tensor] = idx
        # Build ordered list of ONNX layer names for ONNX name lookup
        onnx_layer_names = list(model_data.get('layer_details', {}).keys())

        # Handle missing subgraph directory gracefully
        if subgraph_dir is None:
            print("[WARNING] Subgraph directory not found (optional) - no TIDL subgraph data available")
            tidl_data = {}
        else:
            tidl_parser = TIDLSubgraphParser(subgraph_dir, subgraph_data['node_support'])
            tidl_parser.tensor_to_node_map = tensor_to_node_map
            tidl_parser.onnx_layer_names = onnx_layer_names
            tidl_parser.onnx_layer_details = model_data.get('layer_details', {})
            tidl_data = tidl_parser.parse_all_subgraphs()

        print("\n[5/8] Parsing activation data...")
        activation_data = {}
        if extract_activations:
            try:
                activation_parser = ActivationDataParser(
                    model_dir=work_dirs_path,
                    frame_idx=0,
                    tidl_data=tidl_data
                )
                activation_data = activation_parser.process_all_layers()
                print(f"  Loaded activation data for {len(activation_data)} layers")
            except Exception as e:
                print(f"  WARNING: Could not load activation data: {e}")
                activation_data = {}
        else:
            print("  Skipping activation data - JSON contains only model structure")

        print("\n[6/8] Parsing metrics data...")
        metrics_data = {}
        metrics_xlsx_path = discovered.get('xlsx', None)
        if metrics_xlsx_path and os.path.exists(metrics_xlsx_path):
            metrics_parser = MetricsParser(metrics_xlsx_path)
            metrics_data = metrics_parser.get_metrics()
        else:
            print("WARNING: No analyze.xlsx file found")

        print("\n[7/8] Loading configuration and performance data...")
        config_data = load_config_data(model_dir_path)
        proctime_data = load_proctime_data(model_dir_path)
        cycles_data = load_cycles_data(model_dir_path)
        memory_data = load_memory_data(model_dir_path)

        print("\n[8/8] Combining and saving data...")

        performance_data = {}
        all_subgraphs = set(proctime_data.keys()) | set(cycles_data.keys()) | set(memory_data.keys())

        for subgraph_num in all_subgraphs:
            proctime_lookup = {layer['layer_num']: layer for layer in proctime_data.get(subgraph_num, [])}
            cycles_lookup = {layer['layer_num']: layer for layer in cycles_data.get(subgraph_num, [])}
            memory_lookup = {layer['layer_num']: layer for layer in memory_data.get(subgraph_num, [])}

            all_layer_nums = set(proctime_lookup.keys()) | set(cycles_lookup.keys()) | set(memory_lookup.keys())

            merged_layers = []
            for layer_num in sorted(all_layer_nums):
                layer_entry = {'layer_num': layer_num}

                if layer_num in proctime_lookup:
                    layer_entry['layer_type'] = proctime_lookup[layer_num]['layer_type']
                    layer_entry['proctime_us'] = proctime_lookup[layer_num]['proctime']

                if layer_num in cycles_lookup:
                    if 'layer_type' not in layer_entry:
                        layer_entry['layer_type'] = cycles_lookup[layer_num]['layer_type']
                    layer_entry['kernel_cycles'] = cycles_lookup[layer_num]['kernelOnlyCycles']
                    layer_entry['core_loop_cycles'] = cycles_lookup[layer_num]['coreLoopCycles']
                    layer_entry['layer_cycles'] = cycles_lookup[layer_num].get('layerCycles', 0)
                    layer_entry['io_cycles'] = cycles_lookup[layer_num].get('ioCycles', 0)

                if layer_num in memory_lookup:
                    if 'layer_type' not in layer_entry:
                        layer_entry['layer_type'] = memory_lookup[layer_num]['layer_type']
                    layer_entry['memory'] = {
                        'l2_kb': memory_lookup[layer_num]['l2_usage'],
                        'msmc_kb': memory_lookup[layer_num]['msmc_usage'],
                        'ddr_kb': memory_lookup[layer_num]['ddr_usage'],
                        'total_kb': memory_lookup[layer_num]['total_usage']
                    }

                merged_layers.append(layer_entry)

            performance_data[subgraph_num] = merged_layers

        enhanced_tidl_data = {}

        # Get node_support for lookups
        node_support = subgraph_data.get('node_support', {})

        for subgraph_id, tidl_info in tidl_data.items():
            enhanced_layers = {}  # Changed from list to dict

            subgraph_metrics = metrics_data.get(str(subgraph_id), [])
            metrics_lookup = {m['tidl_layer_id']: m for m in subgraph_metrics if m.get('tidl_layer_id')}

            subgraph_perf = performance_data.get(subgraph_id, [])
            perf_lookup = {p['layer_num']: p for p in subgraph_perf}

            # Get fusion and expansion maps for this subgraph
            fusion_map = tidl_info.get('fusion_map', {})
            expansion_map = tidl_info.get('expansion_map', {})

            # Build buffer→layer registry from netLog data for node-reference inputs/outputs
            netlog_lm = tidl_info.get('netlog_layer_macs', {})

            # buffer_id → {node_id, name, type, tensor_name, shape}
            # First writer wins — do NOT overwrite (prevents buffer-reuse layers like the
            # output DataLayer from stealing ownership of buffer 0 from the input DataLayer)
            buf_to_layer = {}
            for lyr in tidl_info.get('layers', []):
                li = lyr['layer_index']
                lm = netlog_lm.get(li, {})
                ob = lm.get('outbuf_id', li)
                if ob in buf_to_layer:
                    continue  # buffer already claimed by an earlier layer
                out_shape = []
                for o in lyr.get('parameters', {}).get('outputs', []):
                    if 'dims' in o:
                        out_shape = o['dims']
                        break
                buf_to_layer[ob] = {
                    'node_id': li,
                    'name': lyr['layer_name'],
                    'type': lyr['layer_type'],
                    'tensor_name': lm.get('output_name', lyr['layer_name']),
                    'shape': out_shape
                }

            for layer in tidl_info.get('layers', []):
                layer_idx = layer['layer_index']
                tidl_layer_id = str(layer_idx)

                params = layer.get('parameters', {})
                lm = netlog_lm.get(layer_idx, {})

                # Build inputs: each inbuf_id → node reference + tensor_name + shape
                inbuf_ids = lm.get('inbuf_ids', [])
                inputs_list = []
                for bid in inbuf_ids:
                    if bid in buf_to_layer:
                        src = buf_to_layer[bid]
                        inputs_list.append({
                            'node_id': src['node_id'],
                            'name': src['name'],
                            'type': src['type'],
                            'tensor_name': src['tensor_name'],
                            'shape': src['shape']
                        })

                # Build outputs: find all layers that consume this layer's output buffer
                my_outbuf = lm.get('outbuf_id', layer_idx)
                my_tensor_name = lm.get('output_name', layer['layer_name'])
                my_shape = []
                for o in params.get('outputs', []):
                    if 'dims' in o:
                        my_shape = o['dims']
                        break
                outputs_list = []
                # Only emit outputs if this layer is the buffer owner (prevents reuse layers
                # like the output DataLayer from generating spurious output connections)
                is_buffer_owner = buf_to_layer.get(my_outbuf, {}).get('node_id') == layer_idx
                if is_buffer_owner:
                    for other_lyr in tidl_info.get('layers', []):
                        other_li = other_lyr['layer_index']
                        other_lm = netlog_lm.get(other_li, {})
                        if my_outbuf in other_lm.get('inbuf_ids', []):
                            outputs_list.append({
                                'node_id': other_li,
                                'name': other_lyr['layer_name'],
                                'type': other_lyr['layer_type'],
                                'tensor_name': my_tensor_name,
                                'shape': my_shape
                            })

                # Get fusion info for this TIDL layer
                fusion_info = fusion_map.get(layer_idx, None)
                onnx_node_index = layer.get('onnx_node_index', None)
                expansion_info = expansion_map.get(onnx_node_index, None) if onnx_node_index is not None else None

                # Build onnx_node_indices and onnx_node_names as arrays
                if fusion_info:
                    onnx_indices = fusion_info['fused_onnx_indices']
                    onnx_names = fusion_info['fused_onnx_names']
                elif onnx_node_index is not None:
                    onnx_indices = [onnx_node_index]
                    onnx_node_name = layer.get('onnx_node_name')
                    if not onnx_node_name and onnx_node_index in node_support:
                        onnx_node_name = node_support[onnx_node_index].get('node_name')
                    onnx_names = [onnx_node_name] if onnx_node_name else []
                else:
                    onnx_indices = []
                    onnx_names = []

                # Determine mapping_type for onnx_mapping
                if layer['layer_type'] in ('TIDL_DataLayer',):
                    mapping_type = 'data_input'
                elif fusion_info and len(onnx_indices) > 1:
                    mapping_type = 'fusion'
                elif len(onnx_indices) == 1:
                    mapping_type = '1-to-1'
                else:
                    mapping_type = 'none'

                # Determine performance — all layers get the field (null if not available)
                perf_value = perf_lookup.get(layer_idx, None)

                # Determine activation data and bin_files
                layer_type_str = layer.get('layer_type', '')
                act_key = f"{subgraph_id}_{layer_idx}"
                skip_activation = layer_type_str in ('TIDL_DataLayer', 'TIDL_DataConvertLayer')
                if not skip_activation and activation_data and act_key in activation_data:
                    act = activation_data[act_key]
                    act_data = {
                        'histogram': act.get('histogram', {'tidl_bins': [], 'tidl_counts': [], 'notidl_bins': [], 'notidl_counts': []}),
                        'scatter': act.get('scatter', {'x': [], 'y': [], 'sample_size': 0, 'total_points': 0}),
                        'metrics': act.get('metrics', None),
                        'bin_files': act.get('bin_files', None),
                    }
                else:
                    act_data = {
                        'histogram': {'tidl_bins': [], 'tidl_counts': [], 'notidl_bins': [], 'notidl_counts': []},
                        'scatter': {'x': [], 'y': [], 'sample_size': 0, 'total_points': 0},
                        'metrics': None,
                        'bin_files': None,
                    }

                enhanced_layer = {
                    'layer_id': layer_idx,
                    'layer_type': layer['layer_type'],
                    'layer_name': layer['layer_name'],
                    'onnx_mapping': {
                        'onnx_node_indices': onnx_indices,
                        'onnx_node_names': onnx_names,
                        'mapping_type': mapping_type,
                    },
                    'inputs': inputs_list,
                    'outputs': outputs_list,
                    'gmacs': layer.get('gmacs', 0.0),
                    'parameters': layer['parameters'],
                    'performance': perf_value,
                    'activation_data': act_data,
                }

                # Use layer_id as key (like ONNX layers)
                enhanced_layers[str(layer_idx)] = enhanced_layer


            # Convert enhanced_layers dict to array (sorted by layer_id)
            layers_array = []
            for layer_id in sorted(enhanced_layers.keys(), key=lambda x: int(x)):
                layers_array.append(enhanced_layers[layer_id])

            enhanced_tidl_data[subgraph_id] = {
                'subgraph_id': subgraph_id,
                'total_gmacs': tidl_info.get('total_gmacs', 0.0),
                'num_layers': len(layers_array),
                'total_time_us': 0.0,  # Will be calculated later
                'layers': layers_array  # Array format
            }

        # Extract ONNX model inputs/outputs
        from datetime import datetime

        onnx_inputs = []
        onnx_outputs = []
        model_details = model_data.get('model_details', {})

        # Extract inputs from 'input_shape' (not 'inputs')
        for inp in model_details.get('input_shape', []):
            onnx_inputs.append({
                'name': inp.get('name', ''),
                'shape': inp.get('shape', []),
                'dtype': 'float32'  # Default, could be extracted if available
            })

        # Extract outputs from 'output_shape' (not 'outputs')
        for out in model_details.get('output_shape', []):
            onnx_outputs.append({
                'name': out.get('name', ''),
                'shape': out.get('shape', []),
                'dtype': 'float32'  # Default, could be extracted if available
            })

        # Metadata with only essential fields (matching reference schema)
        metadata = {
            'model_name': model_details.get('name', 'Unknown'),
            'task_type': config_data.get('task_type', 'Unknown'),
            'inputs': onnx_inputs,
            'outputs': onnx_outputs
        }

        # Store target_device and tensor_bits for use in each subgraph
        target_device = config_data.get('target_device', 'Unknown')


        # Restructure ONNX layers to match unified schema
        node_support = subgraph_data.get('node_support', {})
        onnx_layers_raw = model_data.get('layer_details', {})
        onnx_layers = {}

        for layer_name, layer_info in onnx_layers_raw.items():
            # Filter inputs to only include non-constant tensors (actual data inputs)
            input_metadata = layer_info.get('input_metadata', [])
            data_inputs = [inp['tensor_name'] for inp in input_metadata if not inp.get('is_constant', False)]

            # Restructure to unified schema format (no inputs/outputs, only input_details/output_details)
            unified_layer = {
                'name': layer_info.get('layer_name', layer_name),
                'type': layer_info.get('type', 'Unknown'),
                'input_details': input_metadata,
                'output_details': layer_info.get('output_metadata', []),
                'attributes': layer_info.get('attributes', {})
            }

            # Add runtime_assignment information
            # Check if this ONNX node failed on any runtime
            # assigned_runtime: "tidl_rt", "tvm_rt", or "arm"
            # reason: failure reason(s) if failed, null if passed
            node_idx = layer_info.get('node_index', None)
            if node_idx is not None and node_idx in node_support:
                support_info = node_support[node_idx]
                is_supported = support_info.get('supported', False)

                if is_supported:
                    # Passed TIDL runtime
                    unified_layer['runtime_assignment'] = {
                        'assigned_runtime': 'tidl_rt',
                        'reason': None
                    }
                else:
                    # Failed on TIDL runtime - assume goes to ARM
                    unified_layer['runtime_assignment'] = {
                        'assigned_runtime': 'arm',
                        'reason': support_info.get('diagInfo', 'TIDL runtime not supported')
                    }
            else:
                # No support information - assume passed TIDL
                unified_layer['runtime_assignment'] = {
                    'assigned_runtime': 'tidl_rt',
                    'reason': None
                }

            onnx_layers[layer_name] = unified_layer

        # Override runtime_assignment for ONNX nodes that are fused into TIDL layers.
        # graphvizInfo.txt may label them as ARM (e.g. "will be delegated in post-processing")
        # but if they appear in onnx_mapping.onnx_node_names of a TIDL layer they actually
        # run on C7x DSP (e.g. TIDL_OdOutputReformatLayer absorbs all OD post-processing).
        for sg_id, tidl_subgraph in enhanced_tidl_data.items():
            sg_runtime = tidl_subgraph.get('runtime', 'tidl_rt')
            for layer in tidl_subgraph.get('layers', []):
                for onnx_name in layer.get('onnx_mapping', {}).get('onnx_node_names', []):
                    if onnx_name in onnx_layers:
                        # Node IS compiled into a TIDL layer — override graphvizInfo
                        # classification and clear misleading reason text.
                        onnx_layers[onnx_name]['runtime_assignment'] = {
                            'assigned_runtime': sg_runtime,
                            'reason': None  # clear "will be delegated in post-processing"
                        }

        # Calculate total_time_us for each TIDL subgraph
        for subgraph_id, tidl_subgraph in enhanced_tidl_data.items():
            total_time = 0.0
            # layers is now an array
            for layer in tidl_subgraph.get('layers', []):
                if layer.get('performance') and 'proctime_us' in layer['performance']:
                    total_time += layer['performance']['proctime_us']
            tidl_subgraph['total_time_us'] = total_time

        # Check for TVM artifacts (optional)
        tvm_data = {}
        tvm_artifacts_path = os.path.join(model_dir_path, 'tvm', 'artifacts')
        tvmrt_artifacts_path = os.path.join(model_dir_path, 'tvmrt_artifacts')

        if os.path.exists(tvm_artifacts_path) or os.path.exists(tvmrt_artifacts_path):
            print("  TVM artifacts detected, parsing TVM data...")
            # TODO: Parse TVM artifacts similar to TIDL
            # For now, create placeholder structure
            tvm_data = {
                'subgraphs': {}
            }
        else:
            print("  No TVM artifacts found, skipping TVM section")
            tvm_data = {}

        # Build subgraphs dict in the format: tidl_0, tidl_1, tvm_0, etc.
        unified_subgraphs = {}

        # Add TIDL subgraphs with correct field order
        for subgraph_id, tidl_subgraph in enhanced_tidl_data.items():
            subgraph_key = f'tidl_{subgraph_id}'

            # Build TIDL subgraph with correct field order matching unified schema
            ordered_tidl_subgraph = {
                'runtime': 'tidl_rt',
                'subgraph_id': tidl_subgraph['subgraph_id'],
                'tidl_tool_version': config_data.get('tidl_tool_version', '9.0.0'),
                'tensor_bits': config_data.get('tensor_bits', 8),
                'total_gmacs': tidl_subgraph['total_gmacs'],
                'num_layers': tidl_subgraph['num_layers'],
                'inputs': [],
                'outputs': [],
                'layers': tidl_subgraph['layers'],
                'target_device': target_device,
            }

            unified_subgraphs[subgraph_key] = ordered_tidl_subgraph

        # Add TVM subgraphs if present
        if tvm_data and 'subgraphs' in tvm_data:
            for subgraph_id, tvm_subgraph in tvm_data['subgraphs'].items():
                subgraph_key = f'tvm_{subgraph_id}'
                # Add target_device to TVM subgraph
                tvm_subgraph['target_device'] = target_device
                unified_subgraphs[subgraph_key] = tvm_subgraph

        # Unified JSON structure v1.0
        combined_data = {
            'metadata': metadata,
            'model': {
                'onnx': {
                    'total_weights': model_details.get('weights', 0),
                    'num_layers': len(onnx_layers),
                    'opset_version': onnx_parser.model.opset_import[0].version if hasattr(onnx_parser.model, 'opset_import') and len(onnx_parser.model.opset_import) > 0 else 0,
                    'ir_version': onnx_parser.model.ir_version if hasattr(onnx_parser.model, 'ir_version') else 0,
                    'layers': onnx_layers
                },
                'tflite': {}
            },
            'runtime': {
                'subgraphs': unified_subgraphs
            }
        }

        print(f"Writing unified JSON to: {output_json_path}")

        class _JSONSafeEncoder(json.JSONEncoder):
            """Fallback encoder: converts any non-serializable object to its str()."""
            def default(self, obj):
                try:
                    return super().default(obj)
                except TypeError:
                    return str(obj)

        # Write single unified JSON for inspection (activation data embedded in TIDL layers)
        print("  Serializing unified data...")
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(combined_data, f, indent=2, cls=_JSONSafeEncoder)

        file_size = os.path.getsize(output_json_path) / (1024 * 1024)
        print(f"Unified JSON saved: {output_json_path} ({file_size:.2f} MB)")
        print(f"  Model structure only (no activation data)")

        print("\n" + "=" * 70)
        print("SUCCESS! Data extraction complete - Unified Schema v1.0")
        print("=" * 70)
        print(f"\nExtracted data summary:")
        print(f"  - ONNX layers: {len(combined_data['model']['onnx']['layers'])}")
        tidl_subgraphs = combined_data.get('runtime', {}).get('subgraphs', {})
        print(f"  - TIDL subgraphs: {len(tidl_subgraphs)}")
        if tvm_data and 'subgraphs' in tvm_data:
            print(f"  - TVM subgraphs: {len(tvm_data['subgraphs'])}")

        # Activation data not included in JSON - only model structure

        # Note: metrics and activation_data are set to null
        # They will be populated by html_generator based on --activation_data flag
        print(f"  - Metrics: null (to be filled by html_generator)")
        print(f"  - Activation data: null (to be filled by html_generator)")

        total_perf_layers = 0
        for subgraph_id, subgraph in combined_data['runtime']['subgraphs'].items():
            # layers is now a list, iterate through it
            layers = subgraph.get('layers', [])
            if isinstance(layers, list):
                total_perf_layers += sum(1 for layer in layers if isinstance(layer, dict) and 'performance' in layer)
            elif isinstance(layers, dict):
                total_perf_layers += sum(1 for layer in layers.values() if 'performance' in layer)
        print(f"  - Layers with performance data: {total_perf_layers}")

        print(f"\nMetadata:")
        print(f"  - Model: {combined_data['metadata']['model_name']}")
        # Get device from first subgraph (moved from metadata to subgraph level)
        first_sg = next(iter(combined_data['runtime']['subgraphs'].values()), {})
        target_device = first_sg.get('target_device', 'Unknown')
        tensor_bits = first_sg.get('tensor_bits', 'Unknown')
        print(f"  - Device: {target_device}")
        print(f"  - Precision: {tensor_bits}-bit")
        print(f"  - Task: {combined_data['metadata']['task_type']}")
        print(f"  - Inputs: {len(combined_data['metadata']['inputs'])}")
        print(f"  - Outputs: {len(combined_data['metadata']['outputs'])}")

        print(f"\nNext step:")
        print(f"  python html_generator.py {output_json_path} template.html output.html")

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def update_with_activations(work_dirs_path, json_path):
    """Patch an existing inspector JSON with activation/trace data from an analyze run.

    Called after the analyze pipeline steps have generated tidl/notidl trace files.
    Only the activation_data field of each TIDL layer is updated; everything else
    (model structure, performance, etc.) is left unchanged.

    Returns True if at least one layer was updated, False otherwise.
    """
    if not os.path.exists(json_path):
        print(f'INFO: JSON not found, cannot update activations: {json_path}')
        return False

    print(f'INFO: Loading inspector JSON for activation update: {json_path}')
    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)

    try:
        activation_parser = ActivationDataParser(model_dir=work_dirs_path, frame_idx=0)
        activation_data = activation_parser.process_all_layers()
    except Exception as e:
        print(f'WARNING: Could not parse activation data: {e}')
        return False

    if not activation_data:
        print('INFO: No activation data found, JSON not updated')
        return False

    skip_types = {'TIDL_DataLayer', 'TIDL_DataConvertLayer'}
    updated_count = 0

    for sg_key, sg_data in data.get('runtime', {}).get('subgraphs', {}).items():
        if not sg_key.startswith('tidl_'):
            continue
        try:
            sg_id = int(sg_key[len('tidl_'):])
        except ValueError:
            continue

        for layer in sg_data.get('layers', []):
            layer_id = layer.get('layer_id')
            if layer_id is None or layer.get('layer_type', '') in skip_types:
                continue
            act_key = f'{sg_id}_{layer_id}'
            if act_key in activation_data:
                act = activation_data[act_key]
                layer['activation_data'] = {
                    'histogram': act.get('histogram', {'tidl_bins': [], 'tidl_counts': [], 'notidl_bins': [], 'notidl_counts': []}),
                    'scatter': act.get('scatter', {'x': [], 'y': [], 'sample_size': 0, 'total_points': 0}),
                    'metrics': act.get('metrics', None),
                    'bin_files': act.get('bin_files', None),
                }
                updated_count += 1

    print(f'INFO: Updated {updated_count} layers with activation data')

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)

    print(f'INFO: Saved updated JSON to: {json_path}')
    return updated_count > 0


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Extract TIDL Artifacts to JSON')
    parser.add_argument('work_dirs', help='Direct path to model directory')
    parser.add_argument('output_json', help='Output JSON file path')
    parser.add_argument('--act_data', action='store_true', help='Extract activation data to separate file')

    args = parser.parse_args()

    main(work_dirs_path=args.work_dirs, output_json_path=args.output_json, extract_activations=args.act_data)