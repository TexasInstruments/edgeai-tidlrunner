#!/usr/bin/env python3
# Copyright (c) 2018-2025, Texas Instruments
# All Rights Reserved.

"""
Data Extractor - Extract TIDL artifact data to JSON

This script discovers and parses all TIDL compilation artifacts from work_dirs/
and outputs a single compressed JSON file containing all extracted data.

Usage:
    python data_extractor.py <work_dirs/> <output.json>

Example:
    python data_extractor.py work_dirs/ model_data.json

Output:
    - Single JSON file with all parsed data (model, subgraphs, TIDL, activation, metrics, config, performance)
"""

import onnx
import json
import sys
import os
import re
import gzip
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional

# Python 3.10+ compatibility: collections.Callable moved to collections.abc.Callable
import collections
import collections.abc
if not hasattr(collections, 'Callable'):
    collections.Callable = collections.abc.Callable

from bs4 import BeautifulSoup

try:
    import onnx_graphsurgeon as gs
    HAS_GRAPHSURGEON = True
except ImportError:
    HAS_GRAPHSURGEON = False
    print("WARNING: onnx-graphsurgeon not installed. Using raw ONNX API (slower).")
    print("  Install with: pip install onnx-graphsurgeon")


class ActivationDataParser:
    """Parser for layer activation binary files with histogram and scatter plot generation

    Uses layer_info.txt files directly to build activation mappings
    """

    def __init__(self, model_dir: str, frame_idx: int = 0):
        """
        Initialize activation data parser using layer_info.txt files

        Args:
            model_dir: Direct path to model directory (e.g., work_dirs/compile/AM69A/cl_onnx_model/)
            frame_idx: Frame index to use for activation data (default: 0)
        """
        self.model_dir = Path(model_dir)
        self.frame_idx = frame_idx
        self.mapping = {}
        self.data_cache = {}

        self._load_from_layer_info()

    def _load_from_layer_info(self):
        """Build activation mapping from layer_info.txt files"""
        import glob

        print(f"Building activation mapping from layer_info.txt files...")

        layer_info_pattern = os.path.join(self.model_dir, 'tidl/artifacts/tempDir/subgraph_*_tidl_net.bin.layer_info.txt')
        layer_info_files = glob.glob(layer_info_pattern, recursive=False)

        if not layer_info_files:
            print(f"  WARNING: No layer_info.txt files found in {layer_info_pattern}")
            return

        print(f"  Found {len(layer_info_files)} layer_info.txt files")

        notidl_outputs_dir = self._find_notidl_outputs()
        tidl_traces_dir = self._find_tidl_traces()

        if not notidl_outputs_dir or not tidl_traces_dir:
            print(f"  WARNING: Could not find output directories")
            return

        total_layers = 0
        total_mapped = 0

        for layer_info_path in sorted(layer_info_files):
            match = re.search(r'subgraph_(\d+)_tidl_net\.bin\.layer_info\.txt', layer_info_path)
            if not match:
                continue

            subgraph_id = int(match.group(1))

            if subgraph_id not in self.mapping:
                self.mapping[subgraph_id] = {}

            with open(layer_info_path, 'r') as f:
                for line in f:
                    parts = line.strip().split()
                    if len(parts) < 3:
                        continue

                    tidl_id_col0 = parts[0]
                    tidl_id_col1 = parts[1]
                    onnx_layer_name = parts[2]

                    if tidl_id_col0 != tidl_id_col1:
                        continue

                    tidl_layer_id = tidl_id_col0
                    total_layers += 1

                    notidl_path = self._build_notidl_path(notidl_outputs_dir, onnx_layer_name)
                    tidl_path = self._build_tidl_path(tidl_traces_dir, subgraph_id, tidl_layer_id)

                    if notidl_path and tidl_path and os.path.exists(notidl_path) and os.path.exists(tidl_path):
                        self.mapping[subgraph_id][tidl_layer_id] = {
                            'notidl': notidl_path,
                            'tidl': tidl_path,
                            'onnx_name': onnx_layer_name
                        }
                        total_mapped += 1

        print(f"  Total layers in layer_info.txt: {total_layers}")
        print(f"  Layers with activation files: {total_mapped}")

        for sg_id in sorted(self.mapping.keys()):
            print(f"    Subgraph {sg_id}: {len(self.mapping[sg_id])} layers")

    def _find_notidl_outputs(self) -> Optional[str]:
        """Find NotIDL outputs directory"""
        import glob
        pattern = os.path.join(self.model_dir, f'notidl/outputs_/{self.frame_idx}')
        matches = glob.glob(pattern, recursive=False)
        return matches[0] if matches else None

    def _find_tidl_traces(self) -> Optional[str]:
        """Find TIDL traces directory"""
        import glob
        pattern = os.path.join(self.model_dir, f'tidl/traces_/{self.frame_idx}')
        matches = glob.glob(pattern, recursive=False)
        return matches[0] if matches else None

    def _build_notidl_path(self, notidl_dir: str, onnx_layer_name: str) -> Optional[str]:
        """Build path to NotIDL output file"""
        onnx_layer_id = onnx_layer_name.replace("/", "_")
        notidl_path = os.path.join(notidl_dir, f"{onnx_layer_id}.bin")

        if os.path.exists(notidl_path):
            return notidl_path

        for filename in os.listdir(notidl_dir):
            if onnx_layer_id in filename and filename.endswith('.bin'):
                return os.path.join(notidl_dir, filename)

        return None

    def _build_tidl_path(self, tidl_dir: str, subgraph_id: int, tidl_layer_id: str) -> Optional[str]:
        """Build path to TIDL trace file"""
        import glob

        tidl_id_padded = tidl_layer_id.zfill(4)
        pattern = os.path.join(tidl_dir, f"tidl_trace_subgraph_{subgraph_id}_{tidl_id_padded}_*_float.bin")
        matches = glob.glob(pattern)

        return matches[0] if matches else None

    def _load_bin_file(self, bin_path: str) -> Optional[np.ndarray]:
        """Load binary file with caching"""
        if bin_path in self.data_cache:
            return self.data_cache[bin_path]

        try:
            if not os.path.exists(bin_path):
                print(f"    Warning: File not found: {bin_path}")
                return None

            data = np.fromfile(bin_path, dtype=float)
            self.data_cache[bin_path] = data
            return data

        except Exception as e:
            print(f"    Error loading {bin_path}: {e}")
            return None

    def _sample_data(self, data: np.ndarray, max_samples: int = 50000) -> np.ndarray:
        """Sample data for visualization"""
        if len(data) <= max_samples:
            return data

        np.random.seed(42)
        indices = np.random.choice(len(data), max_samples, replace=False)
        return data[indices]

    def _sanitize_float(self, value: float) -> float:
        """Convert Infinity/NaN to None for JSON compatibility"""
        if np.isnan(value) or np.isinf(value):
            return None
        return value

    def _calculate_statistics(self, notidl_data: np.ndarray, tidl_data: np.ndarray) -> Dict[str, float]:
        """Calculate accuracy statistics on full data"""
        try:
            notidl_f64 = notidl_data.astype(np.float64)
            tidl_f64 = tidl_data.astype(np.float64)
            diff = notidl_f64 - tidl_f64
            mae = float(np.mean(np.abs(diff)))
            rmse = float(np.sqrt(np.mean(diff ** 2)))
            max_error = float(np.max(np.abs(diff)))

            try:
                correlation = float(np.corrcoef(notidl_f64, tidl_f64)[0, 1])
            except:
                correlation = None

            notidl_min = float(np.min(notidl_data))
            notidl_max = float(np.max(notidl_data))
            notidl_mean = float(np.mean(notidl_data))
            notidl_std = float(np.std(notidl_data))

            tidl_min = float(np.min(tidl_data))
            tidl_max = float(np.max(tidl_data))
            tidl_mean = float(np.mean(tidl_data))
            tidl_std = float(np.std(tidl_data))

            return {
                'mae': self._sanitize_float(mae),
                'rmse': self._sanitize_float(rmse),
                'max_error': self._sanitize_float(max_error),
                'correlation': self._sanitize_float(correlation) if correlation is not None else None,
                'r_squared': self._sanitize_float(correlation ** 2) if correlation is not None else None,
                'total_points': len(notidl_data),
                'notidl_min': self._sanitize_float(notidl_min),
                'notidl_max': self._sanitize_float(notidl_max),
                'notidl_mean': self._sanitize_float(notidl_mean),
                'notidl_std': self._sanitize_float(notidl_std),
                'tidl_min': self._sanitize_float(tidl_min),
                'tidl_max': self._sanitize_float(tidl_max),
                'tidl_mean': self._sanitize_float(tidl_mean),
                'tidl_std': self._sanitize_float(tidl_std)
            }

        except Exception as e:
            print(f"    Error calculating statistics: {e}")
            return {}

    def _generate_histogram_json(self, notidl_data: np.ndarray, tidl_data: np.ndarray,
                              stats: Dict[str, float]) -> Dict[str, Any]:
        """Generate histogram data for visualization"""

        try:
            notidl_data = np.asarray(notidl_data).flatten()
            tidl_data = np.asarray(tidl_data).flatten()

            notidl_counts, notidl_edges = np.histogram(notidl_data, bins=100)
            tidl_counts, tidl_edges = np.histogram(tidl_data, bins=100)

            notidl_centers = (notidl_edges[:-1] + notidl_edges[1:]) / 2
            tidl_centers = (tidl_edges[:-1] + tidl_edges[1:]) / 2

            notidl_centers = np.round(notidl_centers, 4)
            tidl_centers = np.round(tidl_centers, 4)

            notidl_centers_list = notidl_centers.tolist()
            notidl_counts_list = notidl_counts.tolist()
            tidl_centers_list = tidl_centers.tolist()
            tidl_counts_list = tidl_counts.tolist()

        except Exception as e:
            print(f"    Error generating histogram bins: {e}")
            import traceback
            traceback.print_exc()
            return {'traces': [], 'layout': {}}

        traces = [
            {
                'type': 'bar',
                'x': notidl_centers_list,
                'y': notidl_counts_list,
                'name': 'No-TIDL (FP32)',
                'marker': {
                    'color': 'rgba(255, 127, 0, 0.75)',
                    'line': {'color': 'rgba(255, 127, 0, 1)', 'width': 0.5}
                },
                'opacity': 0.75,
                'hovertemplate': '<b>No-TIDL</b><br>Value: %{x:.4f}<br>Count: %{y:,}<extra></extra>'
            },
            {
                'type': 'bar',
                'x': tidl_centers_list,
                'y': tidl_counts_list,
                'name': 'TIDL (INT8)',
                'marker': {
                    'color': 'rgba(255, 152, 0, 0.7)',
                    'line': {'color': 'rgba(255, 152, 0, 1)', 'width': 0.5}
                },
                'opacity': 0.7,
                'hovertemplate': '<b>TIDL</b><br>Value: %{x:.4f}<br>Count: %{y:,}<extra></extra>'
            }
        ]

        layout = {
            'title': {
                'text': 'Activation Distribution Comparison',
                'font': {'size': 14, 'color': '#333'}
            },
            'xaxis': {
                'title': 'Activation Value',
                'showgrid': False,
                'showline': True,
                'linewidth': 1,
                'linecolor': '#e0e0e0'
            },
            'yaxis': {
                'title': 'Frequency',
                'type': 'log',
                'showgrid': True,
                'gridcolor': 'rgba(255, 255, 255, 0.15)'
            },
            'barmode': 'overlay',
            'template': 'plotly_white',
            'showlegend': True,
            'legend': {
                'orientation': 'h',
                'yanchor': 'bottom',
                'y': 1.02,
                'xanchor': 'left',
                'x': 0
            },
            'hovermode': 'x unified',
            'margin': {'l': 50, 'r': 20, 't': 60, 'b': 50},
            'annotations': [
                {
                    'text': f"MAE: {stats.get('mae') or 0:.4f} | " +
                            f"No-TIDL: [{stats.get('notidl_min') or 0:.2f}, {stats.get('notidl_max') or 0:.2f}] | " +
                            f"TIDL: [{stats.get('tidl_min') or 0:.2f}, {stats.get('tidl_max') or 0:.2f}]",
                    'xref': 'paper',
                    'yref': 'paper',
                    'x': 0.5,
                    'y': -0.15,
                    'showarrow': False,
                    'xanchor': 'center',
                    'font': {'size': 10, 'color': '#666'}
                }
            ]
        }

        return {'traces': traces, 'layout': layout}

    def _generate_scatter_plot_d3(self, notidl_data: np.ndarray, tidl_data: np.ndarray,
                                   stats: Dict[str, float]) -> Dict[str, Any]:
        """Generate scatter plot data for D3 visualization"""

        try:
            notidl_flat = np.asarray(notidl_data).flatten()
            tidl_flat = np.asarray(tidl_data).flatten()

            total_points = len(notidl_flat)

            notidl_rounded = np.round(notidl_flat, 4)
            tidl_rounded = np.round(tidl_flat, 4)

            all_values = np.concatenate([notidl_rounded, tidl_rounded])
            min_val = float(np.min(all_values))
            max_val = float(np.max(all_values))

            padding = (max_val - min_val) * 0.05
            axis_min = min_val - padding
            axis_max = max_val + padding

            scatter_data = {
                'points': [
                    {'x': float(tidl_rounded[i]), 'y': float(notidl_rounded[i])}
                    for i in range(len(tidl_rounded))
                ],
                'diagonal': {
                    'start': {'x': axis_min, 'y': axis_min},
                    'end': {'x': axis_max, 'y': axis_max}
                },
                'axis': {
                    'min': axis_min,
                    'max': axis_max
                },
                'stats': {
                    'r_squared': stats.get('r_squared'),
                    'mae': stats.get('mae'),
                    'rmse': stats.get('rmse'),
                    'correlation': stats.get('correlation'),
                    'total_points': total_points
                }
            }

            return scatter_data

        except Exception as e:
            print(f"    Error generating D3 scatter plot: {e}")
            import traceback
            traceback.print_exc()
            return {'points': [], 'stats': {}}

    def process_layer(self, subgraph_id: int, tidl_layer_id: str) -> Optional[Dict[str, Any]]:
        """Process a single layer and generate plot data

        Args:
            subgraph_id: Subgraph ID
            tidl_layer_id: TIDL layer ID within subgraph

        Returns:
            Dict with stats, histogram, and scatter plot data, or None if processing failed
        """

        if subgraph_id not in self.mapping:
            return None

        if tidl_layer_id not in self.mapping[subgraph_id]:
            return None

        layer_data = self.mapping[subgraph_id][tidl_layer_id]

        notidl_path = layer_data['notidl']
        tidl_path = layer_data['tidl']

        notidl_data = self._load_bin_file(notidl_path)
        tidl_data = self._load_bin_file(tidl_path)

        if notidl_data is None or tidl_data is None:
            return None

        if len(notidl_data) != len(tidl_data):
            print(f"  Warning: Size mismatch for subgraph {subgraph_id} layer {tidl_layer_id} (NotIDL: {len(notidl_data)}, TIDL: {len(tidl_data)})")
            return None

        try:
            stats = self._calculate_statistics(notidl_data, tidl_data)
            histogram_data = self._generate_histogram_json(notidl_data, tidl_data, stats)
            scatter_data = self._generate_scatter_plot_d3(notidl_data, tidl_data, stats)

        except Exception as e:
            print(f"  Error processing layer {tidl_layer_id}: {e}")
            import traceback
            traceback.print_exc()
            return None

        del notidl_data, tidl_data

        return {
            'stats': stats,
            'histogram': histogram_data,
            'scatter': scatter_data
        }

    def process_all_layers(self, subgraph_node_mapping: Dict[int, List[int]] = None) -> Dict[str, Any]:
        """
        Process all layers and generate activation data

        Args:
            subgraph_node_mapping: NOT USED - kept for backward compatibility

        Returns:
            Dict with keys "subgraphId_tidlLayerId" -> plot data
        """
        if not self.mapping:
            print("  No activation mapping available, skipping activation analysis")
            return {}

        print("\nProcessing activation data...")

        total_layers = sum(len(layers) for layers in self.mapping.values())
        print(f"  Found {total_layers} TIDL layers with activation files")

        activation_data = {}
        total_processed = 0
        total_failed = 0
        subgraph_stats = {}

        for subgraph_id in sorted(self.mapping.keys()):
            layer_count = len(self.mapping[subgraph_id])
            print(f"\n  Subgraph {subgraph_id}: Processing {layer_count} layers")

            for tidl_layer_id in sorted(self.mapping[subgraph_id].keys()):
                result = self.process_layer(subgraph_id, tidl_layer_id)

                if result:
                    key = f"{subgraph_id}_{tidl_layer_id}"
                    activation_data[key] = result
                    total_processed += 1

                    subgraph_stats[subgraph_id] = subgraph_stats.get(subgraph_id, 0) + 1

                    if total_processed <= 10 or total_processed % 10 == 0:
                        layer_info = self.mapping[subgraph_id][tidl_layer_id]
                        onnx_name = layer_info.get('onnx_name', 'unknown')
                        mae = result.get('stats', {}).get('mae', 0)
                        print(f"    Layer {tidl_layer_id} ({onnx_name}): MAE={mae:.6f}")
                else:
                    total_failed += 1

        print(f"\nActivation processing complete:")
        for sg_id in sorted(subgraph_stats.keys()):
            print(f"  Subgraph {sg_id}: {subgraph_stats[sg_id]} layers processed")
        print(f"  Total Processed: {total_processed}")
        print(f"  Total Failed: {total_failed}")
        print(f"  Data size: {len(json.dumps(activation_data)) / (1024*1024):.2f} MB")

        return activation_data


class MetricsParser:
    """Parser for analyze.xlsx metrics file"""

    def __init__(self, xlsx_path: str):
        """
        Initialize metrics parser

        Args:
            xlsx_path: Path to analyze.xlsx file
        """
        self.xlsx_path = Path(xlsx_path)
        self.metrics_data = {}

        if self.xlsx_path.exists():
            self._load_metrics()
            print(f"Loaded metrics from: {xlsx_path}")
        else:
            print(f"WARNING: Metrics file not found: {xlsx_path}")

    def _load_metrics(self):
        """Load metrics from Excel file"""
        try:
            import openpyxl
        except ImportError:
            print("  Error: openpyxl not installed. Run: pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(self.xlsx_path, read_only=True)

            for sheet_name in wb.sheetnames:
                if sheet_name.startswith('diff_notidl_tidl_'):
                    subgraph_id = sheet_name.split('_')[-1]
                    sheet = wb[sheet_name]

                    headers = [cell.value for cell in sheet[1]]

                    subgraph_metrics = []
                    for row_idx in range(2, sheet.max_row + 1):
                        row = [cell.value for cell in sheet[row_idx]]

                        if len(row) >= 12 and row[1] is not None:
                            metric_entry = {
                                'subgraph': row[0],
                                'serial_num': row[1],
                                'onnx_layer': row[2],
                                'tidl_layer_id': row[3],
                                'mean_abs_rel_diff': float(row[4]) if row[4] is not None else 0.0,
                                'mean_abs_diff': float(row[5]) if row[5] is not None else 0.0,
                                'median_abs_diff': float(row[6]) if row[6] is not None else 0.0,
                                'max_abs_diff': float(row[7]) if row[7] is not None else 0.0,
                                'mean_abs_diff_median': float(row[9]) if row[9] is not None else 0.0,
                                'median_abs_diff_median': float(row[10]) if row[10] is not None else 0.0,
                                'max_abs_diff_median': float(row[11]) if row[11] is not None else 0.0
                            }
                            subgraph_metrics.append(metric_entry)

                    self.metrics_data[subgraph_id] = subgraph_metrics
                    print(f"  Loaded {len(subgraph_metrics)} metrics for subgraph {subgraph_id}")

            wb.close()

        except Exception as e:
            print(f"  Error loading metrics: {e}")
            self.metrics_data = {}

    def get_metrics(self) -> Dict[str, List[Dict[str, Any]]]:
        """Get all metrics data organized by subgraph"""
        return self.metrics_data


class TIDLNetLogParser:
    """Parser for subgraph_X_tidl_net.bin_netLog.txt files"""

    def __init__(self, filepath: str):
        self.filepath = filepath

    def parse(self) -> Dict[str, Any]:
        """Parse netLog.txt file and extract MACS information"""
        print(f"  Parsing netLog: {os.path.basename(self.filepath)}")

        layer_macs = {}
        total_gmacs = 0.0

        try:
            with open(self.filepath, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            in_table = False

            for line in lines:
                line_stripped = line.strip()

                if '----' in line_stripped and len(line_stripped) > 50:
                    in_table = True
                    continue

                if in_table and '----' in line_stripped and len(line_stripped) > 50:
                    in_table = False
                    continue

                if in_table and line_stripped and not line_stripped.startswith('Num|'):
                    parts = [p.strip() for p in line.split('|')]

                    if len(parts) >= 11:
                        try:
                            layer_num = int(parts[0])
                            layer_name = parts[1]
                            output_name = parts[2]
                            macs_str = parts[-1]

                            macs = int(macs_str) if macs_str.isdigit() else 0
                            gmacs = macs / 1_000_000_000.0

                            layer_macs[layer_num] = {
                                'layer_name': layer_name,
                                'output_name': output_name,
                                'macs': macs,
                                'gmacs': gmacs
                            }

                        except (ValueError, IndexError) as e:
                            continue

                if 'Total Giga Macs' in line:
                    match = re.search(r'Total Giga Macs\s*:\s*([\d.]+)', line)
                    if match:
                        total_gmacs = float(match.group(1))

            print(f"    Found MACS info for {len(layer_macs)} layers, Total: {total_gmacs} GMACS")

            return {
                'layer_macs': layer_macs,
                'total_gmacs': total_gmacs
            }

        except Exception as e:
            print(f"    Warning: Failed to parse netLog file: {e}")
            return {
                'layer_macs': {},
                'total_gmacs': 0.0
            }


class TIDLSubgraphParser:
    """Parser for subgraph_X_tidl_net.bin.html files"""

    def __init__(self, base_dir: str, node_support: Dict[int, Dict[str, Any]] = None):
        self.base_dir = base_dir
        self.node_support = node_support or {}

    def _map_tidl_to_onnx_node(self, tidl_output_name: str) -> Optional[int]:
        """
        Map TIDL layer output tensor name to ONNX node index

        Args:
            tidl_output_name: Output tensor name from TIDL layer

        Returns:
            ONNX node index, or None if not found
        """
        if not self.node_support:
            return None

        onnx_node_name = tidl_output_name
        if onnx_node_name.endswith('_output_0'):
            onnx_node_name = onnx_node_name[:-9]
        if onnx_node_name.endswith('_netFormat'):
            onnx_node_name = onnx_node_name[:-10]

        for node_idx, node_data in self.node_support.items():
            if node_data.get('node_name') == onnx_node_name:
                return node_idx

        return None

    def find_subgraph_files(self) -> List[Tuple[int, str, str]]:
        """Find all subgraph HTML files and their corresponding netLog files"""
        pattern = re.compile(r'subgraph_(\d+)_tidl_net\.bin\.html')
        files = []

        for filename in os.listdir(self.base_dir):
            match = pattern.match(filename)
            if match:
                subgraph_idx = int(match.group(1))
                html_filepath = os.path.join(self.base_dir, filename)

                netlog_filename = f'subgraph_{subgraph_idx}_tidl_net.bin_netLog.txt'
                netlog_filepath = os.path.join(self.base_dir, netlog_filename)

                if not os.path.exists(netlog_filepath):
                    netlog_filepath = None

                files.append((subgraph_idx, html_filepath, netlog_filepath))

        files.sort(key=lambda x: x[0])
        return files

    def _parse_output_params(self, text: str, output_dict: Dict[str, Any]):
        """Helper to parse output parameters from text"""
        if not text or not text.strip():
            return

        for param_match in re.finditer(r'(\w+(?:/\w+)?)=([^\s]+)', text):
            param_key = param_match.group(1)
            param_val = param_match.group(2)

            if param_val.isdigit():
                output_dict[param_key] = int(param_val)
            elif param_val.replace('.', '', 1).replace('-', '', 1).isdigit():
                output_dict[param_key] = float(param_val)
            else:
                output_dict[param_key] = param_val

        for array_match in re.finditer(r'(\w+(?:/\w+)?)=\[([^\]]+)\]', text):
            array_key = array_match.group(1)
            array_val = array_match.group(2)

            try:
                output_dict[array_key] = [int(x.strip()) for x in array_val.split(',')]
            except ValueError:
                output_dict[array_key] = array_val

        for paren_match in re.finditer(r'(\w+(?:/\w+)?)\(([^)]+)\)', text):
            paren_key = paren_match.group(1)
            paren_val = paren_match.group(2)
            output_dict[paren_key] = f"({paren_val})"

    def parse_layer_info(self, title_text: str) -> Dict[str, Any]:
        """Parse TIDL layer information from SVG node title"""
        layer_info = {
            'raw_text': title_text,
            'layer_index': None,
            'layer_type': None,
            'layer_name': None,
            'parameters': {}
        }

        lines = title_text.strip().split('\n')
        if not lines:
            return layer_info

        first_line = lines[0].strip()
        layer_match = re.match(r'Layer\s+(\d+):\s+(\w+)\s+"([^"]+)"', first_line)
        if layer_match:
            layer_info['layer_index'] = int(layer_match.group(1))
            layer_info['layer_type'] = layer_match.group(2)
            layer_info['layer_name'] = layer_match.group(3)

        current_section = None
        current_output = None

        for line in lines[1:]:
            line = line.strip()
            if not line:
                continue

            if line.startswith('actParams:'):
                current_section = 'actParams'
                layer_info['parameters']['actParams'] = {}
                remainder = line[len('actParams:'):].strip()
                if remainder:
                    for param_match in re.finditer(r'(\w+(?:/\w+)?)[:=]([^\s]+)', remainder):
                        param_key = param_match.group(1)
                        param_val = param_match.group(2)
                        layer_info['parameters']['actParams'][param_key] = param_val
                continue
            elif line.startswith('Outputs:'):
                current_section = 'outputs'
                if 'outputs' not in layer_info['parameters']:
                    layer_info['parameters']['outputs'] = []

                remainder = line[len('Outputs:'):].strip()
                if remainder:
                    output_match = re.match(r'\[(\d+)\]', remainder)
                    if output_match:
                        current_output = {'index': int(output_match.group(1))}
                        layer_info['parameters']['outputs'].append(current_output)
                        remainder = remainder[output_match.end():].strip()
                    elif not layer_info['parameters']['outputs']:
                        current_output = {'index': 0}
                        layer_info['parameters']['outputs'].append(current_output)
                    else:
                        current_output = layer_info['parameters']['outputs'][-1]

                    if remainder:
                        self._parse_output_params(remainder, current_output)
                continue

            if current_section == 'actParams':
                for param_match in re.finditer(r'(\w+(?:/\w+)?)[:=]([^\s]+)', line):
                    param_key = param_match.group(1)
                    param_val = param_match.group(2)
                    layer_info['parameters']['actParams'][param_key] = param_val

            elif current_section == 'outputs':
                if current_output is None:
                    if not layer_info['parameters']['outputs']:
                        current_output = {'index': 0}
                        layer_info['parameters']['outputs'].append(current_output)
                    else:
                        current_output = layer_info['parameters']['outputs'][-1]

                output_match = re.match(r'\[(\d+)\]', line)
                if output_match:
                    current_output = {'index': int(output_match.group(1))}
                    layer_info['parameters']['outputs'].append(current_output)
                    remainder = line[output_match.end():].strip()
                    if remainder:
                        self._parse_output_params(remainder, current_output)
                else:
                    self._parse_output_params(line, current_output)

            else:
                self._parse_output_params(line, layer_info['parameters'])

        params = layer_info['parameters']
        if 'kernelH/W' in params:
            layer_info['kernelShape'] = params['kernelH/W']

        if 'strideH/W' in params:
            layer_info['strides'] = params['strideH/W']

        if 'padH/W' in params:
            layer_info['pads'] = params['padH/W']

        if 'dilationH/W' in params:
            layer_info['dilations'] = params['dilationH/W']

        if 'numGroups' in params:
            layer_info['groups'] = params['numGroups']

        if 'numInChannels' in params:
            layer_info['numInChannels'] = params['numInChannels']
        if 'numOutChannels' in params:
            layer_info['numOutChannels'] = params['numOutChannels']

        return layer_info

    def _extract_graph_structure(self, soup, layers) -> Tuple[List[Dict], List[Dict]]:
        """Extract graph structure (nodes + edges) from SVG for visualization"""
        graph_nodes = []
        graph_edges = []

        layer_map = {layer['layer_index']: layer for layer in layers}

        svg_nodes = soup.find_all('g', class_='node')

        for node_elem in svg_nodes:
            title = node_elem.find('title')
            if not title or not title.string:
                continue

            first_line = title.string.strip().split('\n')[0]
            layer_match = re.match(r'Layer\s+(\d+):', first_line)
            if not layer_match:
                continue

            layer_idx = int(layer_match.group(1))
            layer_info = layer_map.get(layer_idx)

            if not layer_info:
                continue

            output_dims = []
            if 'parameters' in layer_info and 'outputs' in layer_info['parameters']:
                for output in layer_info['parameters']['outputs']:
                    if 'dims' in output:
                        output_dims.append(output['dims'])

            output_shape_str = 'N/A'
            if output_dims:
                formatted_shapes = [f"[{','.join(map(str, dims))}]" for dims in output_dims]
                output_shape_str = ', '.join(formatted_shapes)

            kernel_shape_str = 'N/A'
            if 'kernelShape' in layer_info and layer_info['kernelShape']:
                kernel_shape_str = str(layer_info['kernelShape'])

            strides_str = 'N/A'
            if 'strides' in layer_info and layer_info['strides']:
                strides_str = str(layer_info['strides'])

            dilations_str = 'N/A'
            if 'dilations' in layer_info and layer_info['dilations']:
                dilations_str = str(layer_info['dilations'])

            pads_str = 'N/A'
            if 'pads' in layer_info and layer_info['pads']:
                pads_str = str(layer_info['pads'])

            groups_str = 'N/A'
            if 'groups' in layer_info:
                groups_str = str(layer_info['groups'])

            activation_str = 'N/A'
            if 'parameters' in layer_info and 'actParams' in layer_info['parameters']:
                act_type = layer_info['parameters']['actParams'].get('actType')
                if act_type:
                    activation_str = act_type

            input_shape_str = 'N/A'
            if 'numInChannels' in layer_info:
                input_shape_str = f"Channels: {layer_info['numInChannels']}"

            layerdetails = {
                'name': layer_info['layer_name'] or f"Layer_{layer_idx}",
                'type': layer_info['layer_type'] or 'Unknown',
                'inputshape': input_shape_str,
                'outputshape': output_shape_str,
                'kernelshape': kernel_shape_str,
                'strides': strides_str,
                'dilations': dilations_str,
                'pads': pads_str,
                'groups': groups_str,
                'activation': activation_str,
                'auto_pad': 'N/A',
                'numInChannels': layer_info.get('numInChannels', 'N/A'),
                'numOutChannels': layer_info.get('numOutChannels', 'N/A')
            }

            if 'parameters' in layer_info:
                params = layer_info['parameters']
                if 'weightsElementSizeInBits' in params:
                    layerdetails['weightsElementSizeInBits'] = params['weightsElementSizeInBits']
                if 'numChannels' in params:
                    layerdetails['numChannels'] = params['numChannels']
                if 'dataQ' in params:
                    layerdetails['dataQ'] = params['dataQ']

            node = {
                'id': f"tidl_layer_{layer_idx}",
                'index': layer_idx,
                'name': layer_info['layer_name'][:20] if layer_info['layer_name'] else f"Layer{layer_idx}",
                'full_name': layer_info['layer_name'] or f"Layer_{layer_idx}",
                'type': layer_info['layer_type'] or 'Unknown',
                'tidl_supported': True,
                'layer_data': layer_info,
                'layerdetails': layerdetails
            }

            graph_nodes.append(node)

        svg_edges = soup.find_all('g', class_='edge')

        for edge_elem in svg_edges:
            title = edge_elem.find('title')
            if not title or not title.string:
                continue

            edge_text = title.string

            arrow_idx = edge_text.find('->')
            if arrow_idx == -1:
                continue

            source_part = edge_text[:arrow_idx]
            source_match = re.search(r'Layer\s+(\d+):', source_part)
            if not source_match:
                continue
            source_idx = int(source_match.group(1))

            target_part = edge_text[arrow_idx+2:]
            target_match = re.search(r'Layer\s+(\d+):', target_part)
            if not target_match:
                continue
            target_idx = int(target_match.group(1))

            edge = {
                'source': f"tidl_layer_{source_idx}",
                'target': f"tidl_layer_{target_idx}",
                'source_node_id': source_idx,
                'target_node_id': target_idx
            }

            graph_edges.append(edge)

        return graph_nodes, graph_edges

    def parse_subgraph_html(self, filepath: str, netlog_filepath: str = None) -> Dict[str, Any]:
        """Parse a single subgraph HTML file and its netLog file"""
        print(f"  Parsing: {os.path.basename(filepath)}")

        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()

        soup = BeautifulSoup(content, 'html.parser')

        layers = []

        svg_nodes = soup.find_all('g', class_='node')

        for node in svg_nodes:
            title = node.find('title')
            if title and title.string:
                layer_info = self.parse_layer_info(title.string)
                if layer_info['layer_index'] is not None:
                    layers.append(layer_info)

        layers.sort(key=lambda x: x['layer_index'])

        print(f"    Found {len(layers)} TIDL layers")

        macs_data = {'layer_macs': {}, 'total_gmacs': 0.0}
        if netlog_filepath and os.path.exists(netlog_filepath):
            netlog_parser = TIDLNetLogParser(netlog_filepath)
            macs_data = netlog_parser.parse()

            for layer in layers:
                layer_idx = layer['layer_index']
                if layer_idx in macs_data['layer_macs']:
                    layer['macs'] = macs_data['layer_macs'][layer_idx]['macs']
                    layer['gmacs'] = macs_data['layer_macs'][layer_idx]['gmacs']

                    tidl_output_name = macs_data['layer_macs'][layer_idx].get('output_name', '')
                    onnx_node_idx = self._map_tidl_to_onnx_node(tidl_output_name)
                    if onnx_node_idx is not None:
                        layer['onnx_node_index'] = onnx_node_idx

        graph_nodes, graph_edges = self._extract_graph_structure(soup, layers)

        print(f"    Extracted graph: {len(graph_nodes)} nodes, {len(graph_edges)} edges")

        return {
            'filepath': filepath,
            'netlog_filepath': netlog_filepath,
            'num_layers': len(layers),
            'layers': layers,
            'total_gmacs': macs_data['total_gmacs'],
            'graph_nodes': graph_nodes,
            'graph_edges': graph_edges
        }

    def parse_all_subgraphs(self) -> Dict[int, Dict[str, Any]]:
        """Parse all subgraph HTML files and netLog files"""
        print("\nParsing TIDL subgraph HTML and netLog files...")

        subgraph_files = self.find_subgraph_files()

        if not subgraph_files:
            print("  No subgraph HTML files found")
            return {}

        print(f"  Found {len(subgraph_files)} subgraph HTML files")

        tidl_data = {}

        for subgraph_idx, html_filepath, netlog_filepath in subgraph_files:
            try:
                tidl_data[subgraph_idx] = self.parse_subgraph_html(html_filepath, netlog_filepath)
            except Exception as e:
                print(f"  Warning: Failed to parse {html_filepath}: {e}")
                continue

        print(f"Successfully parsed {len(tidl_data)} subgraph files")

        return tidl_data


class GraphVizParser:
    """Parser for graphvzInfo.txt"""

    def __init__(self, filepath: str):
        self.filepath = filepath

    def parse(self) -> Dict[int, Dict[str, Any]]:
        """Parse graphvizInfo.txt and extract support status"""
        node_support = {}

        print(f"Parsing graphvizInfo.txt: {self.filepath}")

        with open(self.filepath, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue

                parts = line.split(maxsplit=3)
                if len(parts) < 4:
                    continue

                try:
                    node_id = int(parts[0])
                    node_name = parts[1]
                    node_type = parts[2]

                    diag_match = re.search(r'diagInfo\s+(.+)$', line)
                    diag_info = diag_match.group(1) if diag_match else ''

                    is_supported = 'SUPPORTED' in diag_info and 'UNSUPPORTED' not in diag_info

                    node_support[node_id] = {
                        'supported': is_supported,
                        'diagInfo': diag_info,
                        'node_name': node_name,
                        'node_type': node_type
                    }

                except (ValueError, IndexError) as e:
                    print(f"  Warning: Could not parse line: {line[:50]}...")
                    continue

        supported_count = sum(1 for n in node_support.values() if n['supported'])
        unsupported_count = len(node_support) - supported_count

        print(f"Parsed {len(node_support)} nodes")
        print(f"  Supported: {supported_count}")
        print(f"  Unsupported: {unsupported_count}")

        return node_support


class AllowedNodeParser:
    """Parser for allowednode.txt"""

    def __init__(self, filepath: str):
        self.filepath = filepath

    def parse(self) -> List[Dict[str, Any]]:
        """Parse allowednode.txt and extract subgraph information"""
        print(f"Parsing allowednode.txt: {self.filepath}")

        with open(self.filepath, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f if line.strip()]

        if not lines:
            print("  Warning: Empty allowednode.txt")
            return []

        try:
            num_subgraphs = int(lines[0])
            print(f"  Number of subgraphs: {num_subgraphs}")

            subgraphs = []
            line_idx = 1

            for sg_idx in range(num_subgraphs):
                if line_idx >= len(lines):
                    break

                num_nodes = int(lines[line_idx])
                line_idx += 1

                nodes = []
                for _ in range(num_nodes):
                    if line_idx >= len(lines):
                        break
                    nodes.append(int(lines[line_idx]))
                    line_idx += 1

                subgraphs.append({
                    'id': sg_idx,
                    'nodes': nodes
                })

                print(f"  Subgraph {sg_idx}: {num_nodes} nodes")

            print(f"Parsed {len(subgraphs)} subgraphs")
            return subgraphs

        except (ValueError, IndexError) as e:
            print(f"  Error parsing allowednode.txt: {e}")
            return []


def calculate_node_depths_and_positions(nodes, edges, width=1200, height=800):
    """
    Calculate optimal x,y positions for nodes using Netron-style layout

    Args:
        nodes: Dict of node_name -> node_data
        edges: List of edge dictionaries
        width: Canvas width
        height: Canvas height

    Returns:
        Updated nodes dict with x, y, depth, horizontal_position
    """

    node_list = list(nodes.keys())
    node_to_idx = {name: idx for idx, name in enumerate(node_list)}

    children = {name: [] for name in node_list}
    parents = {name: [] for name in node_list}

    for edge in edges:
        src_name = edge.get('source_node_name')
        tgt_name = edge.get('target_node_name')

        if src_name in children and tgt_name in parents:
            children[src_name].append(tgt_name)
            parents[tgt_name].append(src_name)

    roots = [name for name in node_list if len(parents[name]) == 0]
    if not roots:
        roots = [node_list[0]]

    print(f"  Found {len(roots)} root nodes")

    depths = {}

    def assign_depth(node_name, depth):
        current_depth = depths.get(node_name, -1)
        if depth > current_depth:
            depths[node_name] = depth
            for child_name in children[node_name]:
                assign_depth(child_name, depth + 1)

    actual_roots = []
    constant_nodes = []

    for name in roots:
        node_data = nodes.get(name, {})
        node_type = node_data.get('type', '')

        if node_type in ['Constant', 'Initializer'] or 'Constant' in name:
            constant_nodes.append(name)
        else:
            actual_roots.append(name)

    if not actual_roots:
        actual_roots = roots

    print(f"  Actual roots: {len(actual_roots)}")
    print(f"  Constant nodes: {len(constant_nodes)}")

    for root in actual_roots:
        assign_depth(root, 0)

    for const_name in constant_nodes:
        if const_name not in depths:
            child_depths = [depths.get(child, 0) for child in children[const_name]]
            if child_depths:
                depths[const_name] = max(0, min(child_depths) - 1)
            else:
                depths[const_name] = 0

    for name in node_list:
        if name not in depths:
            depths[name] = 0

    max_depth = max(depths.values()) if depths else 0
    print(f"  Max depth: {max_depth}")

    depth_groups = {}
    for name, depth in depths.items():
        if depth not in depth_groups:
            depth_groups[depth] = []
        depth_groups[depth].append(name)

    horizontal_positions = {}

    for depth in sorted(depth_groups.keys()):
        nodes_at_depth = depth_groups[depth]

        if depth == 0:
            for i, name in enumerate(sorted(nodes_at_depth)):
                horizontal_positions[name] = i
        else:
            node_scores = []
            for name in nodes_at_depth:
                parent_positions = [
                    horizontal_positions.get(p, 0)
                    for p in parents[name]
                    if p in horizontal_positions
                ]
                avg_pos = sum(parent_positions) / len(parent_positions) if parent_positions else 0
                node_scores.append((name, avg_pos))

            node_scores.sort(key=lambda x: x[1])

            for i, (name, _) in enumerate(node_scores):
                horizontal_positions[name] = i

    max_width = max(len(nodes) for nodes in depth_groups.values()) if depth_groups else 1
    print(f"  Max width: {max_width}")

    VERTICAL_SPACING = 150
    HORIZONTAL_SPACING = 200
    PADDING = 100

    for name in node_list:
        depth = depths.get(name, 0)
        h_pos = horizontal_positions.get(name, 0)

        layer_width = len(depth_groups.get(depth, [])) * HORIZONTAL_SPACING
        start_x = (width - layer_width) / 2 + HORIZONTAL_SPACING / 2

        nodes[name]['x'] = start_x + h_pos * HORIZONTAL_SPACING
        nodes[name]['y'] = PADDING + depth * VERTICAL_SPACING
        nodes[name]['depth'] = depth
        nodes[name]['horizontal_position'] = h_pos

    print(f"  Calculated positions for {len(nodes)} nodes")

    return nodes


class ONNXParser:
    """Parser for ONNX models using GraphSurgeon for better readability and efficiency"""

    def __init__(self, model_path: str):
        self.model_path = model_path
        self.model = None
        self.gs_graph = None
        self.use_gs = HAS_GRAPHSURGEON

    def load_model(self):
        """Load ONNX model using GraphSurgeon or raw ONNX"""
        print(f"Loading ONNX model: {self.model_path}")
        self.model = onnx.load(self.model_path)
        onnx.checker.check_model(self.model)

        if self.use_gs:
            self.gs_graph = gs.import_onnx(self.model)
            print(f"Model loaded with GraphSurgeon (nodes: {len(self.gs_graph.nodes)}, tensors: {len(self.gs_graph.tensors())})")
        else:
            print("Model loaded with raw ONNX API")

    def get_tensor_shape(self, tensor) -> List:
        """Extract shape from GraphSurgeon tensor or ONNX ValueInfo

        Args:
            tensor: gs.Variable, gs.Constant, or onnx ValueInfoProto

        Returns:
            List of dimensions (ints or 'var' for dynamic)
        """
        if self.use_gs and (isinstance(tensor, gs.Variable) or isinstance(tensor, gs.Constant)):
            if tensor.shape is None:
                # Try to infer shape from values for Constants
                if isinstance(tensor, gs.Constant) and tensor.values is not None:
                    try:
                        import numpy as np
                        values = np.asarray(tensor.values)
                        return list(values.shape)
                    except:
                        pass
                return []
            shape = []
            for dim in tensor.shape:
                if isinstance(dim, int):
                    shape.append(dim)
                else:
                    shape.append('var')
            return shape
        else:
            if not hasattr(tensor, 'type') or not tensor.type.tensor_type.HasField('shape'):
                return []
            shape = []
            for dim in tensor.type.tensor_type.shape.dim:
                if dim.HasField('dim_value'):
                    shape.append(int(dim.dim_value))
                elif dim.HasField('dim_param'):
                    shape.append('var')
                else:
                    shape.append('var')
            return shape

    def format_shapes(self, shapes: List[List]) -> str:
        """Format shapes for display"""
        if not shapes:
            return 'N/A'

        formatted = []
        for shape in shapes:
            if not shape:
                continue

            shape_str = ','.join(map(str, shape))
            formatted.append(f"[{shape_str}]")

        return ', '.join(formatted) if formatted else 'N/A'

    def get_tensor_dtype(self, tensor) -> str:
        """Get data type of a tensor"""
        try:
            if hasattr(tensor, 'type') and tensor.type:
                if tensor.type.HasField('tensor_type'):
                    elem_type = tensor.type.tensor_type.elem_type
                    dtype_map = {
                        1: 'float32', 2: 'uint8', 3: 'int8', 4: 'uint16',
                        5: 'int16', 6: 'int32', 7: 'int64', 8: 'string',
                        9: 'bool', 10: 'float16', 11: 'float64', 12: 'uint32',
                        13: 'uint64', 14: 'complex64', 15: 'complex128',
                        16: 'bfloat16'
                    }
                    return dtype_map.get(elem_type, f'unknown({elem_type})')
            elif hasattr(tensor, 'data_type'):
                dtype_map = {
                    1: 'float32', 2: 'uint8', 3: 'int8', 4: 'uint16',
                    5: 'int16', 6: 'int32', 7: 'int64', 8: 'string',
                    9: 'bool', 10: 'float16', 11: 'float64', 12: 'uint32',
                    13: 'uint64', 14: 'complex64', 15: 'complex128',
                    16: 'bfloat16'
                }
                return dtype_map.get(tensor.data_type, f'unknown({tensor.data_type})')
        except:
            pass
        return 'unknown'

    def extract_node_attributes(self, node) -> Dict[str, Any]:
        """Extract attributes from GraphSurgeon node or ONNX node"""
        if self.use_gs and isinstance(node, gs.Node):
            attrs = {}
            for key, val in node.attrs.items():
                if isinstance(val, np.ndarray):
                    attrs[key] = val.tolist()
                elif isinstance(val, (list, tuple)):
                    attrs[key] = list(val)
                elif isinstance(val, bytes):
                    attrs[key] = val.decode('utf-8')
                else:
                    attrs[key] = val
            return attrs
        else:
            attrs = {}
            for attr in node.attribute:
                name = attr.name
                if attr.type == onnx.AttributeProto.INTS:
                    attrs[name] = list(attr.ints)
                elif attr.type == onnx.AttributeProto.INT:
                    attrs[name] = int(attr.i)
                elif attr.type == onnx.AttributeProto.FLOATS:
                    attrs[name] = list(attr.floats)
                elif attr.type == onnx.AttributeProto.FLOAT:
                    attrs[name] = float(attr.f)
                elif attr.type == onnx.AttributeProto.STRING:
                    attrs[name] = attr.s.decode('utf-8')
                elif attr.type == onnx.AttributeProto.STRINGS:
                    attrs[name] = [s.decode('utf-8') for s in attr.strings]
                elif attr.type == onnx.AttributeProto.TENSOR:
                    attrs[name] = 'Tensor'
            return attrs

    def parse(self) -> Dict[str, Any]:
        """Parse ONNX model and extract all information using GraphSurgeon or raw ONNX"""
        self.load_model()

        model_name = Path(self.model_path).stem

        if self.use_gs:
            return self._parse_with_graphsurgeon(model_name)
        else:
            return self._parse_with_onnx(model_name)

    def _parse_with_graphsurgeon(self, model_name: str) -> Dict[str, Any]:
        """Parse using GraphSurgeon API"""
        graph = self.gs_graph

        total_params = 0
        for tensor in graph.tensors().values():
            if isinstance(tensor, gs.Constant) and tensor.values is not None:
                total_params += tensor.values.size

        print(f"Total parameters: {total_params:,}")

        tensor_metadata = {}
        tensor_dict = graph.tensors()
        constant_names = set()

        for tensor_name, tensor in tensor_dict.items():
            if isinstance(tensor, gs.Constant):
                constant_names.add(tensor_name)
                shape = self.get_tensor_shape(tensor)
                dtype = 'unknown'
                if tensor.values is not None and hasattr(tensor.values, 'dtype'):
                    dtype_map = {
                        'float32': 'float32', 'uint8': 'uint8', 'int8': 'int8',
                        'int16': 'int16', 'int32': 'int32', 'int64': 'int64',
                        'float16': 'float16', 'float64': 'float64', 'bool': 'bool'
                    }
                    dtype = dtype_map.get(str(tensor.values.dtype), str(tensor.values.dtype))

                tensor_metadata[tensor_name] = {
                    'shape': shape,
                    'dtype': dtype,
                    'is_constant': True
                }

        for inp in graph.inputs:
            if inp.name not in constant_names:
                shape = self.get_tensor_shape(inp)
                dtype = 'unknown'
                if hasattr(inp, 'dtype') and inp.dtype is not None:
                    dtype_map = {
                        1: 'float32', 2: 'uint8', 3: 'int8', 6: 'int32', 7: 'int64',
                        10: 'float16', 11: 'float64', 9: 'bool'
                    }
                    dtype = dtype_map.get(inp.dtype, f'type_{inp.dtype}')

                tensor_metadata[inp.name] = {
                    'shape': shape,
                    'dtype': dtype,
                    'is_constant': False
                }
                print(f"Input: {inp.name} -> {shape} ({dtype})")

        for out in graph.outputs:
            shape = self.get_tensor_shape(out)
            dtype = 'unknown'
            if hasattr(out, 'dtype') and out.dtype is not None:
                dtype_map = {
                    1: 'float32', 2: 'uint8', 3: 'int8', 6: 'int32', 7: 'int64',
                    10: 'float16', 11: 'float64', 9: 'bool'
                }
                dtype = dtype_map.get(out.dtype, f'type_{out.dtype}')

            if out.name not in tensor_metadata:
                tensor_metadata[out.name] = {
                    'shape': shape,
                    'dtype': dtype,
                    'is_constant': False
                }
            print(f"Output: {out.name} -> {shape} ({dtype})")

        for tensor_name, tensor in tensor_dict.items():
            if tensor_name not in tensor_metadata and not isinstance(tensor, gs.Constant):
                shape = self.get_tensor_shape(tensor)
                dtype = 'unknown'
                if hasattr(tensor, 'dtype') and tensor.dtype is not None:
                    dtype_map = {
                        1: 'float32', 2: 'uint8', 3: 'int8', 6: 'int32', 7: 'int64',
                        10: 'float16', 11: 'float64', 9: 'bool'
                    }
                    dtype = dtype_map.get(tensor.dtype, f'type_{tensor.dtype}')

                tensor_metadata[tensor_name] = {
                    'shape': shape,
                    'dtype': dtype,
                    'is_constant': False
                }

        print(f"Total tensors tracked: {len(tensor_metadata)} (constants: {len(constant_names)})")

        shape_lookup = {name: meta['shape'] for name, meta in tensor_metadata.items()}

        model_details = {
            'name': model_name,
            'weights': total_params,
            'no_of_layers': len(graph.nodes),
            'input_shape': [
                {
                    'name': inp.name,
                    'shape': self.get_tensor_shape(inp)
                } for inp in graph.inputs
            ],
            'output_shape': [
                {
                    'name': out.name,
                    'shape': self.get_tensor_shape(out)
                } for out in graph.outputs
            ]
        }

        output_to_node = {}
        for idx, node in enumerate(graph.nodes):
            for out_tensor in node.outputs:
                output_to_node[out_tensor.name] = idx

        layer_details = {}
        edges = []

        print(f"\nProcessing {len(graph.nodes)} nodes...")

        for idx, node in enumerate(graph.nodes):
            node_name = node.name if node.name else f"{node.op}_{idx}"
            attrs = self.extract_node_attributes(node)

            input_names = [inp_tensor.name for inp_tensor in node.inputs]
            output_names = [out_tensor.name for out_tensor in node.outputs]

            input_shapes = []
            for inp_name in input_names:
                shape = shape_lookup.get(inp_name, [])
                if shape:
                    input_shapes.append(shape)

            output_shapes = []
            for out_name in output_names:
                shape = shape_lookup.get(out_name, [])
                if shape:
                    output_shapes.append(shape)

            input_metadata = []
            for inp_name in input_names:
                meta = tensor_metadata.get(inp_name, {})
                input_metadata.append({
                    'name': inp_name,
                    'shape': meta.get('shape', []),
                    'dtype': meta.get('dtype', 'unknown'),
                    'is_constant': meta.get('is_constant', False)
                })

            output_metadata = []
            for out_name in output_names:
                meta = tensor_metadata.get(out_name, {})
                output_metadata.append({
                    'name': out_name,
                    'shape': meta.get('shape', []),
                    'dtype': meta.get('dtype', 'unknown'),
                    'is_constant': False
                })

            layer_details[node_name] = {
                'layer_name': node_name,
                'type': node.op,
                'input': input_names,
                'output': output_names,
                'input_metadata': input_metadata,
                'output_metadata': output_metadata,
                'x': 0,
                'y': 0,
                'depth': 0,
                'horizontal_position': 0,
                'input_shape': {
                    'channels': len(input_names),
                    'shapes': input_shapes,
                    'formatted': self.format_shapes(input_shapes)
                },
                'output_shape': {
                    'channels': len(output_names),
                    'shapes': output_shapes,
                    'formatted': self.format_shapes(output_shapes)
                },
                'kernel_shapes': attrs.get('kernel_shape', 'N/A'),
                'strides': attrs.get('strides', 'N/A'),
                'dilations': attrs.get('dilations', 'N/A'),
                'pads': attrs.get('pads', 'N/A'),
                'groups': attrs.get('group', 1),
                'activation': attrs.get('activation', 'N/A'),
                'auto_pad': attrs.get('auto_pad', 'N/A'),
                'attributes': attrs
            }

            for input_name in input_names:
                source_idx = output_to_node.get(input_name)
                if source_idx is not None and source_idx != idx:
                    source_node = graph.nodes[source_idx]
                    source_name = source_node.name if source_node.name else f"{source_node.op}_{source_idx}"

                    edges.append({
                        'source_node_id': source_idx,
                        'target_node_id': idx,
                        'source_node_name': source_name,
                        'target_node_name': node_name,
                        'connection_info': {
                            'tensor': input_name,
                            'shape': shape_lookup.get(input_name, [])
                        }
                    })

            if (idx + 1) % 10 == 0:
                print(f"  Processed {idx + 1}/{len(graph.nodes)} nodes")

        print(f"Extracted {len(layer_details)} layers and {len(edges)} edges")

        print("\nCalculating node positions...")
        layer_details = calculate_node_depths_and_positions(
            layer_details,
            edges,
            width=1200,
            height=800
        )

        return {
            'model_details': model_details,
            'layer_details': layer_details,
            'edges': edges
        }

    def _parse_with_onnx(self, model_name: str) -> Dict[str, Any]:
        """Fallback: Parse using raw ONNX API"""
        graph = self.model.graph

        total_params = 0
        for initializer in graph.initializer:
            dims = list(initializer.dims)
            if dims:
                size = 1
                for dim in dims:
                    size *= dim
                total_params += size

        print(f"Total parameters: {total_params:,}")

        tensor_metadata = {}
        initializer_names = set()

        for init in graph.initializer:
            numeric_dims = [int(d) for d in init.dims]
            # If dims is empty, try to infer shape from data size
            if not numeric_dims:
                # Check various data fields to infer shape
                data_length = 0
                if init.float_data:
                    data_length = len(init.float_data)
                elif init.int32_data:
                    data_length = len(init.int32_data)
                elif init.int64_data:
                    data_length = len(init.int64_data)
                elif init.raw_data:
                    # Calculate based on data type size
                    type_sizes = {1: 4, 2: 1, 3: 1, 6: 4, 7: 8, 10: 2, 11: 8}  # float32, uint8, int8, int32, int64, float16, float64
                    if init.data_type in type_sizes:
                        data_length = len(init.raw_data) // type_sizes[init.data_type]

                if data_length > 0:
                    numeric_dims = [data_length]

            initializer_names.add(init.name)
            tensor_metadata[init.name] = {
                'shape': numeric_dims,
                'dtype': self.get_tensor_dtype(init),
                'is_constant': True
            }

        for inp in graph.input:
            if inp.name not in initializer_names:
                shape = self.get_tensor_shape(inp)
                tensor_metadata[inp.name] = {
                    'shape': shape,
                    'dtype': self.get_tensor_dtype(inp),
                    'is_constant': False
                }
                print(f"Input: {inp.name} -> {shape} ({self.get_tensor_dtype(inp)})")

        for out in graph.output:
            shape = self.get_tensor_shape(out)
            if out.name not in tensor_metadata:
                tensor_metadata[out.name] = {
                    'shape': shape,
                    'dtype': self.get_tensor_dtype(out),
                    'is_constant': False
                }
            print(f"Output: {out.name} -> {shape} ({self.get_tensor_dtype(out)})")

        for vi in graph.value_info:
            shape = self.get_tensor_shape(vi)
            if vi.name not in tensor_metadata:
                tensor_metadata[vi.name] = {
                    'shape': shape,
                    'dtype': self.get_tensor_dtype(vi),
                    'is_constant': False
                }

        print(f"Total tensors tracked: {len(tensor_metadata)} (constants: {len(initializer_names)})")

        shape_lookup = {name: meta['shape'] for name, meta in tensor_metadata.items()}

        model_details = {
            'name': model_name,
            'weights': total_params,
            'no_of_layers': len(graph.node),
            'input_shape': [
                {
                    'name': inp.name,
                    'shape': self.get_tensor_shape(inp)
                } for inp in graph.input
            ],
            'output_shape': [
                {
                    'name': out.name,
                    'shape': self.get_tensor_shape(out)
                } for out in graph.output
            ]
        }

        output_to_node = {}
        for idx, node in enumerate(graph.node):
            for output in node.output:
                output_to_node[output] = idx

        layer_details = {}
        edges = []

        print(f"\nProcessing {len(graph.node)} nodes...")

        for idx, node in enumerate(graph.node):
            node_name = node.name if node.name else f"{node.op_type}_{idx}"
            attrs = self.extract_node_attributes(node)

            input_shapes = []
            for inp in node.input:
                shape = shape_lookup.get(inp, [])
                if shape:
                    input_shapes.append(shape)

            output_shapes = []
            for out in node.output:
                shape = shape_lookup.get(out, [])
                if shape:
                    output_shapes.append(shape)

            input_metadata = []
            for inp_name in node.input:
                meta = tensor_metadata.get(inp_name, {})
                input_metadata.append({
                    'name': inp_name,
                    'shape': meta.get('shape', []),
                    'dtype': meta.get('dtype', 'unknown'),
                    'is_constant': meta.get('is_constant', False)
                })

            output_metadata = []
            for out_name in node.output:
                meta = tensor_metadata.get(out_name, {})
                output_metadata.append({
                    'name': out_name,
                    'shape': meta.get('shape', []),
                    'dtype': meta.get('dtype', 'unknown'),
                    'is_constant': False
                })

            layer_details[node_name] = {
                'layer_name': node_name,
                'type': node.op_type,
                'input': list(node.input),
                'output': list(node.output),
                'input_metadata': input_metadata,
                'output_metadata': output_metadata,
                'x': 0,
                'y': 0,
                'depth': 0,
                'horizontal_position': 0,
                'input_shape': {
                    'channels': len(node.input),
                    'shapes': input_shapes,
                    'formatted': self.format_shapes(input_shapes)
                },
                'output_shape': {
                    'channels': len(node.output),
                    'shapes': output_shapes,
                    'formatted': self.format_shapes(output_shapes)
                },
                'kernel_shapes': attrs.get('kernel_shape', 'N/A'),
                'strides': attrs.get('strides', 'N/A'),
                'dilations': attrs.get('dilations', 'N/A'),
                'pads': attrs.get('pads', 'N/A'),
                'groups': attrs.get('group', 1),
                'activation': attrs.get('activation', 'N/A'),
                'auto_pad': attrs.get('auto_pad', 'N/A'),
                'attributes': attrs
            }

            for input_name in node.input:
                source_idx = output_to_node.get(input_name)
                if source_idx is not None and source_idx != idx:
                    source_node = graph.node[source_idx]
                    source_name = source_node.name if source_node.name else f"{source_node.op_type}_{source_idx}"

                    edges.append({
                        'source_node_id': source_idx,
                        'target_node_id': idx,
                        'source_node_name': source_name,
                        'target_node_name': node_name,
                        'connection_info': {
                            'tensor': input_name,
                            'shape': shape_lookup.get(input_name, [])
                        }
                    })

            if (idx + 1) % 10 == 0:
                print(f"  Processed {idx + 1}/{len(graph.node)} nodes")

        print(f"Extracted {len(layer_details)} layers and {len(edges)} edges")

        print("\nCalculating node positions...")
        layer_details = calculate_node_depths_and_positions(
            layer_details,
            edges,
            width=1200,
            height=800
        )

        return {
            'model_details': model_details,
            'layer_details': layer_details,
            'edges': edges
        }


def build_hierarchical_tree(layer_details: Dict[str, Any], edges: List[Dict]) -> Dict[str, Any]:
    """Build hierarchical tree structure from layer details based on node naming"""
    print("\nBuilding hierarchical tree structure...")

    node_topo_indices = {name: idx for idx, name in enumerate(layer_details.keys())}

    tree_dict = {}

    for node_name, node_data in layer_details.items():
        node_path = []
        for part in node_name.split('/'):
            node_path.extend(part.split('.'))

        if not node_path:
            continue

        node_info = {
            "name": node_name,
            "op": node_data.get('type', 'Unknown'),
            "is_leaf": True,
            "node_details": {
                "name": node_name,
                "op": node_data.get('type', 'Unknown'),
                "inputs": node_data.get('input', []),
                "outputs": node_data.get('output', []),
                "input_metadata": node_data.get('input_metadata', []),
                "output_metadata": node_data.get('output_metadata', []),
                "attributs": node_data.get('attributes', {})
            },
            "topo_idx": node_topo_indices.get(node_name, 0)
        }

        current_level = tree_dict

        for i, part in enumerate(node_path):
            is_last = (i == len(node_path) - 1)

            if part not in current_level:
                if is_last:
                    current_level[part] = node_info
                else:
                    current_level[part] = {
                        "is_leaf": False,
                        "children": {}
                    }
            elif not is_last:
                current_level[part]["is_leaf"] = False

            if not is_last:
                if "children" not in current_level[part]:
                    current_level[part]["children"] = {}
                current_level = current_level[part]["children"]

    print(f"  Built tree with {len(tree_dict)} top-level modules")

    tree_dict = flatten_single_child_modules(tree_dict)

    return tree_dict


def flatten_single_child_modules(tree: Dict, parent_key: str = '') -> Dict:
    """Recursively flatten modules with only one child module"""
    flattened = {}

    for key, value in tree.items():
        if not value.get("is_leaf", False) and "children" in value:
            value["children"] = flatten_single_child_modules(value["children"], key)
            children = value["children"]

            if len(children) == 1:
                child_key = list(children.keys())[0]
                child_value = children[child_key]

                combined_key = f"{key}/{child_key}"

                if not child_value.get("is_leaf", False) and "children" in child_value:
                    flattened[combined_key] = child_value
                else:
                    flattened[combined_key] = child_value
            else:
                min_topo_idx = float('inf')

                def find_min_topo_idx(node):
                    nonlocal min_topo_idx
                    if node.get("is_leaf", False):
                        if "topo_idx" in node:
                            min_topo_idx = min(min_topo_idx, node["topo_idx"])
                    elif "children" in node:
                        for child_node in node["children"].values():
                            find_min_topo_idx(child_node)

                find_min_topo_idx(value)
                value["min_topo_idx"] = min_topo_idx if min_topo_idx != float('inf') else 0
                flattened[key] = value
        else:
            flattened[key] = value

    return flattened


def discover_files_from_workdir(model_dir_path: str) -> Dict[str, str]:
    """Auto-discover all required files from model directory structure"""
    discovered = {}

    print(f"\nAuto-discovering files in: {model_dir_path}")
    print("=" * 70)

    import glob

    onnx_files = glob.glob(os.path.join(model_dir_path, 'model/*.onnx'), recursive=False)
    if onnx_files:
        discovered['onnx'] = onnx_files[0]
        print(f"[FOUND] ONNX model: {os.path.relpath(discovered['onnx'])}")
    else:
        print("[NOT FOUND] ONNX model not found")

    graphviz_files = glob.glob(os.path.join(model_dir_path, 'artifacts/tempDir/graphvizInfo.txt'), recursive=False)
    if graphviz_files:
        discovered['graphviz'] = graphviz_files[0]
        print(f"[FOUND] graphvizInfo: {os.path.relpath(discovered['graphviz'])}")
    else:
        print("[NOT FOUND] graphvizInfo.txt not found")

    allowednode_files = glob.glob(os.path.join(model_dir_path, 'artifacts/allowedNode.txt'), recursive=False)
    if allowednode_files:
        discovered['allowednode'] = allowednode_files[0]
        print(f"[FOUND] allowedNode: {os.path.relpath(discovered['allowednode'])}")
    else:
        print("[NOT FOUND] allowedNode.txt not found")

    subgraph_dir = None
    subgraph_html_files = glob.glob(os.path.join(model_dir_path, 'artifacts/tempDir/subgraph_*_tidl_net.bin.html'), recursive=False)
    if subgraph_html_files:
        subgraph_dir = os.path.dirname(subgraph_html_files[0])
        discovered['subgraph_dir'] = subgraph_dir
        print(f"[FOUND] {len(subgraph_html_files)} subgraph HTML files in: {os.path.relpath(subgraph_dir)}")
    else:
        print("[NOT FOUND] Subgraph HTML files not found")

    xlsx_files = glob.glob(os.path.join(model_dir_path, 'analyze.xlsx'), recursive=False)
    if xlsx_files:
        discovered['xlsx'] = xlsx_files[0]
        print(f"[FOUND] analyze.xlsx: {os.path.relpath(discovered['xlsx'])}")
    else:
        print("[NOT FOUND] analyze.xlsx not found")

    mapping_files = glob.glob(os.path.join(model_dir_path, 'layer_output_mapping*.yaml'), recursive=False)
    if mapping_files:
        notidl_tidl_files = [f for f in mapping_files if 'notidl_tidl.yaml' in f and 'tidl32' not in f]
        if notidl_tidl_files:
            discovered['activation_yaml'] = notidl_tidl_files[0]
        else:
            discovered['activation_yaml'] = mapping_files[0]
        print(f"[FOUND] activation mapping: {os.path.relpath(discovered['activation_yaml'])}")
    else:
        print("WARNING: Activation mapping YAML not found (optional)")

    print("=" * 70)
    return discovered


def load_config_data(model_dir_path: str) -> Dict[str, Any]:
    """Load configuration data from model directory (config.yaml and result.yaml)"""
    config_data = {}

    import glob
    config_files = glob.glob(os.path.join(model_dir_path, '**/config.yaml'), recursive=True)
    result_files = glob.glob(os.path.join(model_dir_path, '**/result.yaml'), recursive=True)

    config_path = None
    result_path = None

    for path in config_files:
        if '/tidl/' not in path and '/tidl32/' not in path and '/notidl/' not in path:
            config_path = path
            break
    if not config_path:
        for path in config_files:
            if '/tidl/' in path:
                config_path = path
                break
    if not config_path and config_files:
        config_path = config_files[0]

    for path in result_files:
        if '/tidl/' not in path and '/tidl32/' not in path and '/notidl/' not in path:
            result_path = path
            break
    if not result_path:
        for path in result_files:
            if '/tidl/' in path:
                result_path = path
                break
    if not result_path and result_files:
        result_path = result_files[0]

    if config_path and os.path.exists(config_path):
        try:
            import yaml
            with open(config_path, 'r') as f:
                config = yaml.safe_load(f)

            config_data['target_device'] = config.get('session', {}).get('target_device', 'Unknown')
            config_data['task_type'] = config.get('common', {}).get('task_type', 'Unknown')
            config_data['tensor_bits'] = config.get('session', {}).get('runtime_options', {}).get('tensor_bits', 'Unknown')

            metric_ref = config.get('model_info', {}).get('metric_reference', {})
            if 'accuracy_top1%' in metric_ref:
                config_data['accuracy'] = f"{metric_ref['accuracy_top1%']}%"
            else:
                config_data['accuracy'] = 'N/A'

            print(f"Loaded config data from: {config_path}")
        except Exception as e:
            print(f"WARNING: Failed to load config.yaml: {e}")

    if result_path and os.path.exists(result_path):
        try:
            import yaml
            with open(result_path, 'r') as f:
                result = yaml.safe_load(f)

            result_data = result.get('result', {})
            config_data['num_frames'] = result_data.get('num_frames', 'N/A')
            config_data['num_subgraphs'] = result_data.get('num_subgraphs', 'N/A')
            config_data['perfsim_ddr_transfer_mb'] = result_data.get('perfsim_ddr_transfer_mb', 'N/A')
            config_data['perfsim_gmacs'] = result_data.get('perfsim_gmacs', 'N/A')
            config_data['perfsim_time_ms'] = result_data.get('perfsim_time_ms', 'N/A')

            print(f"Loaded result data from: {result_path}")
            print(f"  -> num_frames: {config_data.get('num_frames')}")
            print(f"  -> num_subgraphs: {config_data.get('num_subgraphs')}")
            print(f"  -> perfsim_ddr_transfer_mb: {config_data.get('perfsim_ddr_transfer_mb')}")
            print(f"  -> perfsim_gmacs: {config_data.get('perfsim_gmacs')}")
            print(f"  -> perfsim_time_ms: {config_data.get('perfsim_time_ms')}")
        except Exception as e:
            print(f"WARNING: Failed to load result.yaml: {e}")

    return config_data


def load_proctime_data(model_dir_path: str) -> Dict[int, List[Dict[str, Any]]]:
    """Load proctime data from performance CSV files for each subgraph"""
    proctime_data = {}

    import glob
    csv_files = glob.glob(os.path.join(model_dir_path, '**/tempDir/subgraph_*/tempDir_subgraph_*_tidl_net_*.csv'), recursive=True)

    preferred_files = []
    for path in csv_files:
        if '/tidl/' in path:
            preferred_files.append(path)

    if not preferred_files:
        preferred_files = csv_files

    processed_subgraphs = set()

    for csv_path in preferred_files:
        try:
            import re
            match = re.search(r'subgraph_(\d+)_tidl_net', csv_path)
            if not match:
                continue

            subgraph_num = int(match.group(1))

            # Skip if already processed this subgraph
            if subgraph_num in processed_subgraphs:
                continue

            import csv
            with open(csv_path, 'r') as f:
                reader = csv.DictReader(f)
                layer_data = []

                for row in reader:
                    try:
                        layer_num = int(row[' lyrNum '].strip())
                        layer_type = row[' LyrType'].strip()
                        proctime_str = row['procTime(us)'].strip()

                        proctime = float(proctime_str) if proctime_str else 0.0

                        layer_data.append({
                            'layer_num': layer_num,
                            'layer_type': layer_type,
                            'proctime': proctime
                        })
                    except (ValueError, KeyError) as e:
                        continue

                if layer_data:
                    proctime_data[subgraph_num] = layer_data
                    processed_subgraphs.add(subgraph_num)
                    print(f"Loaded proctime data for subgraph {subgraph_num}: {len(layer_data)} layers")

        except Exception as e:
            print(f"WARNING: Failed to load proctime CSV {csv_path}: {e}")

    return proctime_data


def load_cycles_data(model_dir_path: str) -> Dict[int, List[Dict[str, Any]]]:
    """Load kernel and core loop cycles data from performance CSV files for each subgraph"""
    cycles_data = {}

    import glob
    csv_files = glob.glob(os.path.join(model_dir_path, '**/tempDir/subgraph_*/tempDir_subgraph_*_tidl_net_*.csv'), recursive=True)

    preferred_files = []
    for path in csv_files:
        if '/tidl/' in path:
            preferred_files.append(path)

    if not preferred_files:
        preferred_files = csv_files

    processed_subgraphs = set()

    for csv_path in preferred_files:
        try:
            import re
            match = re.search(r'subgraph_(\d+)_tidl_net', csv_path)
            if not match:
                continue

            subgraph_num = int(match.group(1))

            # Skip if already processed this subgraph
            if subgraph_num in processed_subgraphs:
                continue

            import csv
            with open(csv_path, 'r') as f:
                reader = csv.DictReader(f)
                layer_data = []

                for row in reader:
                    try:
                        layer_num = int(row[' lyrNum '].strip())
                        layer_type = row[' LyrType'].strip()
                        kernel_only_cycles_str = row['kernelOnlyCycles'].strip()
                        core_loop_cycles_str = row['coreLoopCycles'].strip()

                        kernel_only_cycles = float(kernel_only_cycles_str) if kernel_only_cycles_str else 0.0
                        core_loop_cycles = float(core_loop_cycles_str) if core_loop_cycles_str else 0.0

                        layer_data.append({
                            'layer_num': layer_num,
                            'layer_type': layer_type,
                            'kernelOnlyCycles': kernel_only_cycles,
                            'coreLoopCycles': core_loop_cycles
                        })
                    except (ValueError, KeyError) as e:
                        continue

                if layer_data:
                    cycles_data[subgraph_num] = layer_data
                    processed_subgraphs.add(subgraph_num)
                    print(f"Loaded cycles data for subgraph {subgraph_num}: {len(layer_data)} layers")

        except Exception as e:
            print(f"WARNING: Failed to load cycles CSV {csv_path}: {e}")

    return cycles_data


def load_memory_data(model_dir_path: str) -> Dict[int, List[Dict[str, Any]]]:
    """Load memory usage data from performance CSV files for each subgraph"""
    memory_data = {}

    import glob
    csv_files = glob.glob(os.path.join(model_dir_path, '**/tempDir/subgraph_*/tempDir_subgraph_*_tidl_net_*.csv'), recursive=True)

    preferred_files = []
    for path in csv_files:
        if '/tidl/' in path:
            preferred_files.append(path)

    if not preferred_files:
        preferred_files = csv_files

    processed_subgraphs = set()

    for csv_path in preferred_files:
        try:
            import re
            match = re.search(r'subgraph_(\d+)_tidl_net', csv_path)
            if not match:
                continue

            subgraph_num = int(match.group(1))

            # Skip if already processed this subgraph
            if subgraph_num in processed_subgraphs:
                continue

            import csv
            with open(csv_path, 'r') as f:
                reader = csv.DictReader(f)
                layer_data = []

                for row in reader:
                    try:
                        layer_num = int(row[' lyrNum '].strip())
                        layer_type = row[' LyrType'].strip()

                        in_vol = float(row['inVol(KB)'].strip()) if row['inVol(KB)'].strip() else 0.0
                        out_vol = float(row['outVol(KB)'].strip()) if row['outVol(KB)'].strip() else 0.0
                        wt_vol = float(row['wtVol(KB)'].strip()) if row['wtVol(KB)'].strip() else 0.0

                        src_mem_in = row[' srcMem-IN'].strip()
                        dst_mem_in = row[' dstMem-IN'].strip()
                        src_mem_out = row['srcMem-OUT'].strip()
                        dst_mem_out = row['dstMem-OUT'].strip()
                        src_mem_wt = row[' srcMem-WT'].strip()
                        dst_mem_wt = row[' dstMem-WT'].strip()

                        l2_usage = 0.0
                        msmc_usage = 0.0
                        ddr_usage = 0.0

                        for mem_loc, vol in [(src_mem_in, in_vol), (dst_mem_in, in_vol),
                                             (src_mem_out, out_vol), (dst_mem_out, out_vol),
                                             (src_mem_wt, wt_vol), (dst_mem_wt, wt_vol)]:
                            if 'L2' in mem_loc:
                                l2_usage += vol / 2
                            elif 'MSMC' in mem_loc or 'L3' in mem_loc:
                                msmc_usage += vol / 2
                            elif 'DDR' in mem_loc:
                                ddr_usage += vol / 2

                        layer_data.append({
                            'layer_num': layer_num,
                            'layer_type': layer_type,
                            'l2_usage': l2_usage,
                            'msmc_usage': msmc_usage,
                            'ddr_usage': ddr_usage,
                            'total_usage': l2_usage + msmc_usage + ddr_usage
                        })
                    except (ValueError, KeyError) as e:
                        continue

                if layer_data:
                    memory_data[subgraph_num] = layer_data
                    processed_subgraphs.add(subgraph_num)
                    print(f"Loaded memory data for subgraph {subgraph_num}: {len(layer_data)} layers")

        except Exception as e:
            print(f"WARNING: Failed to load memory CSV {csv_path}: {e}")

    return memory_data


def main(work_dirs_path, output_json_path):
    """Main function to extract all artifact data to JSON"""
    if len(sys.argv) < 3:
        print("=" * 70)
        print("Data Extractor - Extract TIDL Artifacts to JSON")
        print("=" * 70)
        print("\nUsage: python data_extractor.py <model_dir/> <output.json>")
        print("\nArguments:")
        print("  model_dir/   - Direct path to model directory (e.g., work_dirs/compile/AM69A/cl_onnx_model_name/)")
        print("  output.json  - Output JSON file path (will be compressed)")
        print("\nExample:")
        print("  python data_extractor.py work_dirs/compile/AM69A/cl-ort-resnet18/ model_data.json")
        print("\nThe script will automatically discover and parse:")
        print("  - ONNX model from <model_dir>/model/*.onnx")
        print("  - GraphViz info from <model_dir>/artifacts/tempDir/graphvizInfo.txt")
        print("  - Allowed nodes from <model_dir>/artifacts/allowedNode.txt")
        print("  - Subgraph files from <model_dir>/artifacts/tempDir/")
        print("  - Metrics from <model_dir>/analyze.xlsx")
        print("  - Activation data from layer_info.txt and binary files")
        print("  - Config/Result from <model_dir>/tidl/*.yaml")
        print("  - Performance data from <model_dir>/tidl/artifacts/tempDir/**/*.csv")
        print("=" * 70)
        sys.exit(1)

    model_dir_path = sys.argv[1]
    output_json_path = sys.argv[2]

    if not os.path.exists(model_dir_path):
        print(f"ERROR: Model directory not found: {model_dir_path}")
        sys.exit(1)

    if not os.path.isdir(model_dir_path):
        print(f"ERROR: {model_dir_path} is not a directory")
        sys.exit(1)

    discovered = discover_files_from_workdir(model_dir_path)

    required_keys = ['onnx', 'graphviz', 'allowednode', 'subgraph_dir']
    missing = [key for key in required_keys if key not in discovered]
    if missing:
        print(f"\nERROR: Required files not found: {', '.join(missing)}")
        sys.exit(1)

    onnx_path = discovered['onnx']
    graphviz_path = discovered['graphviz']
    allowednode_path = discovered['allowednode']
    subgraph_dir = discovered['subgraph_dir']

    print("\n" + "=" * 70)
    print("Data Extractor - Parsing Artifacts")
    print("=" * 70)
    print(f"ONNX Model:      {onnx_path}")
    print(f"GraphViz Info:   {graphviz_path}")
    print(f"Allowed Nodes:   {allowednode_path}")
    print(f"Subgraph Dir:    {subgraph_dir}")
    print(f"Output JSON:     {output_json_path}")
    print("=" * 70)

    try:
        print("\n[1/8] Parsing ONNX model...")
        onnx_parser = ONNXParser(onnx_path)
        model_data = onnx_parser.parse()

        tree_structure = build_hierarchical_tree(
            model_data['layer_details'],
            model_data['edges']
        )
        model_data['tree_structure'] = tree_structure

        subgraph_data = {
            'subgraphs': [],
            'node_support': {}
        }

        print("\n[2/8] Parsing GraphViz info...")
        if os.path.exists(graphviz_path):
            graphviz_parser = GraphVizParser(graphviz_path)
            subgraph_data['node_support'] = graphviz_parser.parse()
        else:
            print(f"WARNING: {graphviz_path} not found")

        print("\n[3/8] Parsing allowed nodes...")
        if os.path.exists(allowednode_path):
            allowednode_parser = AllowedNodeParser(allowednode_path)
            subgraph_data['subgraphs'] = allowednode_parser.parse()
        else:
            print(f"WARNING: {allowednode_path} not found")

        print("\n[4/8] Parsing TIDL subgraph HTML files...")
        tidl_parser = TIDLSubgraphParser(subgraph_dir, subgraph_data['node_support'])
        tidl_data = tidl_parser.parse_all_subgraphs()

        print("\n[5/8] Parsing activation data...")
        activation_data = {}
        try:
            activation_parser = ActivationDataParser(model_dir_path, frame_idx=0)
            activation_data = activation_parser.process_all_layers()
        except Exception as e:
            print(f"WARNING: Failed to parse activation data: {e}")
            import traceback
            traceback.print_exc()

        print("\n[6/8] Parsing metrics data...")
        metrics_data = {}
        metrics_xlsx_path = discovered.get('xlsx', None)
        if metrics_xlsx_path and os.path.exists(metrics_xlsx_path):
            metrics_parser = MetricsParser(metrics_xlsx_path)
            metrics_data = metrics_parser.get_metrics()
        else:
            print("WARNING: No analyze.xlsx file found")

        print("\n[7/8] Loading configuration and performance data...")
        config_data = load_config_data(model_dir_path)
        proctime_data = load_proctime_data(model_dir_path)
        cycles_data = load_cycles_data(model_dir_path)
        memory_data = load_memory_data(model_dir_path)

        print("\n[8/8] Combining and saving data...")

        performance_data = {}
        all_subgraphs = set(proctime_data.keys()) | set(cycles_data.keys()) | set(memory_data.keys())

        for subgraph_num in all_subgraphs:
            proctime_lookup = {layer['layer_num']: layer for layer in proctime_data.get(subgraph_num, [])}
            cycles_lookup = {layer['layer_num']: layer for layer in cycles_data.get(subgraph_num, [])}
            memory_lookup = {layer['layer_num']: layer for layer in memory_data.get(subgraph_num, [])}

            all_layer_nums = set(proctime_lookup.keys()) | set(cycles_lookup.keys()) | set(memory_lookup.keys())

            merged_layers = []
            for layer_num in sorted(all_layer_nums):
                layer_entry = {'layer_num': layer_num}

                if layer_num in proctime_lookup:
                    layer_entry['layer_type'] = proctime_lookup[layer_num]['layer_type']
                    layer_entry['proctime_us'] = proctime_lookup[layer_num]['proctime']

                if layer_num in cycles_lookup:
                    if 'layer_type' not in layer_entry:
                        layer_entry['layer_type'] = cycles_lookup[layer_num]['layer_type']
                    layer_entry['kernel_cycles'] = cycles_lookup[layer_num]['kernelOnlyCycles']
                    layer_entry['core_loop_cycles'] = cycles_lookup[layer_num]['coreLoopCycles']

                if layer_num in memory_lookup:
                    if 'layer_type' not in layer_entry:
                        layer_entry['layer_type'] = memory_lookup[layer_num]['layer_type']
                    layer_entry['memory'] = {
                        'l2_kb': memory_lookup[layer_num]['l2_usage'],
                        'msmc_kb': memory_lookup[layer_num]['msmc_usage'],
                        'ddr_kb': memory_lookup[layer_num]['ddr_usage'],
                        'total_kb': memory_lookup[layer_num]['total_usage']
                    }

                merged_layers.append(layer_entry)

            performance_data[subgraph_num] = merged_layers

        enhanced_tidl_data = {}

        onnx_nodes_by_subgraph = {}
        for subgraph_info in subgraph_data.get('subgraphs', []):
            sg_id = subgraph_info.get('id')
            onnx_nodes_by_subgraph[sg_id] = subgraph_info.get('nodes', [])

        for subgraph_id, tidl_info in tidl_data.items():
            enhanced_layers = []

            subgraph_metrics = metrics_data.get(str(subgraph_id), [])
            metrics_lookup = {m['tidl_layer_id']: m for m in subgraph_metrics if m.get('tidl_layer_id')}

            subgraph_perf = performance_data.get(subgraph_id, [])
            perf_lookup = {p['layer_num']: p for p in subgraph_perf}

            for layer in tidl_info.get('layers', []):
                layer_idx = layer['layer_index']
                tidl_layer_id = str(layer_idx)

                enhanced_layer = {
                    'index': layer_idx,
                    'type': layer['layer_type'],
                    'name': layer['layer_name'],
                    'parameters': layer['parameters']
                }

                if 'macs' in layer:
                    enhanced_layer['macs'] = layer['macs']
                if 'gmacs' in layer:
                    enhanced_layer['gmacs'] = layer['gmacs']

                if 'onnx_node_index' in layer:
                    enhanced_layer['onnx_node_index'] = layer['onnx_node_index']

                if layer_idx in perf_lookup:
                    enhanced_layer['performance'] = perf_lookup[layer_idx]

                if tidl_layer_id in metrics_lookup:
                    metric = metrics_lookup[tidl_layer_id]
                    enhanced_layer['metrics'] = {
                        'mae': metric['mean_abs_diff'],
                        'mean_abs_rel_diff': metric['mean_abs_rel_diff'],
                        'median_abs_diff': metric['median_abs_diff'],
                        'max_abs_diff': metric['max_abs_diff']
                    }

                activation_key = f"{subgraph_id}_{tidl_layer_id}"
                if activation_key in activation_data:
                    enhanced_layer['activation'] = activation_data[activation_key]

                enhanced_layers.append(enhanced_layer)

            enhanced_tidl_data[subgraph_id] = {
                'id': subgraph_id,
                'onnx_nodes': onnx_nodes_by_subgraph.get(subgraph_id, []),
                'layers': enhanced_layers,
                'total_gmacs': tidl_info.get('total_gmacs', 0.0),
                'graph': {
                    'nodes': tidl_info.get('graph_nodes', []),
                    'edges': tidl_info.get('graph_edges', [])
                }
            }

        metadata = {
            'target_device': config_data.get('target_device', 'Unknown'),
            'task_type': config_data.get('task_type', 'Unknown'),
            'tensor_bits': config_data.get('tensor_bits', 'Unknown'),
            'model_accuracy': config_data.get('accuracy', 'N/A')
        }

        performance_summary = {
            'num_frames': config_data.get('num_frames', 'N/A'),
            'total_gmacs': config_data.get('perfsim_gmacs', 'N/A'),
            'total_time_ms': config_data.get('perfsim_time_ms', 'N/A'),
            'ddr_transfer_mb': config_data.get('perfsim_ddr_transfer_mb', 'N/A'),
            'num_subgraphs': config_data.get('num_subgraphs', 'N/A')
        }

        combined_data = {
            'metadata': metadata,
            'model': {
                'details': model_data.get('model_details', {}),
                'layers': model_data.get('layer_details', {}),
                'tree_structure': model_data.get('tree_structure', {}),
                'graph': {
                    'edges': model_data.get('edges', [])
                }
            },
            'compilation': {
                'node_support': subgraph_data.get('node_support', {}),
                'tidl_subgraphs': enhanced_tidl_data
            },
            'performance': performance_summary
        }

        print(f"Writing compressed JSON to: {output_json_path}")
        json_str = json.dumps(combined_data)

        if not output_json_path.endswith('.gz'):
            output_json_path = output_json_path + '.gz'

        with gzip.open(output_json_path, 'wt', encoding='utf-8') as f:
            f.write(json_str)

        file_size = os.path.getsize(output_json_path) / (1024 * 1024)
        print(f"JSON data saved (compressed): {output_json_path} ({file_size:.2f} MB)")

        print("\n" + "=" * 70)
        print("SUCCESS! Data extraction complete.")
        print("=" * 70)
        print(f"\nExtracted data summary:")
        print(f"  - ONNX layers: {len(combined_data['model']['layers'])}")
        print(f"  - TIDL subgraphs: {len(combined_data['compilation']['tidl_subgraphs'])}")

        total_activation_layers = 0
        for subgraph_id, subgraph in combined_data['compilation']['tidl_subgraphs'].items():
            total_activation_layers += sum(1 for layer in subgraph['layers'] if 'activation' in layer)
        print(f"  - Layers with activation data: {total_activation_layers}")

        total_metrics_layers = 0
        for subgraph_id, subgraph in combined_data['compilation']['tidl_subgraphs'].items():
            total_metrics_layers += sum(1 for layer in subgraph['layers'] if 'metrics' in layer)
        print(f"  - Layers with metrics: {total_metrics_layers}")

        total_perf_layers = 0
        for subgraph_id, subgraph in combined_data['compilation']['tidl_subgraphs'].items():
            total_perf_layers += sum(1 for layer in subgraph['layers'] if 'performance' in layer)
        print(f"  - Layers with performance data: {total_perf_layers}")

        print(f"\nMetadata:")
        print(f"  - Device: {combined_data['metadata']['target_device']}")
        print(f"  - Precision: {combined_data['metadata']['tensor_bits']}")
        print(f"  - Task: {combined_data['metadata']['task_type']}")

        print(f"\nNext step:")
        print(f"  python html_generator.py {output_json_path} template.html output.html")

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main(work_dirs_path = sys.argv[1], output_json_path = sys.argv[2])
